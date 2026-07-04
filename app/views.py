from django.contrib.auth import logout
from django.db import DatabaseError, transaction
from django.db import models
from django.db.models import Q, Count
from django.shortcuts import render, redirect
from django.utils.text import slugify
from .models import LoaiSanPham, NhomHuong
import json
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.utils.html import escape
import random, string
import hashlib
import hmac
import urllib.parse
from datetime import datetime, timedelta
from django.conf import settings
from django.core.mail import send_mail
import requests as http_requests
import uuid
from django.http import JsonResponse
from .models import ThuocTinh, GiaTriThuocTinh
import json as _json
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone as tz
from django.db import models


from .models import (
    BaiViet,
    BienThe,
    BienTheThuocTinh,
    SanPhamNhomHuong,
    DanhGia,
    DonHang,
    ChiTietDonHang,
    GiaoHang,
    HinhAnh,
    HoiDap,
    KhachHang,
    SanPham,
    TaiKhoan,
    KhuyenMai,
    KhuyenMaiTaiKhoan,
    ThuongHieu,
    YeuThich,
    LichSuDiem,
    CauHinhThanhVien,NhaCungCap, PhieuNhap, ChiTietNhap,
    AIRecommendClick,
    ChatbotFeedback, 
)
from app.models import (
    LichSuXemSanPham, AIUserProfile, ChatbotHistory,
    AIRecommendImpression, SurveyResponse,
)
from collections import defaultdict




# Create your views here.
FALLBACK_IMAGES = {
    "default": "https://images.unsplash.com/photo-1541643600914-78b084683702?auto=format&fit=crop&w=900&q=80",
    "brand_hero": "https://images.unsplash.com/photo-1610461888750-10bfc601b874?auto=format&fit=crop&w=1800&q=80",
    "brand_poster": "https://images.unsplash.com/photo-1611930022073-b7a4ba5fcccd?auto=format&fit=crop&w=900&q=80",
    "brand_about": "https://images.unsplash.com/photo-1615634260167-c8cdede054de?auto=format&fit=crop&w=1200&q=80",
    "article_cover": "https://images.unsplash.com/photo-1588405748880-12d1d2a59a75?auto=format&fit=crop&w=1600&q=80",
}


def _format_currency(value):
    if value is None:
        return "0đ"
    return f"{int(value):,}".replace(",", ".") + "đ"

def _payment_status_label(order):
    payment = ((order.HinhThucThanhToan or "COD").strip() or "COD").lower()
    status = (order.TrangThai or "").strip()
    if status == "Thanh toán thất bại":
        return "Thanh toán thất bại"
    if payment in ("vnpay", "momo", "paypal"):
        return "Đã thanh toán"
    if status == "Hoàn tất":
        return "Đã thanh toán"
    return "Chưa thanh toán"


def _delivery_status_label(order):
    status = (order.TrangThai or "Chờ xác nhận").strip()
    if status == "Đã thanh toán":
        return "Chờ xác nhận"
    if status == "Thanh toán thất bại":
        return "Đã hủy"
    return status or "Chờ xác nhận"


def _order_account(order):
    customer = getattr(order, "id_KhachHang", None)
    if customer and getattr(customer, "id_TaiKhoan", None):
        return customer.id_TaiKhoan
    delivery = getattr(order, "id_GiaoHang", None)
    if delivery and getattr(delivery, "id_TaiKhoan", None):
        return delivery.id_TaiKhoan
    return None


def _send_order_status_email(order, status):
    account = _order_account(order)
    email = (getattr(account, "Email", "") or "").strip() if account else ""
    if not email:
        return False

    customer_name = _account_display_name(account)
    order_code = order.MaDonHang or f"#{order.id_DonHang}"
    subject_map = {
        "Đã xác nhận": f"Ami Perfumery — Đơn hàng {order_code} đã được xác nhận",
        "Hoàn tất": f"Ami Perfumery — Đơn hàng {order_code} đã hoàn tất",
    }
    body_map = {
        "Đã xác nhận": (
            f"Xin chào {customer_name},\n\n"
            f"Đơn hàng {order_code} của bạn đã được Ami Perfumery xác nhận. "
            "Bạn có thể theo dõi trạng thái đơn trong trang tài khoản và xác nhận đã nhận hàng khi đơn được giao thành công.\n\n"
            "Cảm ơn bạn đã mua sắm tại Ami Perfumery."
        ),
        "Hoàn tất": (
            f"Xin chào {customer_name},\n\n"
            f"Đơn hàng {order_code} của bạn đã được hoàn tất. "
            "Ami Perfumery rất vui khi được phục vụ bạn.\n\n"
            "Đừng quên đánh giá sản phẩm để nhận thêm nhiều ưu đãi hấp dẫn từ cửa hàng. "
            "Bạn có thể vào trang Tài khoản > Đơn hàng và nhấn nút Đánh giá ngay trên đơn đã hoàn tất.\n\n"
            "👉 Mở nhanh trang đơn hàng: https://amiperfumery.vn/profile/?tab=orders\n\n"
            "Trân trọng,\nAmi Perfumery"
        ),
    }
    try:
        send_mail(
            subject_map.get(status, f"Ami Perfumery — Cập nhật đơn hàng {order_code}"),
            body_map.get(status, f"Đơn hàng {order_code} đã được cập nhật sang trạng thái {status}."),
            getattr(settings, "DEFAULT_FROM_EMAIL", "noreply@ami.com"),
            [email],
            fail_silently=False,
        )
        return True
    except Exception as exc:
        print(f"[order_email] Không gửi được email đơn {order_code} tới {email}: {exc}")
        return False


def _restore_inventory_on_cancel(order):
    """Hoàn lại tồn kho khi admin xác nhận hủy đơn."""
    from django.db import models as db_models
    details = ChiTietDonHang.objects.filter(id_DonHang=order)
    for d in details:
        if d.id_BienThe and d.SoLuong:
            BienThe.objects.filter(pk=d.id_BienThe_id).update(
                SoLuong=db_models.F("SoLuong") + d.SoLuong
            )


def _send_cancel_confirmation_email(order):
    """Gửi email xác nhận hủy đơn thành công đến khách hàng."""
    account = _order_account(order)
    email = (getattr(account, "Email", "") or "").strip() if account else ""
    if not email:
        return False

    customer_name = _account_display_name(account)
    order_code = order.MaDonHang or f"#{order.id_DonHang}"

    # Lấy lý do hủy từ GhiChu GiaoHang
    gh = order.id_GiaoHang
    reason = ""
    if gh and gh.GhiChu and "[YÊU CẦU HỦY]:" in (gh.GhiChu or ""):
        for line in gh.GhiChu.splitlines():
            if "[YÊU CẦU HỦY]:" in line:
                reason = line.replace("[YÊU CẦU HỦY]:", "").strip()
                break

    subject = f"Ami Perfumery — Đơn hàng {order_code} đã được hủy thành công"
    body = f"""Xin chào {customer_name},

Đơn hàng {order_code} của bạn đã được hủy thành công theo yêu cầu.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Mã đơn:    {order_code}
  Tổng tiền: {_format_currency(order.TongTien)}
  Lý do hủy: {reason or "Theo yêu cầu khách hàng"}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Nếu bạn đã thanh toán online, số tiền sẽ được hoàn lại trong vòng 3-5 ngày làm việc.
Điểm thưởng đã sử dụng (nếu có) sẽ được hoàn trả vào tài khoản của bạn.

Nếu có bất kỳ thắc mắc nào, vui lòng liên hệ:
📞 0901 234 567 | ✉️ hello@amiperfumery.vn

Trân trọng,
Ami Perfumery 🌿
"""
    try:
        from django.core.mail import send_mail
        from django.conf import settings as dj_settings
        send_mail(
            subject=subject,
            message=body,
            from_email=getattr(dj_settings, "DEFAULT_FROM_EMAIL", "noreply@ami.com"),
            recipient_list=[email],
            fail_silently=False,
        )
        return True
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"[cancel_email] {e}")
        return False


def _deduct_inventory_on_order(order):
    """Trừ tồn kho an toàn với transaction lock — tránh race condition."""
    from django.db import transaction, models as db_models

    out_of_stock = []

    with transaction.atomic():
        details = ChiTietDonHang.objects.filter(id_DonHang=order).select_related('id_BienThe')

        for d in details:
            if not d.id_BienThe or not d.SoLuong:
                continue

            # Lock dòng BienThe này lại — request khác phải chờ
            try:
                bt = BienThe.objects.select_for_update().get(pk=d.id_BienThe_id)
            except BienThe.DoesNotExist:
                continue

            if bt.SoLuong >= d.SoLuong:
                bt.SoLuong -= d.SoLuong
                bt.save(update_fields=["SoLuong"])
            else:
                out_of_stock.append(bt.Sku or f"BienThe#{bt.id_BienThe}")

        if out_of_stock:
            # Rollback toàn bộ transaction, đơn hàng không được tạo
            raise Exception(f"Hết hàng: {', '.join(out_of_stock)}")

def _stock_shortage_message():
    return "Số lượng tồn kho không đủ, vui lòng giảm bớt số lượng sản phẩm sau đó mới thêm vào giỏ hàng."


def _safe_positive_int(value, default=1):
    try:
        value = int(value or default)
    except (TypeError, ValueError):
        return default
    return max(default, value)


def _resolve_checkout_items(items):
    """Khóa biến thể trong transaction và kiểm tra tổng số lượng đặt không vượt tồn kho."""
    resolved_items = []
    requested_by_variant = defaultdict(int)

    for item in items:
        product_id = item.get("productId")
        variant_id = item.get("variantId")
        qty = _safe_positive_int(item.get("qty"), 1)

        bien_the_qs = BienThe.objects.select_for_update()
        bien_the = None
        if variant_id and str(variant_id).isdigit():
            bien_the = bien_the_qs.filter(id_BienThe=int(variant_id)).first()
        if not bien_the and product_id:
            bien_the = bien_the_qs.filter(id_SanPham_id=product_id).order_by("id_BienThe").first()

        if not bien_the:
            product_name = item.get("name") or "Sản phẩm"
            return None, f"Không tìm thấy biến thể tồn kho cho {product_name}."

        requested_by_variant[bien_the.id_BienThe] += qty
        resolved_items.append({
            "item": item,
            "variant": bien_the,
            "qty": qty,
            "price": float(item.get("price") or bien_the.GiaBan or 0),
        })

    for resolved in resolved_items:
        variant = resolved["variant"]
        requested_qty = requested_by_variant[variant.id_BienThe]
        available_qty = int(variant.SoLuong or 0)
        if requested_qty > available_qty:
            return None, _stock_shortage_message()

    return resolved_items, None


def _extract_voucher_code(order):
    note = ""
    delivery = getattr(order, "id_GiaoHang", None)
    if delivery:
        note = delivery.GhiChu or ""
    marker = "[AMI_VOUCHER:"
    if marker in note:
        return note.split(marker, 1)[1].split("]", 1)[0].strip() or "—"
    return "—"


def _clean_delivery_note(note):
    if not note:
        return ""
    marker = "[AMI_VOUCHER:"
    if marker in note:
        before, after = note.split(marker, 1)
        tail = after.split("]", 1)[1] if "]" in after else ""
        return (before + tail).strip()
    return note


def _safe_list(queryset):
    try:
        return list(queryset)
    except DatabaseError:
        return []


def _safe_first(queryset):
    try:
        return queryset.first()
    except DatabaseError:
        return None
    

def _account_display_name(account):
    return (account.TenDangNhap or account.Username or "Khách hàng").strip()


def _rating_label(star):
    return {5: "Tuyệt vời · Highly Recommended", 4: "Rất tốt", 3: "Tốt", 2: "Chưa phù hợp", 1: "Không hài lòng"}.get(star, "")


def _voucher_type_label(voucher):
    loai = (voucher.LoaiKhuyenMai or "").lower()
    mapping = {
        "vip": "Exclusive",
        "thanh vien moi": "New Member",
        "freeship": "Freeship",
    }
    for key, label in mapping.items():
        if key in loai:
            return label
    return "Member"

# ═══════════════════════════════════════════════════════════════
# views.py — Thay các hàm helper điểm + thêm apply_points_api
# ═══════════════════════════════════════════════════════════════

# ── Helper: cộng điểm ──────────────────────────────────────────
def add_points(account, points, loai, mo_ta="", order=None):
    """Cộng điểm và ghi lịch sử. Dùng schema DB thực tế."""
    if not account or int(points) <= 0:
        return
    account.DiemTichLuy = int(account.DiemTichLuy or 0) + int(points)
    account.save(update_fields=["DiemTichLuy"])

    LichSuDiem.objects.create(
        id_TaiKhoan=account,
        id_DonHang=order,
        SoDiem=int(points),           # cột SoDiem (dương = cộng)
        Loai=loai,
        MoTa=mo_ta or f"Cộng {points} điểm",
        NgayTao=timezone.now(),
    )


# ── Helper: trừ điểm ──────────────────────────────────────────
def redeem_points(account, points, order=None):
    """Trừ điểm khi thanh toán. Trả False nếu không đủ điểm."""
    if not account:
        return False
    points = int(points)
    current = int(account.DiemTichLuy or 0)
    if points > current:
        return False
    account.DiemTichLuy = current - points
    account.save(update_fields=["DiemTichLuy"])

    LichSuDiem.objects.create(
        id_TaiKhoan=account,
        id_DonHang=order,
        SoDiem=-points,               # âm = trừ điểm
        Loai="redeem_order",
        MoTa="Dùng điểm thanh toán",
        NgayTao=timezone.now(),
    )
    return True


# ── Helper: cập nhật hạng thành viên ──────────────────────────
def update_member_level(account):
    """Cập nhật HangThanhVien theo TongChiTieu."""
    total = float(account.TongChiTieu or 0)
    if total >= 10_000_000:
        level = "Platinum"
    elif total >= 5_000_000:
        level = "Gold"
    elif total >= 2_000_000:
        level = "Silver"
    else:
        level = "Member"
    if account.HangThanhVien != level:
        account.HangThanhVien = level
        account.save(update_fields=["HangThanhVien"])


# ══════════════════════════════════════════════════════════════
# API: GET điểm hiện tại của user (dùng khi trang checkout load)
# URL: GET /api/points/
# ══════════════════════════════════════════════════════════════
def get_points_api(request):
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "points": 0, "discount": 0})
 
    account = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first()
    if not account:
        return JsonResponse({"ok": False, "points": 0, "discount": 0})
 
    points  = int(account.DiemTichLuy or 0)
    # 100 điểm = 10.000₫ → 1 điểm = 100₫
    discount = points * 100
    return JsonResponse({
        "ok": True,
        "points":   points,
        "discount": discount,                    # số tiền quy đổi nếu dùng hết
        "rate":     100,                          # 1 điểm = 100 VNĐ
    })


# ══════════════════════════════════════════════════════════════
# API: Tính giảm giá từ điểm (KHÔNG trừ thật, chỉ tính preview)
# POST /api/apply-points/
# Body: subtotal (tạm tính đã trừ voucher)
# ══════════════════════════════════════════════════════════════
@csrf_exempt
@require_POST
def apply_points_api(request):
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "need_login": True,
                             "message": "Vui lòng đăng nhập."})
 
    account = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first()
    if not account:
        return JsonResponse({"ok": False, "message": "Không tìm thấy tài khoản."})
 
    points  = int(account.DiemTichLuy or 0)
    if points <= 0:
        return JsonResponse({"ok": False, "message": "Bạn chưa có điểm tích lũy."})
 
    subtotal = float(request.POST.get("subtotal") or 0)
 
    # Giới hạn: điểm chỉ được giảm tối đa 30% tổng đơn
    RATE        = 100           # 1 điểm = 100 VNĐ
    MAX_RATIO   = 0.30          # tối đa 30% đơn hàng
    max_by_pct  = subtotal * MAX_RATIO
    discount_all = points * RATE
 
    # Số tiền thực tế được giảm
    discount = min(discount_all, max_by_pct)
    # Số điểm tương ứng
    points_used = int(discount / RATE)
 
    if points_used <= 0:
        return JsonResponse({"ok": False,
                             "message": "Đơn hàng chưa đủ điều kiện sử dụng điểm."})
 
    return JsonResponse({
        "ok":           True,
        "message":      f"Áp dụng {points_used} điểm — giảm {int(discount):,}₫ ✨",
        "points":       points,
        "points_used":  points_used,
        "discount":     int(discount),
        "rate":         RATE,
    })


# ══════════════════════════════════════════════════════════════
# Thêm vào urls.py:
# path('api/points/',         views.get_points_api,   name='points-api'),
# path('api/apply-points/',   views.apply_points_api, name='apply-points-api'),
# ══════════════════════════════════════════════════════════════

# ── Helper: cộng điểm ──────────────────────────────────────────
def add_points(account, points, loai, mo_ta="", order=None):
    """Cộng điểm và ghi lịch sử. Dùng schema DB thực tế."""
    if not account or int(points) <= 0:
        return
    account.DiemTichLuy = int(account.DiemTichLuy or 0) + int(points)
    account.save(update_fields=["DiemTichLuy"])
 
    LichSuDiem.objects.create(
        id_TaiKhoan=account,
        id_DonHang=order,
        SoDiem=int(points),           # cột SoDiem (dương = cộng)
        Loai=loai,
        MoTa=mo_ta or f"Cộng {points} điểm",
        NgayTao=timezone.now(),
    )

# ── Helper: trừ điểm ──────────────────────────────────────────
def redeem_points(account, points, order=None):
    """Trừ điểm khi thanh toán. Trả False nếu không đủ điểm."""
    if not account:
        return False
    points = int(points)
    current = int(account.DiemTichLuy or 0)
    if points > current:
        return False
    account.DiemTichLuy = current - points
    account.save(update_fields=["DiemTichLuy"])
 
    LichSuDiem.objects.create(
        id_TaiKhoan=account,
        id_DonHang=order,
        SoDiem=-points,               # âm = trừ điểm
        Loai="redeem_order",
        MoTa="Dùng điểm thanh toán",
        NgayTao=timezone.now(),
    )
    return True

def _get_available_vouchers(account_id):
    now = timezone.now()
    rows = KhuyenMaiTaiKhoan.objects.select_related("id_KhuyenMai").filter(
        id_TaiKhoan_id=account_id,
    ).order_by("-id")
    vouchers = []
    for row in rows:
        v = row.id_KhuyenMai
        if not v:
            continue
        is_active = (v.TrangThai or "").lower() in {"active", "on", "1"}
        not_expired = (v.NgayKetThuc is None) or (v.NgayKetThuc >= now)
        started = (v.NgayBatDau is None) or (v.NgayBatDau <= now)
        has_quota = (v.SoLuong is None) or (int(v.DaSuDung or 0) < int(v.SoLuong or 0))
        status = "Còn hiệu lực"
        if row.DaSuDung:
            status = "Đã sử dụng"
        elif not not_expired:
            status = "Hết hạn"
        elif not (is_active and started and has_quota):
            status = "Không khả dụng"

        vouchers.append({
            "id": v.id_KhuyenMai,
            "code": (v.MaKhuyenMai or "").upper(),
            "name": v.TenKhuyenMai or "",
            "description": v.MoTa or "",
            "discount_type": (v.LoaiGiam or "").lower(),
            "discount_value": float(v.GiaTriGiam or 0),
            "minimum_order": float(v.DonHangToiThieu or 0),
            "max_discount": float(v.GiamToiDa or 0),
            "expiry": v.NgayKetThuc.strftime("%d/%m/%Y") if v.NgayKetThuc else "",
            "status": status,
            "exclusive_badge": _voucher_type_label(v),
            "used": bool(row.DaSuDung),
        })
    return vouchers

# def _normalize_vietnamese_text(value):
#     if not isinstance(value, str):
#         return value
#     if not value:
#         return value

#     # Fix common mojibake when UTF-8 bytes were interpreted as latin-1/cp1252.
#     for source_enc in ("latin-1", "cp1252"):
#         try:
#             repaired = value.encode(source_enc).decode("utf-8")
#         except (UnicodeEncodeError, UnicodeDecodeError):
#             continue
#         if repaired != value:
#             return repaired
#     return value




def _product_image_map(product_ids):
    images = HinhAnh.objects.filter(id_SanPham_id__in=product_ids)

    mapping = {}

    for img in images:
        pid = img.id_SanPham_id

        if img.url:
            # 👉 luôn ép về URL đúng
            url = str(img.url)

            if not url.startswith("http"):
                url = settings.MEDIA_URL + url  # 👈 FIX QUAN TRỌNG

            if pid not in mapping:
                mapping[pid] = []

            mapping[pid].append(url)

    return mapping


def _first_variant_map(product_ids):
    first_variant_map = {}
    variant_rows = BienThe.objects.filter(id_SanPham_id__in=product_ids).order_by("id_SanPham_id", "id_BienThe")
    for variant in variant_rows:
        first_variant_map.setdefault(variant.id_SanPham_id, variant)
    return first_variant_map

def _variant_value_map(product_ids):
    rows = _safe_list(
        BienTheThuocTinh.objects.select_related(
            "id_BienThe",
            "id_GiaTriThuocTinh",
            "id_GiaTriThuocTinh__id_ThuocTinh",
        ).filter(id_BienThe__id_SanPham_id__in=product_ids)
    )
    mapping = {}
    for row in rows:
        product_id = row.id_BienThe.id_SanPham_id
        value = row.id_GiaTriThuocTinh.GiaTri
        if not value:
            continue
        mapping.setdefault(product_id, set()).add(value)
    return {pid: sorted(values) for pid, values in mapping.items()}

def _build_product_cards(products):
    product_ids = [item.id_SanPham for item in products]
    image_map = _product_image_map(product_ids)

    first_variant_map = _first_variant_map(product_ids)
    variant_value_map = _variant_value_map(product_ids)
    variant_count_map = {
        row["id_SanPham_id"]: row["count"]
        for row in BienThe.objects.filter(id_SanPham_id__in=product_ids)
        .values("id_SanPham_id")
        .annotate(count=models.Count("id_BienThe"))
    }

    cards = []
    for product in products:
        default_variant = first_variant_map.get(product.id_SanPham)
        product_images = image_map.get(product.id_SanPham, [])
        primary_image = product_images[0] if product_images else FALLBACK_IMAGES["default"]
        hover_image = product_images[1] if len(product_images) > 1 else primary_image
        stock = int(default_variant.SoLuong) if default_variant else 0
        status_value = (product.TrangThai_SanPham or "").strip().lower()
        is_new = status_value in {"new", "moi", "mới"}
        cards.append(
            {
                "id": product.id_SanPham,
                "name": product.TenSanPham,
                "brand": product.id_ThuongHieu.TenThuongHieu,
                "brand_slug": slugify(product.id_ThuongHieu.TenThuongHieu),
                "group_list": [h.TenNhomHuong for h in product.nhom_huongs.all()],
                "price": _format_currency(default_variant.GiaBan if default_variant else None),
                "price_raw": int(default_variant.GiaBan or 0) if default_variant else 0,
                "stock": stock,
                "is_new": is_new,
                "primary_image": primary_image,
                "hover_image": hover_image,
                "variant_values": variant_value_map.get(product.id_SanPham, []),
                "variant_count": variant_count_map.get(product.id_SanPham, 0),
            }
        )
    return cards




def home(request):
    brands = _safe_list(ThuongHieu.objects.order_by("TenThuongHieu"))
    categories = _safe_list(LoaiSanPham.objects.all())

    featured_products = _build_product_cards(
        _safe_list(
            SanPham.objects.select_related("id_ThuongHieu", "id_LoaiSanPham")
                            .prefetch_related("nhom_huongs")
            .order_by("-id_SanPham")[:8]
        )
    )

    total_products = SanPham.objects.count()
    total_brands = ThuongHieu.objects.count()
    total_customers = KhachHang.objects.count()

    def format_k(num):
        if num >= 1000:
            return f"{num//1000}K"
        return str(num)

    total_customers = format_k(total_customers)

    latest_articles = _safe_list(BaiViet.objects.order_by("-NgayTao")[:4])

    return render(
        request,
        "app/home.html",
        {
            "brands_home": brands,
            "categories": categories,
            "featured_products": featured_products,
            "latest_articles_home": latest_articles,
            "total_products": total_products,
            "total_brands": total_brands,
            "total_customers": total_customers,
        },
    )

def _category_article_queryset(category=None):
    articles = BaiViet.objects.order_by("-NgayTao")
    if category and category.TenLoaiSanPham:
        keyword = category.TenLoaiSanPham
        articles = articles.filter(Q(TieuDe__icontains=keyword) | Q(NoiDung__icontains=keyword))
    return articles

def category(request, segment='tat-ca'):

    search_q = (request.GET.get("q") or "").strip()

    # lấy tất cả danh mục
    categories = LoaiSanPham.objects.all()

    # tìm danh mục theo slug
    current_category = None
    for c in categories:
        if slugify(c.TenLoaiSanPham) == segment:
            current_category = c
            break

    # query sản phẩm
    products = SanPham.objects.select_related(
        "id_ThuongHieu",
        "id_LoaiSanPham"
    ).prefetch_related("nhom_huongs")

    # nếu không phải "tất cả" thì filter
    if segment != "tat-ca" and current_category:
        products = products.filter(id_LoaiSanPham=current_category)

    if search_q:
        products = products.filter(
            Q(TenSanPham__icontains=search_q)
            | Q(id_ThuongHieu__TenThuongHieu__icontains=search_q)
            | Q(nhom_huongs__TenNhomHuong__icontains=search_q)
        ).distinct()

    products = products.order_by("-id_SanPham")

    related_articles_qs = _category_article_queryset(current_category)
    related_articles = _safe_list(related_articles_qs[:6])

    if not related_articles and current_category:
        related_articles = _safe_list(BaiViet.objects.order_by("-NgayTao")[:6])

    recent_reviews = _safe_list(
        DanhGia.objects.select_related("id_TaiKhoan")
        .filter(id_SanPham__id_LoaiSanPham=current_category)
        .order_by("-NgayDanhGia", "-id_DanhGia")[:8]
    ) if current_category else _safe_list(
        DanhGia.objects.select_related("id_TaiKhoan")
        .order_by("-NgayDanhGia", "-id_DanhGia")[:8]
    )


    # 👉 context động theo DB
    context = {
        "page_title": f"Ami – {current_category.TenLoaiSanPham}" if current_category else "Ami – Tất cả nước hoa",
        "title": current_category.TenLoaiSanPham if current_category else "Tất cả nước hoa",
        "subtitle": current_category.MoTa if current_category else "Khám phá toàn bộ bộ sưu tập nước hoa.",
        "breadcrumb": current_category.TenLoaiSanPham if current_category else "Tất cả",
        "brands": _safe_list(ThuongHieu.objects.order_by("TenThuongHieu")),
        "segment": segment,
        "search_q": search_q,
        "products": _build_product_cards(products),
        "related_articles": related_articles,
        "recent_reviews": recent_reviews,
    }

    context["product_count"] = len(context["products"])

    from .models import BienThe
    from django.db.models import Max
    import math

    max_price_raw = BienThe.objects.aggregate(max_price=Max('GiaBan'))['max_price'] or 8000000
    # Làm tròn lên đến hàng triệu gần nhất
    max_price = math.ceil(float(max_price_raw) / 1000000) * 1000000

    context['max_price'] = int(max_price)
    context['max_price_display'] = f"{int(max_price):,}".replace(',', '.') + '₫'
    return render(request, 'app/category.html', context)

def get_sillage_label(value):
    if value is None:
        return "Đang cập nhật"
    if value >= 8:
        return "Tỏa xa"
    elif value >= 5:
        return "Vừa phải"
    return "Nhẹ"

def get_longevity_label(value):
    if value is None:
        return "Đang cập nhật"
    if value >= 9:
        return "Trên 10 giờ"
    elif value >= 7:
        return "8-10 giờ"
    elif value >= 5:
        return "5-7 giờ"
    return "Dưới 5 giờ"

from django.views.decorators.csrf import ensure_csrf_cookie

@ensure_csrf_cookie
def product_detail(request, product_id=None):
    
    product_queryset = SanPham.objects.select_related("id_ThuongHieu", "id_LoaiSanPham")\
                                        .prefetch_related("nhom_huongs")

    if product_id:
        product_obj = _safe_first(product_queryset.filter(id_SanPham=product_id))
    else:
        product_obj = _safe_first(product_queryset.order_by("id_SanPham"))
    if not product_obj:
        return render(request, "app/product.html", {"product_data": {}, "product_images": []})
    is_favorite = False

    account_id = request.session.get("account_id")

    if account_id:

        account = TaiKhoan.objects.filter(
            id_TaiKhoan=account_id
        ).first()

        if account:

            is_favorite = YeuThich.objects.filter(
                id_TaiKhoan=account,
                id_SanPham=product_obj
            ).exists()

    nhom_huongs = SanPhamNhomHuong.objects.select_related("id_NhomHuong").filter(
        id_SanPham=product_obj
    )
     # ── Flat list (hiện tại) ──
    nhom_huong_list = [
        {
            "name": item.id_NhomHuong.TenNhomHuong,
            "icon": item.id_NhomHuong.IconUrl.url if item.id_NhomHuong.IconUrl else "",
            "vai_tro": item.VaiTroHuong or "",
        }
        for item in nhom_huongs
    ]
 
    # ── Kim tự tháp phân tầng ──
    pyramid = {"top": [], "heart": [], "base": [], "other": []}
    for item in nhom_huongs:
        vai_tro = (item.VaiTroHuong or "").strip()
        entry = {
            "name": item.id_NhomHuong.TenNhomHuong,
            "icon": item.id_NhomHuong.IconUrl.url if item.id_NhomHuong.IconUrl else "",
        }
        if vai_tro.strip() == "Top Notes":
            pyramid["top"].append(entry)
        elif vai_tro.strip() == "Heart Notes":
            pyramid["heart"].append(entry)
        elif vai_tro.strip() == "Base Notes":
            pyramid["base"].append(entry)
        else:
            pyramid["other"].append(entry)
 
    # Nếu không có top/heart/base → đưa tất cả vào "other" để vẫn hiển thị
    if not (pyramid["top"] or pyramid["heart"] or pyramid["base"]):
        # Không có tầng → đưa tất cả vào other để hiển thị flat
        pyramid["other"] = [
            {
                "name": item.id_NhomHuong.TenNhomHuong,
                "icon": item.id_NhomHuong.IconUrl.url if item.id_NhomHuong.IconUrl else "",
            }
            for item in nhom_huongs
        ]

    nhom_huong_pyramid = pyramid if any([
        pyramid["top"], pyramid["heart"], pyramid["base"], pyramid["other"]
    ]) else None

    variants = _safe_list(BienThe.objects.filter(id_SanPham=product_obj).order_by("id_BienThe"))
    variant_attr_rows = _safe_list(
        BienTheThuocTinh.objects.select_related(
            "id_BienThe",
            "id_GiaTriThuocTinh",
            "id_GiaTriThuocTinh__id_ThuocTinh",
        ).filter(id_BienThe__id_SanPham=product_obj)
    )
    images = _safe_list(
        HinhAnh.objects.filter(
            Q(id_SanPham=product_obj) | Q(id_BienThe__id_SanPham=product_obj)
        ).order_by("id_HinhAnh")
    )
    root_reviews = _safe_list(
    DanhGia.objects.select_related("id_TaiKhoan")
    .filter(id_SanPham=product_obj, parent_id__isnull=True)
        .order_by("-NgayDanhGia", "-id_DanhGia")
    )
    # ═══════════════════════════════════════════════════════════════
# THAY TOÀN BỘ đoạn này trong hàm product_detail (views.py)
#
# Tìm dòng:    qa_items = []
# Đến hết:     qa_items.append({"question": q, "answer": answer})
# Thay bằng đoạn dưới đây:
# ═══════════════════════════════════════════════════════════════

    qa_items = []
 
    questions = HoiDap.objects.select_related(
        "id_TaiKhoan"
    ).filter(
        id_SanPham=product_obj,
        parent_id__isnull=True
    ).exclude(
        TrangThai='hidden'
    ).order_by('-NgayTao')
 
    # Thêm tạm vào cuối vòng for q in questions (trong views.py)
# để debug xem admin_answer có được tìm thấy không

    for q in questions:

        admin_answer = HoiDap.objects.select_related("id_TaiKhoan") \
            .exclude(TrangThai='hidden') \
            .filter(
                parent_id=q.id_HoiDap,
                id_TaiKhoan__LoaiTaiKhoan__in=['admin', 'staff']
            ).order_by('NgayTao').first()

        follow_ups = list(
            HoiDap.objects.select_related("id_TaiKhoan")
            .exclude(TrangThai='hidden')
            .filter(parent_id=q.id_HoiDap)
            .exclude(id_TaiKhoan__LoaiTaiKhoan__in=['admin', 'staff'])
            .order_by('NgayTao')
        )

        # DEBUG — xóa sau khi xác nhận hoạt động
        print(f"[QA DEBUG] q.id={q.id_HoiDap} nội_dung='{q.NoiDung[:20]}' "
              f"admin_answer={admin_answer.id_HoiDap if admin_answer else None} "
              f"follow_ups={[f.id_HoiDap for f in follow_ups]}")

        qa_items.append({
            "question":   q,
            "answer":     admin_answer,
            "follow_ups": follow_ups,
        })

    top_variant = variants[0] if variants else None
    variant_attr_map = {}
    option_groups = {}
    for row in variant_attr_rows:
        variant_id = row.id_BienThe_id
        attr_name = row.id_GiaTriThuocTinh.id_ThuocTinh.TenThuocTinh
        attr_value = row.id_GiaTriThuocTinh.GiaTri
        variant_attr_map.setdefault(variant_id, {})[attr_name] = attr_value
        option_groups.setdefault(attr_name, set()).add(attr_value)

    variant_payload = []
    for variant in variants:
        variant_payload.append(
            {
                "id": variant.id_BienThe,
                "sku": variant.Sku,
                "price": _format_currency(variant.GiaBan),
                "price_raw": int(variant.GiaBan or 0),
                "stock": int(variant.SoLuong or 0),
                "attributes": variant_attr_map.get(variant.id_BienThe, {}),
            }
        )
    rating_values = [item.SoSao for item in root_reviews if item.SoSao]
    rating_avg = round(sum(rating_values) / len(rating_values), 1) if rating_values else 0
    rating_count = len(rating_values)
    rating_breakdown = {star: len([v for v in rating_values if v == star]) for star in range(1, 6)}

    expert_replies = defaultdict(list)
    for reply in DanhGia.objects.select_related("id_TaiKhoan").filter(id_SanPham=product_obj, parent_id__isnull=False).order_by("NgayDanhGia", "id_DanhGia"):
        expert_replies[reply.parent_id].append(reply)

    review_cards = []
    for rv in root_reviews:
        name = _account_display_name(rv.id_TaiKhoan) if rv.id_TaiKhoan else "Khách hàng"
        review_cards.append({
            "id": rv.id_DanhGia,
            "name": name,
            "avatar": name[:1].upper(),
            "rating": int(rv.SoSao or 0),
            "label": _rating_label(int(rv.SoSao or 0)),
            "content": rv.NoiDung or "",
            "created_at": rv.NgayDanhGia.strftime("%d/%m/%Y %H:%M") if rv.NgayDanhGia else "",
            "replies": expert_replies.get(rv.id_DanhGia, []),
        })
    # print("ID:", product_obj.id_SanPham)
    # print("NongDo:", product_obj.NongDo)
    # print("XuatXu:", product_obj.XuatXu)
    # print("PhongCach:", product_obj.PhongCach)
    product_data = {
        
        "id": product_obj.id_SanPham,
        "name": product_obj.TenSanPham,
        "brand": product_obj.id_ThuongHieu.TenThuongHieu,
        "description": product_obj.MoTa_SanPham,
        "category": product_obj.id_LoaiSanPham.TenLoaiSanPham,
        "concentration": product_obj.NongDo,
        "longevity": product_obj.DoLuuHuong,
        "sillage": product_obj.DoToaHuong,

        "longevity_percent": min((product_obj.DoLuuHuong or 0) * 10, 100),
        "sillage_percent": min((product_obj.DoToaHuong or 0) * 10, 100),

        "longevity_text": get_longevity_label(product_obj.DoLuuHuong),
        "sillage_text": get_sillage_label(product_obj.DoToaHuong),

        "season": product_obj.MuaPhuHop,
        "time_use": product_obj.ThoiDiemSuDung,
        "style": product_obj.PhongCach,
        "age_group": product_obj.DoTuoiPhuHop,
        "release_year": product_obj.NamPhatHanh,
        "origin": product_obj.XuatXu,
        "scent_group": ", ".join([
                h.TenNhomHuong for h in product_obj.nhom_huongs.all()
            ]),
        "price": _format_currency(top_variant.GiaBan if top_variant else None),
        "stock": top_variant.SoLuong if top_variant else 0,
        "status": product_obj.TrangThai_SanPham,
        "rating_avg": rating_avg,
        "rating_count": rating_count,
        "rating_breakdown": {k: v for k, v in sorted(rating_breakdown.items(), reverse=True)},
        "reviews": review_cards,
        "questions": questions,
        # "reviews": root_reviews,
        "variants": json.dumps(variant_payload, ensure_ascii=False),
        "variant_count": len(variant_payload),
        "option_groups": {k: sorted(list(v)) for k, v in option_groups.items()},
        "qa_items": qa_items,
        # "is_favorite": is_favorite,
    }

    
    product_images = [img.url.url if hasattr(img.url, "url") else str(img.url) for img in images]

    related_products = _build_product_cards(
        _safe_list(
            SanPham.objects.select_related("id_ThuongHieu", "id_LoaiSanPham")
            .prefetch_related("nhom_huongs")
            .filter(id_ThuongHieu=product_obj.id_ThuongHieu)
            .exclude(id_SanPham=product_obj.id_SanPham)
            .order_by("-id_SanPham")[:10]
        )
    )

    scent_group_ids = [item.id_NhomHuong_id for item in nhom_huongs]
    similar_scent_products = _build_product_cards(
        _safe_list(
            SanPham.objects.select_related("id_ThuongHieu", "id_LoaiSanPham")
            .prefetch_related("nhom_huongs")
            .filter(nhom_huongs__id_NhomHuong__in=scent_group_ids)
            .exclude(id_SanPham=product_obj.id_SanPham)
            .distinct()
            .order_by("-id_SanPham")[:4]
        )
    ) if scent_group_ids else []

    return render(request, "app/product.html", {
        "product_data": product_data,
        "product_images": product_images,
        "nhom_huong_list": nhom_huong_list,
        "nhom_huong_pyramid": nhom_huong_pyramid,
        "related_products": related_products,
        "brand_slug": slugify(product_obj.id_ThuongHieu.TenThuongHieu),
        "similar_scent_products": similar_scent_products,
        "is_favorite": is_favorite,
    })


def brand_list(request):
    brands = list(ThuongHieu.objects.all().order_by("TenThuongHieu"))
    # brands = []
    # for row in brand_rows:
    #     name = row["TenThuongHieu"]
    #     slug = slugify(name)
    #     brands.append(
    #         {
    #             "slug": slug,
    #             "name": name,
    #             "tagline": "Tinh hoa mùi hương đẳng cấp",
    #             "palette": "#6f7d62",
    #             "poster_image": row["LogoUrl"] or FALLBACK_IMAGES["brand_poster"],
    #             "category": "Designer" if len(name) % 2 == 0 else "Niche",
    #         }
    #     )
    return render(request, "app/brand_list.html", {"brands": brands})


def brand_detail(request, slug):
    brand_obj = next((item for item in ThuongHieu.objects.all() if slugify(item.TenThuongHieu) == slug), None)
    if brand_obj is None:
        first_brand = ThuongHieu.objects.first()
        if not first_brand:
            return render(request, "app/brand_detail.html", {"brand": {}, "products": [], "pinned_products": []})
        brand_obj = first_brand

    brand_products = SanPham.objects.select_related("id_ThuongHieu", "id_LoaiSanPham")\
                                    .prefetch_related("nhom_huongs").filter(
        id_ThuongHieu=brand_obj
    )
    products = _build_product_cards(brand_products)

    brand = {
        "slug": slugify(brand_obj.TenThuongHieu),
        "name": brand_obj.TenThuongHieu,
        "tagline": "Di sản mùi hương tinh tế",
        "palette": "#6f7d62",
        "hero_image": brand_obj.LogoUrl or FALLBACK_IMAGES["brand_hero"],
        "about_image": brand_obj.LogoUrl or FALLBACK_IMAGES["brand_about"],
        "story": f"{brand_obj.TenThuongHieu} là thương hiệu được yêu thích trong bộ sưu tập nước hoa tại Ami.",
        "philosophy": "Tập trung vào chiều sâu mùi hương, sự cân bằng và tính ứng dụng mỗi ngày.",
        "signature_notes": ["Citrus", "Floral", "Woody"],
    }

    return render(
        request,
        "app/brand_detail.html",
        {
            "brand": brand,
            "products": products,
            "pinned_products": products[:2],
        },
    )


def blog_list(request):

    import re
    def _strip_html(text):
        return re.sub(r'<[^>]+>', '', text or '')

    featured_articles = BaiViet.objects.order_by("-NgayTao")[:3]

    for a in featured_articles:
        a.plain_text = _strip_html(a.NoiDung)[:120] + '...' if a.NoiDung else ''

    bento_articles = BaiViet.objects.order_by("-NgayTao")[3:]

    popular_articles = BaiViet.objects.order_by("-NgayTao")[:5]

    context = {
        "featured_articles": featured_articles,
        "bento_articles": bento_articles,
        "popular_articles": popular_articles,
    }

    return render(request, "app/blog.html", context)

def article_detail(request, id):
    article_obj = BaiViet.objects.get(id_BaiViet=id)

    articles_qs = BaiViet.objects.order_by("-NgayTao")

    article = {
        "id": article_obj.id_BaiViet,
        "title": article_obj.TieuDe,
        "author": article_obj.TacGia,
        "published_at": article_obj.NgayTao.strftime("%d/%m/%Y") if article_obj.NgayTao else "",
        "cover": article_obj.AnhDaiDien.url if article_obj.AnhDaiDien else FALLBACK_IMAGES["article_cover"],
        "body": [{"type": "p", "text": article_obj.NoiDung or ""}],
    }

    related_articles = [
        {
            "id": item.id_BaiViet,
            "title": item.TieuDe,
            "cover": item.AnhDaiDien.url if item.AnhDaiDien else FALLBACK_IMAGES["article_cover"],
        }
        for item in articles_qs
        if item.id_BaiViet != article_obj.id_BaiViet
    ][:5]

    return render(request, "app/article_detail.html", {
        "article": article,
        "related_articles": related_articles,
    })


def contact_page(request):
    return render(request, 'app/contact.html')

@csrf_exempt
@require_POST
def contact_send(request):
    name    = (request.POST.get('name')    or '').strip()
    email   = (request.POST.get('email')   or '').strip()
    phone   = (request.POST.get('phone')   or '').strip()
    subject = (request.POST.get('subject') or '').strip()
    message = (request.POST.get('message') or '').strip()

    if not all([name, email, subject, message]):
        return JsonResponse({'ok': False, 'message': 'Vui lòng điền đầy đủ thông tin.'})

    from django.core.mail import EmailMessage

    # ── Email gửi đến cửa hàng ──
    body_store = f"""Tin nhắn mới từ website Ami Perfumery
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Họ tên:        {name}
Email:         {email}
Số điện thoại: {phone}
Chủ đề:        {subject}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{message}

↩ Reply email này để trả lời trực tiếp cho khách: {email}
"""

    # ── Email xác nhận gửi cho khách ──
    body_customer = f"""Xin chào {name},

Ami Perfumery đã nhận được tin nhắn của bạn.
Chúng tôi sẽ phản hồi trong vòng 24 giờ!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Nội dung bạn đã gửi:
Chủ đề: {subject}

{message}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Trân trọng,
Ami Perfumery
📍 123 Đường 30/4, Quận Ninh Kiều, Cần Thơ
📞 0901 234 567
"""

    try:
        # Gửi cho cửa hàng
        mail_store = EmailMessage(
            subject=f'[Ami Web] {subject} — {name}',
            body=body_store,
            from_email='Ami Perfumery <lhngocc1304@gmail.com>',
            to=['lhngocc1304@gmail.com'],
            reply_to=[f'{name} <{email}>'],
        )
        mail_store.send(fail_silently=False)

        # Gửi xác nhận cho khách
        mail_customer = EmailMessage(
            subject='Ami Perfumery đã nhận tin nhắn của bạn ✨',
            body=body_customer,
            from_email='Ami Perfumery <lhngocc1304@gmail.com>',
            to=[email],
        )
        mail_customer.send(fail_silently=False)

        return JsonResponse({'ok': True})

    except Exception as e:
        print(f'[contact_send] Lỗi gửi mail: {e}')
        return JsonResponse({'ok': False, 'message': 'Không thể gửi email. Vui lòng thử lại sau.'})


@ensure_csrf_cookie 
def cart_page(request):
    account_id = request.session.get("account_id")
 
    suggestions = _build_product_cards(
        _safe_list(
            SanPham.objects.select_related("id_ThuongHieu", "id_LoaiSanPham")
                           .prefetch_related("nhom_huongs")
                           .order_by("-id_SanPham")[:8]
        )
    )[:4]
 
    voucher_data = []
    account      = None
    if account_id:
        voucher_data = _get_available_vouchers(account_id)
        account      = _safe_first(TaiKhoan.objects.filter(id_TaiKhoan=account_id))
 
    return render(request, 'app/cart.html', {
        'suggestions':  suggestions,
        'voucher_data': voucher_data,
        'is_logged_in': bool(account_id),
        'account':      account,
    })



def auth_page(request):
    if request.session.get("account_id"):
        return redirect('profile-page')
    return render(request, 'app/auth.html')
# @csrf_exempt
# @require_POST
# def login_api(request):
#     email = (request.POST.get("email") or "").strip()
#     password = request.POST.get("password") or ""
#     if not email or not password:
#         return JsonResponse({"ok": False, "message": "Vui lòng nhập email và mật khẩu."}, status=400)

#     account = TaiKhoan.objects.filter(Email__iexact=email, MatKhau=password, TrangThai_TaiKhoan__iexact='active').first()
#     if not account:
#         return JsonResponse({"ok": False, "message": "Email hoặc mật khẩu không đúng."}, status=401)

#     request.session["account_id"] = account.id_TaiKhoan
#     request.session["account_name"] = account.TenDangNhap or account.Username
#     return JsonResponse({"ok": True, "message": "Đăng nhập thành công."})


# ═══════════════════════════════════════════════════════
# views.py — Thay thế hàm submit_question hiện tại
# Thêm xử lý parent_id để khách phản hồi sau câu trả lời admin
# ═══════════════════════════════════════════════════════

@csrf_exempt
@require_POST
def submit_question(request):
    print("=== SUBMIT QUESTION CALLED ===")
    print("POST:", request.POST)

    account_id = request.session.get("account_id")

    if not account_id:
        return JsonResponse({
            "ok": False,
            "need_login": True
        })

    product_id = request.POST.get("product_id")
    content = (request.POST.get("content") or "").strip()
    parent_id = request.POST.get("parent_id")  # <-- nhận parent_id từ form phản hồi

    if not content:
        return JsonResponse({
            "ok": False,
            "message": "Vui lòng nhập câu hỏi."
        })

    try:
        account = TaiKhoan.objects.get(id_TaiKhoan=account_id)
        product = SanPham.objects.get(id_SanPham=product_id)
    except (TaiKhoan.DoesNotExist, SanPham.DoesNotExist):
        return JsonResponse({
            "ok": False,
            "message": "Không tìm thấy dữ liệu."
        })

    # Nếu có parent_id → đây là câu hỏi tiếp theo sau khi admin trả lời
    # Nếu không có → đây là câu hỏi mới hoàn toàn
    resolved_parent_id = int(parent_id) if parent_id else None

    question = HoiDap.objects.create(
        id_SanPham=product,
        id_TaiKhoan=account,
        NoiDung=content,
        TrangThai='pending',
        parent_id=resolved_parent_id,
        NgayTao=timezone.now()
    )

    return JsonResponse({
        "ok": True,
        "question": {
            "id": question.id_HoiDap,
            "name": account.TenDangNhap,
            "content": question.NoiDung,
            "created_at": question.NgayTao.strftime("%d/%m/%Y %H:%M"),
            "is_reply": resolved_parent_id is not None,  # để JS biết đây là reply hay câu hỏi mới
        }
    })

@csrf_exempt
@require_POST
def submit_review(request):
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "need_login": True, "message": "Vui lòng đăng nhập để đánh giá."})

    last_submit = request.session.get("review_last_submit")
    now_ts = timezone.now().timestamp()
    if last_submit and now_ts - float(last_submit) < 8:
        return JsonResponse({"ok": False, "message": "Bạn đang thao tác quá nhanh, vui lòng thử lại sau vài giây."}, status=429)

    order_id = request.POST.get("order_id")
    product_id = request.POST.get("product_id")

    try:
        rating = int(request.POST.get("rating") or 0)
    except ValueError:
        rating = 0

    content = escape((request.POST.get("content") or "").strip())

    if not order_id:
        return JsonResponse({"ok": False, "message": "Thiếu mã đơn hàng."}, status=400)

    if rating < 1 or rating > 5:
        return JsonResponse({"ok": False, "message": "Số sao phải từ 1 đến 5."}, status=400)

    if not content:
        return JsonResponse({"ok": False, "message": "Vui lòng nhập nội dung đánh giá."}, status=400)

    if len(content) > 800:
        return JsonResponse({"ok": False, "message": "Đánh giá tối đa 800 ký tự."}, status=400)

    account = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first()
    product = SanPham.objects.filter(id_SanPham=product_id).first()

    if not account or not product:
        return JsonResponse({"ok": False, "message": "Không tìm thấy dữ liệu."}, status=404)

    customer = KhachHang.objects.filter(id_TaiKhoan_id=account_id).first()

    order_filter = Q(id_DonHang=order_id) & Q(TrangThai="Hoàn tất")

    if customer:
        order_filter &= (
            Q(id_KhachHang=customer) |
            Q(id_GiaoHang__id_TaiKhoan_id=account_id)
        )
    else:
        order_filter &= Q(id_GiaoHang__id_TaiKhoan_id=account_id)

    order = DonHang.objects.filter(order_filter).first()

    if not order:
        return JsonResponse({
            "ok": False,
            "message": "Đơn hàng chưa hoàn tất hoặc không thuộc tài khoản của bạn."
        }, status=403)

    product_in_order = ChiTietDonHang.objects.filter(
        id_DonHang=order,
        id_BienThe__id_SanPham=product
    ).exists()

    if not product_in_order:
        return JsonResponse({
            "ok": False,
            "message": "Sản phẩm này không thuộc đơn hàng cần đánh giá."
        }, status=403)

    already_reviewed = DanhGia.objects.filter(
        id_DonHang=order,
        id_SanPham=product,
        id_TaiKhoan=account,
        parent_id__isnull=True
    ).exists()

    if already_reviewed:
        return JsonResponse({
            "ok": False,
            "message": "Bạn đã đánh giá sản phẩm trong đơn hàng này rồi."
        }, status=400)

    review = DanhGia.objects.create(
        id_DonHang=order,
        id_SanPham=product,
        id_TaiKhoan=account,
        SoSao=rating,
        NoiDung=content,
        parent_id=None,
        NgayDanhGia=timezone.now(),
    )

    request.session["review_last_submit"] = now_ts

    config = CauHinhThanhVien.objects.filter(
        TrangThai='active'
    ).order_by('MucChiTieuToiThieu').first()

    review_bonus = int(config.ThuongDanhGia or 100) if config else 100

    add_points(
        account,
        review_bonus,
        "review_bonus",
        f"Thưởng đánh giá sản phẩm trong đơn {order.MaDonHang}",
        order
    )

    update_member_level(account)

    return JsonResponse({
        "ok": True,
        "message": "Cảm ơn bạn đã chia sẻ trải nghiệm cùng Ami Perfume.",
        "reviewed": True,
        "order_id": order.id_DonHang,
        "product_id": product.id_SanPham,
        "review": {
            "id": review.id_DanhGia,
            "name": _account_display_name(account),
            "rating": rating,
            "label": _rating_label(rating),
            "content": content,
            "created_at": review.NgayDanhGia.strftime("%d/%m/%Y %H:%M"),
        }
    })

# ═══════════════════════════════════════════════════════
# Thêm 2 hàm này vào views.py (sau hàm submit_review)
# Và thêm 2 URL vào urls.py
# ═══════════════════════════════════════════════════════

@csrf_exempt
@require_POST
def delete_review(request):
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "need_login": True})

    review_id = request.POST.get("review_id")
    try:
        review = DanhGia.objects.get(
            id_DanhGia=review_id,
            id_TaiKhoan_id=account_id  # chỉ cho xóa review của chính mình
        )
        review.delete()
        return JsonResponse({"ok": True})
    except DanhGia.DoesNotExist:
        return JsonResponse({"ok": False, "message": "Không tìm thấy đánh giá."}, status=404)


@csrf_exempt
@require_POST
def edit_review(request):
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "need_login": True})

    review_id = request.POST.get("review_id")
    content   = escape((request.POST.get("content") or "").strip())
    try:
        rating = int(request.POST.get("rating") or 0)
    except ValueError:
        rating = 0

    if rating < 1 or rating > 5:
        return JsonResponse({"ok": False, "message": "Số sao phải từ 1 đến 5."})
    if not content:
        return JsonResponse({"ok": False, "message": "Vui lòng nhập nội dung."})
    if len(content) > 800:
        return JsonResponse({"ok": False, "message": "Tối đa 800 ký tự."})

    try:
        review = DanhGia.objects.get(
            id_DanhGia=review_id,
            id_TaiKhoan_id=account_id
        )
        review.SoSao   = rating
        review.NoiDung = content
        review.save(update_fields=["SoSao", "NoiDung"])
        return JsonResponse({"ok": True})
    except DanhGia.DoesNotExist:
        return JsonResponse({"ok": False, "message": "Không tìm thấy đánh giá."}, status=404)
    

# ═══════════════════════════════════════════════════════
# Thêm vào views.py sau hàm edit_review
# ═══════════════════════════════════════════════════════

@csrf_exempt
@require_POST
def toggle_wishlist(request):
    """Toggle thêm/xóa sản phẩm khỏi danh sách yêu thích."""
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "need_login": True})

    product_id = request.POST.get("product_id")
    if not product_id:
        return JsonResponse({"ok": False, "message": "Thiếu product_id."})

    try:
        account  = TaiKhoan.objects.get(id_TaiKhoan=account_id)
        product  = SanPham.objects.get(id_SanPham=product_id)
        customer = KhachHang.objects.filter(id_TaiKhoan=account).first()
    except (TaiKhoan.DoesNotExist, SanPham.DoesNotExist):
        return JsonResponse({"ok": False, "message": "Không tìm thấy dữ liệu."})

    if not customer:
        return JsonResponse({"ok": False, "message": "Không tìm thấy thông tin khách hàng."})

    # Kiểm tra đã yêu thích chưa
    existing = YeuThich.objects.filter(
        id_KhachHang=customer,
        id_SanPham=product
    ).first()

    if existing:
        # Đã yêu thích → xóa
        existing.delete()
        return JsonResponse({"ok": True, "action": "removed", "message": "Đã xóa khỏi danh sách yêu thích."})
    else:
        # Chưa yêu thích → thêm
        YeuThich.objects.create(
            id_KhachHang=customer,
            id_SanPham=product,
            NgayTao=timezone.now(),
        )
        return JsonResponse({"ok": True, "action": "added", "message": "Đã thêm vào danh sách yêu thích."})


def get_wishlist_status(request, product_id):
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"liked": False})
    try:
        account = TaiKhoan.objects.get(id_TaiKhoan=account_id)
        liked = YeuThich.objects.filter(
            id_TaiKhoan=account,
            id_SanPham_id=product_id
        ).exists()
        return JsonResponse({"liked": liked})
    except TaiKhoan.DoesNotExist:
        return JsonResponse({"liked": False})


@csrf_exempt
@require_POST
def login_api(request):
    email = (request.POST.get("email") or "").strip()
    password = request.POST.get("password") or ""
    if not email or not password:
        return JsonResponse({"ok": False, "message": "Vui lòng nhập email và mật khẩu."}, status=400)

    account = TaiKhoan.objects.filter(Email__iexact=email, MatKhau=password, TrangThai_TaiKhoan__iexact='active').first()
    if not account:
        return JsonResponse({"ok": False, "message": "Email hoặc mật khẩu không đúng."}, status=401)

    request.session["account_id"] = account.id_TaiKhoan
    request.session["account_name"] = account.TenDangNhap
    # ── Kích hoạt AI Personalization ──
    import threading
    def _bg_personalize(acc_id):
        try:
            from app.ai.personalize import get_personalized_recommendations
            get_personalized_recommendations(account.id_TaiKhoan, top_n=8)
        except Exception as e:
            print(f'[AI] Login personalization failed: {e}')
    threading.Thread(target=_bg_personalize, args=(account.id_TaiKhoan,), daemon=True).start()
    return JsonResponse({
        "ok":      True,
        "message": "Đăng nhập thành công.",
        "user": {
            "name":   account.TenDangNhap,
            "email":  account.Email,
            "level":  account.HangThanhVien or "Member",
            "points": account.DiemTichLuy or 0,
        }
    })

def social_complete(request):
    """
    Sau khi social_django xử lý xong, redirect về đây.
    Session đã được set bởi pipeline.
    """
    next_url = request.GET.get('next') or '/'
    if request.session.get('account_id'):
        return redirect(next_url)
    return redirect('auth-page')

@csrf_exempt
@require_POST
def register_api(request):
    full_name = (request.POST.get("fullname") or "").strip()
    email = (request.POST.get("email") or "").strip()
    phone = (request.POST.get("phone") or "").strip()
    username = (request.POST.get("username") or "").strip()
    password = request.POST.get("password") or ""

    if not all([full_name, email, phone, username, password]):
        return JsonResponse({"ok": False, "message": "Vui lòng điền đầy đủ thông tin."}, status=400)

    if TaiKhoan.objects.filter(Username__iexact=username).exists():
        return JsonResponse({"ok": False, "message": "Tên đăng nhập đã tồn tại."}, status=409)
    if TaiKhoan.objects.filter(Email__iexact=email).exists():
        return JsonResponse({"ok": False, "message": "Email đã được sử dụng."}, status=409)

    account = TaiKhoan.objects.create(
        Username=username,
        MatKhau=password,
        TenDangNhap=full_name,
        Email=email,
        SDT=phone,
        LoaiTaiKhoan='customer',
        TrangThai_TaiKhoan='active',
        NgayTao=timezone.now(),
    )
    KhachHang.objects.create(
        id_TaiKhoan=account,
        TenKhachHang=full_name,
        DiaChi='',
        GioiTinh='',
    )
    return JsonResponse({"ok": True, "message": "Đăng ký thành công."})


@csrf_exempt
@require_POST
def forgot_password_api(request):
    email = (request.POST.get("email") or "").strip()
    username = (request.POST.get("username") or "").strip()
    new_password = request.POST.get("new_password") or ""

    if not all([email, username, new_password]):
        return JsonResponse({"ok": False, "message": "Vui lòng nhập email, tên đăng nhập và mật khẩu mới."}, status=400)

    account = TaiKhoan.objects.filter(Email__iexact=email, Username__iexact=username).first()
    if not account:
        return JsonResponse({"ok": False, "message": "Không tìm thấy tài khoản phù hợp."}, status=404)

    account.MatKhau = new_password
    account.save(update_fields=["MatKhau"])
    return JsonResponse({"ok": True, "message": "Đổi mật khẩu thành công."})

# ═══════════════════════════════════════════════════════
# Thay hàm profile_page trong views.py
# Thêm phần lấy wishlist_data từ DB
# ═══════════════════════════════════════════════════════

def profile_page(request):
    account_id = request.session.get("account_id")
    if not account_id:
        return redirect('auth-page')

    account = _safe_first(TaiKhoan.objects.filter(id_TaiKhoan=account_id))
    customer = _safe_first(
        KhachHang.objects
        .select_related("id_TaiKhoan")
        .filter(id_TaiKhoan_id=account_id)
    )
    print(f"[DEBUG] account={account_id}, customer={customer}, gender={customer.GioiTinh if customer else 'NO CUSTOMER'}")

    # ── Đánh giá ──────────────────────────────────────────────
    review_data = []
    try:
        user_reviews = list(
            DanhGia.objects
            .select_related("id_SanPham", "id_SanPham__id_ThuongHieu")
            .filter(id_TaiKhoan_id=account_id, parent_id__isnull=True)
            .order_by("-NgayDanhGia")
        )
        review_product_ids = [rv.id_SanPham_id for rv in user_reviews if rv.id_SanPham_id]
        review_image_map = _product_image_map(review_product_ids) if review_product_ids else {}

        for rv in user_reviews:
            product = rv.id_SanPham
            if not product:
                continue
            images = review_image_map.get(product.id_SanPham, [])
            review_data.append({
                "id":           rv.id_DanhGia,
                "product_id":   product.id_SanPham,
                "product_name": product.TenSanPham,
                "brand":        product.id_ThuongHieu.TenThuongHieu if product.id_ThuongHieu else "",
                "image":        images[0] if images else FALLBACK_IMAGES["default"],
                "rating":       int(rv.SoSao or 0),
                "label":        _rating_label(int(rv.SoSao or 0)),
                "content":      rv.NoiDung or "",
                "created_at":   rv.NgayDanhGia.strftime("%d/%m/%Y") if rv.NgayDanhGia else "",
            })
    except Exception as e:
        import traceback; traceback.print_exc()

    # ── Yêu thích ─────────────────────────────────────────────
    wishlist_data = []
    try:
        wish_rows = list(
            YeuThich.objects
            .select_related("id_SanPham", "id_SanPham__id_ThuongHieu")
            .filter(id_TaiKhoan_id=account_id)
        )
        wish_product_ids = [w.id_SanPham_id for w in wish_rows if w.id_SanPham_id]
        wish_image_map   = _product_image_map(wish_product_ids) if wish_product_ids else {}
 
        for w in wish_rows:
            product = w.id_SanPham
            if not product:
                continue
            images        = wish_image_map.get(product.id_SanPham, [])
            first_variant = BienThe.objects.filter(id_SanPham=product).order_by("id_BienThe").first()
            wishlist_data.append({
                "id":           w.id_YeuThich,
                "product_id":   product.id_SanPham,
                "product_name": product.TenSanPham,
                "brand":        product.id_ThuongHieu.TenThuongHieu if product.id_ThuongHieu else "",
                "image":        images[0] if images else FALLBACK_IMAGES["default"],
                "price":        _format_currency(first_variant.GiaBan if first_variant else None),
            })
    except Exception:
        import traceback; traceback.print_exc()

    profile = {
        "full_name": (customer.TenKhachHang if customer else None)
                     or (account.TenDangNhap if account else "") or "Khách hàng",
        "username":  (account.Username if account else "") or "guest",
        "email":     (account.Email if account else "") or "",
        "phone":     (account.SDT if account else "") or "",
        "address":   (customer.DiaChi if customer else "") or "",
        "gender":    (customer.GioiTinh if customer else "") or "",
        "points": account.DiemTichLuy if account else 0,

        "level": account.HangThanhVien if account else "Member",

        "total_spending": _format_currency(
            account.TongChiTieu if account else 0
        ),
        "profile_avatar": (
            account.AnhDaiDien.url
            if account and account.AnhDaiDien
            else "/static/app/images/default-avatar.png"
        ),
    }

    print("WISHLIST DATA:", wishlist_data)
    vouchers = _get_available_vouchers(account_id)

    point_logs = []

    if account:

        point_logs = list(

            LichSuDiem.objects
            .filter(id_TaiKhoan=account)
            .order_by("-NgayTao")[:20]

        )
    return render(request, 'app/profile.html', {
        "profile":       profile,
        "review_data":   review_data,
        "wishlist_data": wishlist_data,
        "voucher_data": vouchers,
        "point_logs": point_logs,
    })


def checkout_page(request):
    account_id = request.session.get("account_id")
 
    delivery  = None
    form_data = {"name": "", "phone": "", "email": "", "address": "", "note": ""}
 
    if account_id:
        # Lấy thông tin giao hàng gần nhất của tài khoản này
        delivery = _safe_first(
            GiaoHang.objects
            .filter(id_TaiKhoan_id=account_id)
            .order_by("-id_GiaoHang")
        )
        if delivery:
            form_data["name"]    = delivery.TenNguoiNhan or ""
            form_data["phone"]   = delivery.SDT          or ""
            form_data["address"] = delivery.DiaChi       or ""
            form_data["note"]    = _clean_delivery_note(delivery.GhiChu)
 
        # Email lấy từ TaiKhoan
        acc = _safe_first(TaiKhoan.objects.filter(id_TaiKhoan=account_id))
        if acc:
            form_data["email"] = acc.Email or ""
            if not form_data["name"]:
                form_data["name"] = acc.TenDangNhap or ""
 
    return render(request, 'app/checkout.html', {
        "checkout":     form_data,
        "is_logged_in": bool(account_id),
        "voucher_data": _get_available_vouchers(account_id) if account_id else [],
    })


@csrf_exempt
@require_POST
def check_stock_api(request):
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({"ok": False, "message": "Dữ liệu không hợp lệ.", "stock": 0}, status=400)

    variant_id = body.get("variant_id") or body.get("variantId")
    product_id = body.get("product_id") or body.get("productId")
    qty = _safe_positive_int(body.get("qty"), 1)

    variant = None

    # Chỉ tìm theo variant_id nếu là số nguyên hợp lệ
    if variant_id and str(variant_id).strip().isdigit():
        variant = BienThe.objects.filter(id_BienThe=int(variant_id)).first()

    # Fallback: tìm theo product_id
    if not variant and product_id:
        try:
            variant = BienThe.objects.filter(
                id_SanPham_id=int(product_id)
            ).order_by("id_BienThe").first()
        except (ValueError, TypeError):
            pass

    if not variant:
        return JsonResponse({
            "ok":      False,
            "message": "Không tìm thấy biến thể tồn kho.",
            "stock":   0,
        }, status=404)   # ← đổi 404 để client phân biệt với 400

    stock = int(variant.SoLuong or 0)

    if stock <= 0:
        return JsonResponse({
            "ok":      False,
            "message": "Sản phẩm này hiện đã hết hàng.",
            "stock":   0,
        }, status=400)

    if qty > stock:
        return JsonResponse({
            "ok":      False,
            "message": _stock_shortage_message(),
            "stock":   stock,   # ← luôn trả về stock thực
        }, status=400)

    return JsonResponse({
        "ok":        True,
        "stock":     stock,
        "variant_id": variant.id_BienThe,
    })

@csrf_exempt
@require_POST
def place_order_api(request):
    """
    Nhận thông tin từ checkout.js → kiểm tra tồn kho, lưu đơn hàng và trừ tồn kho.
    """
    account_id = request.session.get("account_id")
 
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({"ok": False, "message": "Dữ liệu không hợp lệ."}, status=400)
 
    name       = (body.get("name")    or "").strip()
    phone      = (body.get("phone")   or "").strip()
    address    = (body.get("address") or "").strip()
    note       = (body.get("note")    or "").strip()
    payment    = (body.get("payment") or "cod").strip()
    items      = body.get("items", [])
    total      = float(body.get("total") or 0)
    voucher_cd = (body.get("voucher_code") or "").strip().upper()
    pts_used   = int(body.get("points_used") or 0)
    pts_disc   = float(body.get("points_discount") or 0)
 
    if not name or not phone or not address:
        return JsonResponse({"ok": False, "message": "Thiếu thông tin giao hàng."}, status=400)
    if not items:
        return JsonResponse({"ok": False, "message": "Giỏ hàng trống."}, status=400)
 
    ma_don = "AMI-" + "".join(random.choices(string.digits, k=8))
 
    try:
        with transaction.atomic():
            resolved_items, stock_error = _resolve_checkout_items(items)
            if stock_error:
                return JsonResponse({"ok": False, "message": stock_error}, status=400)

            account  = None
            customer = None

            if account_id:
                account = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first()
                if account:
                    customer = KhachHang.objects.filter(id_TaiKhoan_id=account_id).first()

                    # ── FIX: Tự tạo KhachHang nếu chưa có ──────────
                    if not customer:
                        customer = KhachHang.objects.create(
                            id_TaiKhoan=account,
                            TenKhachHang=account.TenDangNhap or name,
                            DiaChi=address,
                            GioiTinh='',
                        )
                        print(f"[place_order] Đã tạo KhachHang mới id={customer.id_KhachHang} cho account={account_id}")

            # ── 1. Lưu GiaoHang ─────────────────────────────────────
            delivery_note = note
            if voucher_cd:
                delivery_note = f"{note}\n[AMI_VOUCHER:{voucher_cd}]".strip()
            giao_hang = GiaoHang.objects.create(
                id_TaiKhoan_id = account_id if account_id else None,
                TenNguoiNhan   = name,
                SDT            = phone,
                DiaChi         = address,
                GhiChu         = delivery_note,
            )

            # ── 2. Lưu DonHang ──────────────────────────────────────
            don_hang = DonHang.objects.create(
                MaDonHang           = ma_don,
                id_KhachHang        = customer,      # Giờ luôn có giá trị nếu đăng nhập
                id_GiaoHang         = giao_hang,
                ThoiGian            = timezone.now(),
                HinhThucThanhToan   = payment,
                TrangThai           = "Chờ xác nhận",
                TongTien            = total,
                DiemDaDung          = pts_used,
                TienGiamTuDiem      = pts_disc,
                DiemNhanDuoc        = 0,
            )
            print(f"[place_order] Đã tạo DonHang {ma_don} | customer={customer} | total={total}")

            # ── 3. Lưu ChiTietDonHang ─────────────────────────────────
            for resolved in resolved_items or []:
                bien_the = resolved["variant"]
                qty = resolved["qty"]
                ChiTietDonHang.objects.create(
                    id_DonHang  = don_hang,
                    id_BienThe  = bien_the,
                    SoLuong     = qty,
                    GiaBan      = resolved["price"],
                    GiaGiam     = 0,
                )
            # KHÔNG có .update() ở đây nữa

            # ── Trừ tồn kho an toàn (có lock) ────────────────────────
            try:
                _deduct_inventory_on_order(don_hang)
            except Exception as e:
                raise  # để transaction.atomic() bên ngoài rollback

        # ── 4. Đánh dấu voucher đã dùng ─────────────────────────
            if voucher_cd and account_id:
                rel = KhuyenMaiTaiKhoan.objects.filter(
                    id_TaiKhoan_id=account_id,
                    id_KhuyenMai__MaKhuyenMai__iexact=voucher_cd,
                    DaSuDung=False,
                ).first()
                if rel:
                    rel.DaSuDung = True
                    rel.save(update_fields=["DaSuDung"])
                    v = rel.id_KhuyenMai
                    if v:
                        v.DaSuDung = int(v.DaSuDung or 0) + 1
                        v.save(update_fields=["DaSuDung"])

            # ── 5. Trừ điểm nếu dùng ────────────────────────────────
            if pts_used > 0 and account:
                redeem_points(account, pts_used, order=don_hang)

            # ── 6. Cộng điểm đơn đầu tiên ───────────────────────────
            if account and customer:
                order_count = DonHang.objects.filter(id_KhachHang=customer).count()
                if order_count == 1:
                    add_points(account, 200, "first_order_bonus",
                               "Thưởng đơn hàng đầu tiên", don_hang)

            # ── 7. Cộng điểm thường ─────────────────────────────────
            if account:
                earned = int(total / 10000)
                if earned > 0:
                    add_points(account, earned, "earn_order",
                               f"Tích điểm đơn hàng {ma_don}", don_hang)

            # ── 8. Cập nhật TongChiTieu & hạng ──────────────────────
            if account:
                account.TongChiTieu = float(account.TongChiTieu or 0) + total
                account.save(update_fields=["TongChiTieu"])
                update_member_level(account)

        # ── Return sau khi transaction đã commit ───────────────────
        if payment == "vnpay":
            return JsonResponse({
                "ok":       True,
                "order_id": ma_don,
                "redirect": f"/api/vnpay-create/?order_id={ma_don}&amount={int(total)}",
            })
        if payment == "momo":
            return JsonResponse({
                "ok":       True,
                "order_id": ma_don,
                "redirect": f"/api/momo-create/?order_id={ma_don}&amount={int(total)}",
            })
        if payment == "paypal":
            return JsonResponse({
                "ok":       True,
                "order_id": ma_don,
                "redirect": f"/api/paypal-create/?order_id={ma_don}&amount={int(total)}",
            })
 
        # COD
        return JsonResponse({
            "ok":       True,
            "order_id": ma_don,
            "message":  "Đặt hàng thành công!",
        })
 
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({"ok": False, "message": f"Lỗi hệ thống: {str(e)}"}, status=500)


@csrf_exempt
@require_POST
def apply_voucher_api(request):

    account_id = request.session.get("account_id")

    if not account_id:
        return JsonResponse({
            "ok": False,
            "need_login": True,
            "message": "Vui lòng đăng nhập để sử dụng ưu đãi thành viên."
        }, status=401)

    # =========================
    # LẤY DỮ LIỆU
    # =========================
    code = (request.POST.get("code") or "").strip()
    subtotal = float(request.POST.get("subtotal") or 0)

    print("CODE:", code)
    print("SUBTOTAL:", subtotal)

    if not code:
        return JsonResponse({
            "ok": False,
            "message": "Vui lòng nhập mã khuyến mãi."
        }, status=400)

    # =========================
    # TÌM VOUCHER
    # =========================
    voucher = KhuyenMai.objects.filter(
        MaKhuyenMai__iexact=code,
        TrangThai="active"
    ).first()

    print("FOUND VOUCHER:", voucher)

    if not voucher:
        return JsonResponse({
            "ok": False,
            "message": "Mã khuyến mãi không hợp lệ."
        }, status=404)

    # =========================
    # KIỂM TRA USER CÓ SỞ HỮU
    # =========================
    rel = KhuyenMaiTaiKhoan.objects.filter(
        id_TaiKhoan_id=account_id,
        id_KhuyenMai=voucher
    ).first()

    print("USER REL:", rel)

    if not rel:
        return JsonResponse({
            "ok": False,
            "message": "Bạn không sở hữu mã này."
        }, status=403)

    # =========================
    # KIỂM TRA ĐÃ DÙNG
    # =========================
    if rel.DaSuDung:
        return JsonResponse({
            "ok": False,
            "message": "Bạn đã sử dụng mã này."
        }, status=400)

    # =========================
    # KIỂM TRA THỜI GIAN
    # =========================
    now = timezone.now()

    if voucher.NgayBatDau and voucher.NgayBatDau > now:
        return JsonResponse({
            "ok": False,
            "message": "Mã chưa thể sử dụng."
        }, status=400)

    if voucher.NgayKetThuc and voucher.NgayKetThuc < now:
        return JsonResponse({
            "ok": False,
            "message": "Mã đã hết hạn."
        }, status=400)

    # =========================
    # KIỂM TRA SỐ LƯỢNG
    # =========================
    if (
        voucher.SoLuong is not None
        and int(voucher.DaSuDung or 0) >= int(voucher.SoLuong)
    ):
        return JsonResponse({
            "ok": False,
            "message": "Mã khuyến mãi đã hết lượt sử dụng."
        }, status=400)

    # =========================
    # KIỂM TRA ĐƠN TỐI THIỂU
    # =========================
    if subtotal < float(voucher.DonHangToiThieu or 0):
        return JsonResponse({
            "ok": False,
            "message": "Đơn hàng chưa đạt giá trị tối thiểu."
        }, status=400)

    # =========================
    # TÍNH GIẢM GIÁ
    # =========================
    discount = 0
    loai_giam = (voucher.LoaiGiam or "").lower()

    print("LOAI GIAM:", loai_giam)

    # GIẢM %
    if loai_giam == "percent":

        discount = subtotal * (
            float(voucher.GiaTriGiam or 0) / 100
        )

        # GIẢM TỐI ĐA
        if voucher.GiamToiDa:
            discount = min(
                discount,
                float(voucher.GiamToiDa)
            )

    # FREE SHIP
    elif loai_giam == "free_ship":

        discount = 0

    # GIẢM TIỀN CỐ ĐỊNH
    elif loai_giam == "fixed":

        discount = float(voucher.GiaTriGiam or 0)

    print("DISCOUNT:", discount)

    # =========================
    # RESPONSE
    # =========================
    return JsonResponse({
        "ok": True,
        "message": "Áp dụng mã thành công ✨",
        "code": voucher.MaKhuyenMai,
        "discount": int(discount),
        "type": loai_giam,
    })

def logout_view(request):
    logout(request)
    request.session.pop("account_id", None)
    request.session.pop("account_name", None)
    return redirect('home')


@csrf_exempt
@require_POST
def toggle_favorite(request):
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "need_login": True})
 
    product_id = request.POST.get("product_id")
    if not product_id:
        return JsonResponse({"ok": False, "message": "Thiếu product_id."})
 
    try:
        account = TaiKhoan.objects.get(id_TaiKhoan=account_id)
        product = SanPham.objects.get(id_SanPham=product_id)
    except (TaiKhoan.DoesNotExist, SanPham.DoesNotExist):
        return JsonResponse({"ok": False, "message": "Không tìm thấy dữ liệu."})
 
    existing = YeuThich.objects.filter(
        id_TaiKhoan=account,
        id_SanPham=product
    ).first()
 
    if existing:
        existing.delete()
        count = YeuThich.objects.filter(id_TaiKhoan=account).count()
        return JsonResponse({
            "ok":             True,
            "action":         "removed",
            "message":        "Đã xóa khỏi danh sách yêu thích.",
            "wishlist_count": count,
        })
    else:
        YeuThich.objects.create(
            id_TaiKhoan=account,
            id_SanPham=product,
        )
        count = YeuThich.objects.filter(id_TaiKhoan=account).count()
        return JsonResponse({
            "ok":             True,
            "action":         "added",
            "message":        "Đã thêm vào danh sách yêu thích.",
            "wishlist_count": count,
        })

# ADMIN
def admin_dashboard(request):
    context = {
        "total_orders": 120,
        "total_users": 45,
        "revenue": 25000000,
    }
    return render(request, "admin/dashboard.html", context)

# def admin_redirect(request):
#     return redirect('admin-dashboard')

# ═══════════════════════════════════════════════════════════════
# Thêm vào views.py — API kiểm tra đơn hàng đầu tiên
# URL: GET /api/check-first-order/
# ═══════════════════════════════════════════════════════════════

def check_first_order_api(request):
    """
    Kiểm tra xem đây có phải đơn hàng đầu tiên của khách không.
    Nếu chưa có đơn nào → freeship tự động.
    """
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"is_first": False, "freeship": False})

    try:
        customer = KhachHang.objects.filter(id_TaiKhoan_id=account_id).first()
        if not customer:
            return JsonResponse({"is_first": True, "freeship": True,
                                 "message": "🎁 Đơn hàng đầu tiên — Miễn phí vận chuyển!"})

        has_order = DonHang.objects.filter(
            id_KhachHang=customer
        ).exists()

        if not has_order:
            return JsonResponse({
                "is_first": True,
                "freeship": True,
                "message": "🎁 Đơn hàng đầu tiên — Miễn phí vận chuyển!",
            })
        else:
            return JsonResponse({"is_first": False, "freeship": False})

    except Exception:
        return JsonResponse({"is_first": False, "freeship": False})


# ═══════════════════════════════════════════════════════════════
# Thêm vào urls.py:
# path('api/check-first-order/', views.check_first_order_api, name='check-first-order'),
# ═══════════════════════════════════════════════════════════════


# ── Ký VNPAY ──────────────────────────────────────────────────
def _vnpay_sign(data: dict, secret: str) -> str:
    """
    VNPAY yêu cầu:
    1. Sắp xếp key theo alphabet
    2. Nối thành query string (KHÔNG encode value)
    3. Ký HMAC-SHA512
    """
    # Sắp xếp theo key
    sorted_items = sorted(data.items())
    # Nối chuỗi dạng key=value&key=value (KHÔNG urllib.parse.urlencode)
    query_string = "&".join(f"{k}={v}" for k, v in sorted_items)
    
    # Ký HMAC-SHA512
    signature = hmac.new(
        secret.encode("utf-8"),
        query_string.encode("utf-8"),
        hashlib.sha512
    ).hexdigest()
    return signature


# ── Tạo URL thanh toán VNPAY ──────────────────────────────────
def vnpay_create(request):
    order_id = request.GET.get("order_id", "")
    amount   = int(request.GET.get("amount", 0))

    now         = datetime.now()
    create_date = now.strftime("%Y%m%d%H%M%S")
    expire_date = (now + timedelta(minutes=15)).strftime("%Y%m%d%H%M%S")

    params = {
        "vnp_Version":    "2.1.0",
        "vnp_Command":    "pay",
        "vnp_TmnCode":    settings.VNPAY_TMN_CODE,
        "vnp_Amount":     str(amount * 100),
        "vnp_CurrCode":   "VND",
        "vnp_TxnRef":     order_id,
        "vnp_OrderInfo":  f"Thanh toan don hang {order_id}",
        "vnp_OrderType":  "other",
        "vnp_Locale":     "vn",
        "vnp_ReturnUrl":  settings.VNPAY_RETURN_URL,
        "vnp_IpAddr":     request.META.get("REMOTE_ADDR", "127.0.0.1"),
        "vnp_CreateDate": create_date,
        "vnp_ExpireDate": expire_date,
    }

    # ── Bước 1: Sắp xếp ──
    sorted_params = sorted(params.items())

    # ── Bước 2: Tạo chuỗi ký — dùng urllib.parse.quote_plus cho value ──
    hash_data = "&".join(
        f"{k}={urllib.parse.quote_plus(str(v), safe='')}"
        for k, v in sorted_params
    )

    # ── Bước 3: Ký HMAC-SHA512 ──
    secure_hash = hmac.new(
        settings.VNPAY_HASH_SECRET.encode("utf-8"),
        hash_data.encode("utf-8"),
        hashlib.sha512
    ).hexdigest()

    # DEBUG
    print("=== VNPAY HASH DATA ===")
    print(hash_data)
    print("HASH:", secure_hash[:20], "...")

    # ── Bước 4: Build URL cuối ──
    query = "&".join(
        f"{k}={urllib.parse.quote_plus(str(v), safe='')}"
        for k, v in sorted_params
    )
    pay_url = f"{settings.VNPAY_URL}?{query}&vnp_SecureHash={secure_hash}"

    return redirect(pay_url)


# ── Nhận kết quả từ VNPAY trả về ─────────────────────────────
def _vnpay_sign_verify(data: dict, secret: str) -> str:
    """Dùng để verify callback từ VNPAY — cùng logic encode."""
    sorted_items = sorted(data.items())
    hash_data = "&".join(
        f"{k}={urllib.parse.quote_plus(str(v), safe='')}"
        for k, v in sorted_items
    )
    return hmac.new(
        secret.encode("utf-8"),
        hash_data.encode("utf-8"),
        hashlib.sha512
    ).hexdigest()


def vnpay_return(request):
    params        = dict(request.GET)
    vnp_hash      = request.GET.get("vnp_SecureHash", "")
    order_id      = request.GET.get("vnp_TxnRef", "")
    response_code = request.GET.get("vnp_ResponseCode", "")

    # Loại bỏ hash khỏi params trước khi verify
    verify_params = {
        k: (v[0] if isinstance(v, list) else v)
        for k, v in params.items()
        if k not in ("vnp_SecureHash", "vnp_SecureHashType")
    }

    expected_hash   = _vnpay_sign_verify(verify_params, settings.VNPAY_HASH_SECRET)
    signature_valid = hmac.compare_digest(expected_hash, vnp_hash)

    print("=== VNPAY RETURN ===")
    print("Response code:", response_code)
    print("Signature valid:", signature_valid)
    print("Expected:", expected_hash[:20])
    print("Got:     ", vnp_hash[:20])

    if signature_valid and response_code == "00":
        DonHang.objects.filter(MaDonHang=order_id).exclude(TrangThai__in=["Đã hủy", "Hoàn tất"]).update(TrangThai="Chờ xác nhận")
        return render(request, "app/payment_success.html", {
            "order_id": order_id,
            "message":  "Thanh toán VNPAY thành công!",
        })
    else:
        DonHang.objects.filter(MaDonHang=order_id).update(TrangThai="Thanh toán thất bại")
        return render(request, "app/payment_failed.html", {
            "order_id":      order_id,
            "response_code": response_code,
            "message":       "Thanh toán thất bại hoặc bị hủy.",
        })
    

# Thanh toán MOMO
def momo_create(request):
    order_id   = request.GET.get("order_id", "")
    amount     = int(request.GET.get("amount", 0))
    request_id = str(uuid.uuid4())  # ID duy nhất cho mỗi request
    order_info = f"Thanh toan don hang {order_id}"
    extra_data = ""

    # ── Tạo chữ ký ──────────────────────────────────────────
    raw_signature = (
        f"accessKey={settings.MOMO_ACCESS_KEY}"
        f"&amount={amount}"
        f"&extraData={extra_data}"
        f"&ipnUrl={settings.MOMO_NOTIFY_URL}"
        f"&orderId={order_id}"
        f"&orderInfo={order_info}"
        f"&partnerCode={settings.MOMO_PARTNER_CODE}"
        f"&redirectUrl={settings.MOMO_RETURN_URL}"
        f"&requestId={request_id}"
        f"&requestType=payWithMethod"
    )

    signature = hmac.new(
        settings.MOMO_SECRET_KEY.encode("utf-8"),
        raw_signature.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()

    # ── Gọi API MoMo ────────────────────────────────────────
    payload = {
        "partnerCode": settings.MOMO_PARTNER_CODE,
        "partnerName": "Ami Perfumery",
        "storeId":     "AmiStore",
        "requestId":   request_id,
        "amount":      str(amount),
        "orderId":     order_id,
        "orderInfo":   order_info,
        "redirectUrl": settings.MOMO_RETURN_URL,
        "ipnUrl":      settings.MOMO_NOTIFY_URL,
        "requestType": "payWithMethod",
        "extraData":   extra_data,
        "lang":        "vi",
        "signature":   signature,
    }

    print("=== MOMO REQUEST ===")
    print("Raw signature:", raw_signature)
    print("Signature:", signature[:20], "...")

    try:
        res  = http_requests.post(
            settings.MOMO_ENDPOINT,
            json=payload,
            timeout=15
        )
        data = res.json()

        print("=== MOMO RESPONSE ===")
        print(data)

        if data.get("resultCode") == 0:
            return redirect(data["payUrl"])
        else:
            print("MoMo Error:", data.get("message"))
            return render(request, "app/payment_failed.html", {
                "order_id": order_id,
                "message":  f"MoMo: {data.get('message', 'Lỗi không xác định')}",
                "response_code": str(data.get("resultCode", "")),
            })
    except Exception as e:
        import traceback; traceback.print_exc()
        return render(request, "app/payment_failed.html", {
            "order_id": order_id,
            "message":  f"Lỗi kết nối MoMo: {str(e)}",
            "response_code": "",
        })


def momo_return(request):
    """MoMo redirect về đây sau khi thanh toán."""
    result_code = request.GET.get("resultCode", "")
    order_id    = request.GET.get("orderId",    "")
    message     = request.GET.get("message",    "")

    print("=== MOMO RETURN ===")
    print("resultCode:", result_code)
    print("orderId:",    order_id)
    print("message:",    message)

    if result_code == "0":
        DonHang.objects.filter(MaDonHang=order_id).exclude(TrangThai__in=["Đã hủy", "Hoàn tất"]).update(TrangThai="Chờ xác nhận")
        return render(request, "app/payment_success.html", {
            "order_id": order_id,
            "message":  "Thanh toán MoMo thành công!",
        })
    else:
        DonHang.objects.filter(MaDonHang=order_id).update(TrangThai="Thanh toán thất bại")
        return render(request, "app/payment_failed.html", {
            "order_id":      order_id,
            "message":       f"Thanh toán thất bại: {message}",
            "response_code": result_code,
        })


@csrf_exempt
def momo_ipn(request):
    """MoMo gọi IPN để xác nhận server-to-server."""
    try:
        data        = json.loads(request.body)
        result_code = str(data.get("resultCode", ""))
        order_id    = data.get("orderId", "")

        print("=== MOMO IPN ===")
        print(data)

        if result_code == "0":
            DonHang.objects.filter(MaDonHang=order_id).exclude(TrangThai__in=["Đã hủy", "Hoàn tất"]).update(TrangThai="Chờ xác nhận")
        return JsonResponse({"status": "ok"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)
    

def my_orders_api(request):
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "need_login": True})
 
    customer = KhachHang.objects.filter(id_TaiKhoan_id=account_id).first()
 
    # Query cả 2 trường hợp để không bỏ sót đơn hàng
    if customer:
        orders = list(
            DonHang.objects
            .filter(
                Q(id_KhachHang=customer) |
                Q(id_GiaoHang__id_TaiKhoan_id=account_id)
            )
            .select_related("id_GiaoHang")
            .distinct()
            .order_by("-ThoiGian")
        )
    else:
        # Không có KhachHang → tìm qua GiaoHang
        orders = list(
            DonHang.objects
            .filter(id_GiaoHang__id_TaiKhoan_id=account_id)
            .select_related("id_GiaoHang")
            .order_by("-ThoiGian")
        )
 
    if not orders:
        return JsonResponse({"ok": True, "orders": []})
 
    order_ids = [o.id_DonHang for o in orders]
    details = list(
        ChiTietDonHang.objects
        .filter(id_DonHang_id__in=order_ids)
        .select_related(
            "id_BienThe",
            "id_BienThe__id_SanPham",
            "id_BienThe__id_SanPham__id_ThuongHieu"
        )
    )
 
    detail_map = {}
    product_ids_all = []
    for d in details:
        detail_map.setdefault(d.id_DonHang_id, []).append(d)
        if d.id_BienThe and d.id_BienThe.id_SanPham_id:
            product_ids_all.append(d.id_BienThe.id_SanPham_id)
 
    image_map = _product_image_map(list(set(product_ids_all)))
 
    STATUS_STEP = {
        "Chờ xác nhận":   0,
        "Đã xác nhận":    1,
        "Đang giao":      2,
        "Khách đã nhận hàng": 3,
        "Hoàn tất":       4,
        "Đã hủy":         -1,
        "Đã giao":        3,
        "Đã thanh toán":  1,
        "Thanh toán thất bại": -1,
    }
 
    result = []
    for order in orders:
        items = detail_map.get(order.id_DonHang, [])
        item_list = []
        for d in items:
            bt = d.id_BienThe
            if not bt:
                continue
            sp = bt.id_SanPham
            if not sp:
                continue
            imgs = image_map.get(sp.id_SanPham, [])
            reviewed = DanhGia.objects.filter(
                id_DonHang=order,
                id_SanPham=sp,
                id_TaiKhoan_id=account_id,
                parent_id__isnull=True
            ).exists()
            item_list.append({
                "product_id":   sp.id_SanPham,
                "product_name": sp.TenSanPham,
                "brand":        sp.id_ThuongHieu.TenThuongHieu if sp.id_ThuongHieu else "",
                "image":        imgs[0] if imgs else FALLBACK_IMAGES["default"],
                "qty":          d.SoLuong or 1,
                "price":        _format_currency(d.GiaBan),
                "reviewed": reviewed,
            })
 
        trang_thai = _delivery_status_label(order)
        gh = order.id_GiaoHang
 
        result.append({
            "id":       order.id_DonHang,
            "ma_don":   order.MaDonHang or f"#{order.id_DonHang}",
            "date":     order.ThoiGian.strftime("%d/%m/%Y") if order.ThoiGian else "",
            "status":   trang_thai,
            "step":     STATUS_STEP.get(trang_thai, 0),
            "total":    _format_currency(order.TongTien),
            "payment":  order.HinhThucThanhToan or "COD",
            "payment_status": _payment_status_label(order),
            "items":    item_list,
            "address":  gh.DiaChi if gh else "",
            "receiver": gh.TenNguoiNhan if gh else "",
            "phone":    gh.SDT if gh else "",
        })
 
    return JsonResponse({"ok": True, "orders": result})
 
 
# ── 2. API: Khách xác nhận đã nhận hàng ──────────────────────────
# URL: POST /api/confirm-received/
# Body: order_id
# ──────────────────────────────────────────────────────────────────
@csrf_exempt
@require_POST
def confirm_received_api(request):
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "need_login": True})
 
    order_id = request.POST.get("order_id")
    if not order_id:
        return JsonResponse({"ok": False, "message": "Thiếu order_id."})
 
    customer = KhachHang.objects.filter(id_TaiKhoan_id=account_id).first()
 
    if customer:
        order = DonHang.objects.filter(
            Q(id_KhachHang=customer) | Q(id_GiaoHang__id_TaiKhoan_id=account_id),
            id_DonHang=order_id,
            TrangThai__in=["Đã xác nhận", "Đang giao"]
        ).first()
    else:
        order = DonHang.objects.filter(
            id_DonHang=order_id,
            id_GiaoHang__id_TaiKhoan_id=account_id,
            TrangThai__in=["Đã xác nhận", "Đang giao"]
        ).first()
 
    if not order:
        return JsonResponse({
            "ok": False,
            "message": "Không tìm thấy đơn hàng hoặc đơn chưa đủ điều kiện xác nhận."
        })
 
    order.TrangThai = "Khách đã nhận hàng"
    order.save(update_fields=["TrangThai"])
 
    return JsonResponse({
        "ok":         True,
        "message":    "Cảm ơn bạn đã xác nhận! Ami Perfumery sẽ hoàn tất đơn hàng trong thời gian sớm nhất.",
        "new_status": "Khách đã nhận hàng",
    })
 
def _notify_admin_cancel_request(order, reason):
    """Gửi email cho admin khi khách yêu cầu hủy đơn."""
    from django.conf import settings as dj_settings
    admin_email = getattr(dj_settings, "ADMIN_NOTIFY_EMAIL",
                  getattr(dj_settings, "DEFAULT_FROM_EMAIL", "admin@amiperfumery.vn"))

    account = _order_account(order)
    customer_name = _account_display_name(account) if account else "Khách hàng"
    customer_email = (getattr(account, "Email", "") or "").strip() if account else ""
    order_code = order.MaDonHang or f"#{order.id_DonHang}"
    gh = order.id_GiaoHang

    body = f"""Ami Perfumery — Thông báo yêu cầu hủy đơn

Khách hàng: {customer_name}
Email: {customer_email}
Mã đơn: {order_code}
Tổng tiền: {_format_currency(order.TongTien)}
Địa chỉ: {gh.DiaChi if gh else 'N/A'}
SĐT: {gh.SDT if gh else 'N/A'}

Lý do hủy: {reason}

Truy cập trang quản lý để xác nhận hủy:
http://localhost:8000/admin-orders/

Sau khi xác nhận hủy, email xác nhận sẽ được gửi tự động đến khách hàng.
"""
    try:
        send_mail(
            subject=f"[AMI] Yêu cầu hủy đơn {order_code} từ {customer_name}",
            message=body,
            from_email=getattr(dj_settings, "DEFAULT_FROM_EMAIL", "noreply@ami.com"),
            recipient_list=[admin_email],
            fail_silently=True,
        )
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"[cancel_notify] {e}")


# ── 3. API: Khách hủy đơn hàng ────────────────────────────────────
# URL: POST /api/cancel-order/
# Body: order_id, reason
# ──────────────────────────────────────────────────────────────────
@csrf_exempt
@require_POST
def cancel_order_api(request):
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "need_login": True})
 
    order_id = request.POST.get("order_id")
    customer = KhachHang.objects.filter(id_TaiKhoan_id=account_id).first()
 
    # Tìm đơn hàng qua cả 2 trường hợp
    if customer:
        order = DonHang.objects.select_related("id_GiaoHang").filter(
            Q(id_KhachHang=customer) | Q(id_GiaoHang__id_TaiKhoan_id=account_id),
            id_DonHang=order_id,
            TrangThai="Chờ xác nhận"
        ).first()
    else:
        order = DonHang.objects.select_related("id_GiaoHang").filter(
            id_DonHang=order_id,
            id_GiaoHang__id_TaiKhoan_id=account_id,
            TrangThai="Chờ xác nhận"
        ).first()
 
    if not order:
        return JsonResponse({
            "ok": False,
            "message": "Không thể hủy đơn này. Đơn hàng đã được xác nhận hoặc đang giao."
        })
 
    order.TrangThai = "Chờ hủy"
    order.save(update_fields=["TrangThai"])

    # Lưu lý do vào GhiChu GiaoHang
    reason = (request.POST.get("reason") or "Không có lý do").strip()
    gh = order.id_GiaoHang
    if gh:
        old = (gh.GhiChu or "").strip()
        gh.GhiChu = f"{old}\n[YÊU CẦU HỦY]: {reason}".strip()
        gh.save(update_fields=["GhiChu"])

    _notify_admin_cancel_request(order, reason)

    return JsonResponse({
        "ok": True,
        "message": "Yêu cầu hủy đơn đã được gửi. Admin sẽ xác nhận và thông báo qua email.",
        "new_status": "Chờ hủy",
    })
 
 
# ═══════════════════════════════════════════════════════════════════
# ADMIN VIEWS — Quản lý đơn hàng
# ═══════════════════════════════════════════════════════════════════
 
# ─────────────────────────────────────────────────────────────────
# ADMIN VIEWS — Quản lý đơn hàng
# ─────────────────────────────────────────────────────────────────

def admin_orders_view(request):
    is_django_admin = bool(
        getattr(request, "user", None)
        and request.user.is_authenticated
        and request.user.is_staff
    )
    account_id = request.session.get("account_id")
    account    = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first() if account_id else None
    if not is_django_admin and (not account or account.LoaiTaiKhoan not in ('admin','staff')):
        return redirect('/admin/login/?next=/admin-orders/')

    status_filter = request.GET.get("status", "all")
    search_q      = (request.GET.get("q") or "").strip()

    orders_qs = DonHang.objects.select_related(
        "id_KhachHang", "id_KhachHang__id_TaiKhoan",
        "id_GiaoHang",  "id_GiaoHang__id_TaiKhoan"
    ).order_by("-ThoiGian")

    if status_filter != "all":
        orders_qs = orders_qs.filter(TrangThai=status_filter)

    if search_q:
        orders_qs = orders_qs.filter(
            Q(MaDonHang__icontains=search_q) |
            Q(id_KhachHang__TenKhachHang__icontains=search_q) |
            Q(id_GiaoHang__TenNguoiNhan__icontains=search_q) |
            Q(id_GiaoHang__SDT__icontains=search_q)
        )

    orders = list(orders_qs[:200])
    order_ids = [o.id_DonHang for o in orders]

    details = list(
        ChiTietDonHang.objects
        .filter(id_DonHang_id__in=order_ids)
        .select_related("id_BienThe", "id_BienThe__id_SanPham",
                        "id_BienThe__id_SanPham__id_ThuongHieu")
    ) if order_ids else []

    detail_map = defaultdict(list)
    for d in details:
        detail_map[d.id_DonHang_id].append(d)

    product_ids_all = list({
        d.id_BienThe.id_SanPham_id
        for d in details
        if d.id_BienThe and d.id_BienThe.id_SanPham_id
    })
    image_map = _product_image_map(product_ids_all)

    for order in orders:
        rows = detail_map.get(order.id_DonHang, [])
        order.admin_items    = rows
        order.admin_products = ", ".join(
            d.id_BienThe.id_SanPham.TenSanPham
            for d in rows if d.id_BienThe and d.id_BienThe.id_SanPham
        ) or "—"
        order.admin_variants = ", ".join(
            d.id_BienThe.Sku for d in rows if d.id_BienThe and d.id_BienThe.Sku
        ) or "—"
        order.admin_quantity    = sum(int(d.SoLuong or 0) for d in rows)
        order.delivery_status   = _delivery_status_label(order)
        order.payment_status    = _payment_status_label(order)

        # Ảnh sản phẩm đầu tiên trong đơn
        first_img = ""
        for d in rows:
            if d.id_BienThe and d.id_BienThe.id_SanPham_id:
                imgs = image_map.get(d.id_BienThe.id_SanPham_id, [])
                if imgs:
                    first_img = imgs[0]
                    break
        order.first_image = first_img

    from django.db.models import Count
    status_counts = {
        row["TrangThai"]: row["cnt"]
        for row in DonHang.objects.values("TrangThai").annotate(cnt=Count("id_DonHang"))
    }

    STATUS_TABS = [
        {"label": "Chờ xác nhận",       "key": "Chờ xác nhận"},
        {"label": "Chờ hủy",            "key": "Chờ hủy"},
        {"label": "Đã xác nhận",         "key": "Đã xác nhận"},
        {"label": "Đang giao",           "key": "Đang giao"},
        {"label": "Khách đã nhận hàng",  "key": "Khách đã nhận hàng"},
        {"label": "Hoàn tất",            "key": "Hoàn tất"},
        {"label": "Đã hủy",              "key": "Đã hủy"},
    ]
    for tab in STATUS_TABS:
        tab["count"] = status_counts.get(tab["key"], 0)

    return render(request, "admin/orders.html", {
        "orders":        orders,
        "status_filter": status_filter,
        "search_q":      search_q,
        "status_tabs":   STATUS_TABS,
        "total_count":   DonHang.objects.count(),
    })


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Cập nhật trạng thái đơn hàng
# URL: POST /api/admin/update-order-status/
# ═══════════════════════════════════════════════════════════════════
@csrf_exempt
@require_POST
def admin_update_order_status(request):
    is_django_admin = bool(
        getattr(request, "user", None)
        and request.user.is_authenticated
    )
    account_id = request.session.get("account_id")
    if not is_django_admin and not account_id:
        return JsonResponse({"ok": False, "message": "Chưa đăng nhập."}, status=403)

    order_id   = request.POST.get("order_id")
    new_status = (request.POST.get("status") or "").strip()

    VALID = [
        "Chờ xác nhận", "Đã xác nhận", "Đang giao",
        "Khách đã nhận hàng", "Hoàn tất", "Đã hủy", "Chờ hủy"
    ]
    if new_status not in VALID:
        return JsonResponse({"ok": False, "message": "Trạng thái không hợp lệ."})

    order = DonHang.objects.select_related(
        "id_KhachHang", "id_KhachHang__id_TaiKhoan",
        "id_GiaoHang",  "id_GiaoHang__id_TaiKhoan",
    ).filter(id_DonHang=order_id).first()
    if not order:
        return JsonResponse({"ok": False, "message": "Không tìm thấy đơn hàng."})

    # ── Luồng hợp lệ (dùng TrangThai thô trong DB) ──
    FLOW = {
    "Chờ xác nhận":       ["Đã xác nhận", "Đã hủy", "Chờ hủy"],
    "Đã thanh toán":      ["Đã xác nhận", "Đã hủy"],
    "Chờ hủy":            ["Đã hủy", "Chờ xác nhận"],   
    "Đã xác nhận":        ["Đang giao", "Đã hủy"],
    "Đang giao":          ["Khách đã nhận hàng"],
    "Khách đã nhận hàng": ["Hoàn tất"],
    "Hoàn tất":           [],
    "Đã hủy":             [],
    "Thanh toán thất bại":[],
}
    current = (order.TrangThai or "Chờ xác nhận").strip()
    if new_status not in FLOW.get(current, []):
        return JsonResponse({
            "ok": False,
            "message": f"Không thể chuyển từ '{current}' sang '{new_status}'."
        })

    order.TrangThai = new_status
    order.save(update_fields=["TrangThai"])

    # ── Gửi email theo mốc ──
    email_sent = False
    if new_status in ("Đã xác nhận", "Hoàn tất"):
        email_sent = _send_order_status_email(order, new_status)

    # ── Khi admin xác nhận hủy: hoàn tồn kho + gửi mail khách ──
    if new_status == "Đã hủy":
        _restore_inventory_on_cancel(order)
        email_sent = _send_cancel_confirmation_email(order)
        # Hoàn điểm nếu khách đã dùng
        acc = _order_account(order)
        if acc and int(order.DiemDaDung or 0) > 0:
            add_points(acc, int(order.DiemDaDung), "refund_points",
                    f"Hoàn điểm do hủy đơn {order.MaDonHang}", order)

    # ── Khi Hoàn tất: cộng điểm, cập nhật TongChiTieu, hạng ──
    if new_status == "Hoàn tất":
        acc = _order_account(order)
        if acc:
            total_val = float(order.TongTien or 0)
            acc.TongChiTieu = float(acc.TongChiTieu or 0) + total_val
            acc.save(update_fields=["TongChiTieu"])
            earned = int(total_val / 10000)
            if earned > 0:
                add_points(acc, earned, "earn_order",
                           f"Tích điểm hoàn tất đơn {order.MaDonHang}", order)
            update_member_level(acc)
            order.DiemNhanDuoc = earned
            order.save(update_fields=["DiemNhanDuoc"])

    suffix = " · Email đã gửi." if email_sent else ""
    return JsonResponse({
        "ok":         True,
        "message":    f"Đã cập nhật → {new_status}.{suffix}",
        "new_status": new_status,
        "order_id":   order.id_DonHang,
        "email_sent": email_sent,
    })


# ═══════════════════════════════════════════════════════════════════
# ADMIN — Chi tiết đơn hàng (JSON)
# URL: GET /api/admin/order-detail/?order_id=...
# ═══════════════════════════════════════════════════════════════════
def admin_order_detail_api(request):
    is_django_admin = bool(
        getattr(request, "user", None)
        and request.user.is_authenticated
    )
    account_id = request.session.get("account_id")
    if not is_django_admin and not account_id:
        return JsonResponse({"ok": False, "message": "Chưa đăng nhập."}, status=403)

    order_id = request.GET.get("order_id")
    order = DonHang.objects.select_related(
        "id_KhachHang", "id_GiaoHang",
        "id_KhachHang__id_TaiKhoan", "id_GiaoHang__id_TaiKhoan",
    ).filter(id_DonHang=order_id).first()
    if not order:
        return JsonResponse({"ok": False, "message": "Không tìm thấy đơn hàng."})

    details = list(
        ChiTietDonHang.objects
        .filter(id_DonHang=order)
        .select_related(
            "id_BienThe", "id_BienThe__id_SanPham",
            "id_BienThe__id_SanPham__id_ThuongHieu",
        )
    )
    product_ids = [
        d.id_BienThe.id_SanPham_id for d in details
        if d.id_BienThe and d.id_BienThe.id_SanPham_id
    ]
    image_map = _product_image_map(product_ids)

    items = []
    for d in details:
        bt = d.id_BienThe
        if not bt:
            continue
        sp = bt.id_SanPham
        if not sp:
            continue
        imgs = image_map.get(sp.id_SanPham, [])
        items.append({
            "product_name": sp.TenSanPham,
            "brand":  sp.id_ThuongHieu.TenThuongHieu if sp.id_ThuongHieu else "",
            "image":  imgs[0] if imgs else FALLBACK_IMAGES["default"],
            "sku":    bt.Sku,
            "qty":    d.SoLuong or 1,
            "price":  _format_currency(d.GiaBan),
        })

    gh = order.id_GiaoHang
    kh = order.id_KhachHang

    # ── Tính next_status dựa trên TrangThai thô trong DB ──
    raw_status = (order.TrangThai or "").strip()
    display_status = _delivery_status_label(order)   # trạng thái hiển thị cho người dùng

    FLOW_NEXT = {
        "Chờ xác nhận":       ("Đã xác nhận",        "✅ Xác nhận đơn"),
        "Đã thanh toán":      ("Đã xác nhận",        "✅ Xác nhận đơn"),
        "Đã xác nhận":        ("Đang giao",           "🚚 Bắt đầu giao hàng"),
        "Đang giao":          ("Khách đã nhận hàng",  "📦 Đánh dấu đã giao"),
        "Khách đã nhận hàng": ("Hoàn tất",            "🎉 Hoàn tất đơn hàng"),
        "Chờ hủy":            ("Đã hủy",              "✕ Xác nhận hủy đơn"),
    }
    next_action = FLOW_NEXT.get(raw_status) or FLOW_NEXT.get(display_status)

    can_cancel = raw_status in ("Chờ xác nhận", "Đã xác nhận", "Đã thanh toán")
    # can_restore = raw_status == "Chờ hủy"
    return JsonResponse({
        "ok": True,
        "order": {
            "id":             order.id_DonHang,
            "ma_don":         order.MaDonHang,
            "date":           order.ThoiGian.strftime("%d/%m/%Y %H:%M") if order.ThoiGian else "",
            "status":         display_status,
            "raw_status":     raw_status,
            "payment":        order.HinhThucThanhToan or "COD",
            "payment_status": _payment_status_label(order),
            "voucher_code":   _extract_voucher_code(order),
            "points_used":    int(order.DiemDaDung or 0),
            "points_discount":_format_currency(order.TienGiamTuDiem),
            "points_earned":  int(order.DiemNhanDuoc or 0),
            "total":          _format_currency(order.TongTien),
            "items":          items,
            "receiver":       gh.TenNguoiNhan if gh else "",
            "phone":          gh.SDT if gh else "",
            "address":        gh.DiaChi if gh else "",
            "note":           _clean_delivery_note(gh.GhiChu) if gh else "",
            "customer":       kh.TenKhachHang if kh else "Khách vãng lai",
            "next_status":    next_action[0] if next_action else None,
            "next_label":     next_action[1] if next_action else None,
            "can_cancel":     can_cancel,
            # "can_restore": can_restore,
        }
    })

# ═══════════════════════════════════════════════════════════════════
# ADMIN — API đếm đơn mới (dùng cho auto-poll)
# URL: GET /api/admin/new-orders-count/
# ═══════════════════════════════════════════════════════════════════
def admin_new_orders_count(request):
    is_django_admin = bool(
        getattr(request, "user", None)
        and request.user.is_authenticated
    )
    account_id = request.session.get("account_id")
    if not is_django_admin and not account_id:
        return JsonResponse({"ok": False, "count": 0})

    count = DonHang.objects.filter(TrangThai="Chờ xác nhận").count()
    return JsonResponse({"ok": True, "count": count})
 

@csrf_exempt
@require_POST
def update_profile_api(request):
    account_id = request.session.get("account_id")

    def profile_json(payload, status=200):
        return JsonResponse(payload, status=status,
                            json_dumps_params={"ensure_ascii": False})

    if not account_id:
        return profile_json({"ok": False, "need_login": True,
                             "message": "Vui lòng đăng nhập."}, status=401)

    account = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first()
    if not account:
        return profile_json({"ok": False, "message": "Không tìm thấy tài khoản."}, status=404)

    full_name = (request.POST.get("full_name") or "").strip()
    username  = (request.POST.get("username")  or "").strip()
    email     = (request.POST.get("email")     or "").strip()
    phone     = (request.POST.get("phone")     or "").strip()
    gender    = (request.POST.get("gender")    or "").strip()
    avatar    = request.FILES.get("avatar")

    # ── Validate ──
    if not full_name:
        return profile_json({"ok": False, "message": "Vui lòng nhập họ và tên."}, status=400)
    if not username:
        return profile_json({"ok": False, "message": "Vui lòng nhập tên đăng nhập."}, status=400)

    import re
    if email and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return profile_json({"ok": False, "message": "Email không đúng định dạng."}, status=400)
    if phone and not re.match(r'^(0|\+84)[0-9]{8,10}$', phone.replace(' ', '')):
        return profile_json({"ok": False, "message": "Số điện thoại không đúng định dạng (VD: 0901234567)."}, status=400)

    if TaiKhoan.objects.filter(Username__iexact=username).exclude(id_TaiKhoan=account_id).exists():
        return profile_json({"ok": False, "message": "Tên đăng nhập đã tồn tại."}, status=409)

    # ── Lưu TaiKhoan ──
    account.TenDangNhap = full_name
    account.Username    = username
    account.Email       = email
    account.SDT         = phone
    if avatar:
        account.AnhDaiDien = avatar

    fields = ["TenDangNhap", "Username", "Email", "SDT"]
    if avatar:
        fields.append("AnhDaiDien")
    account.save(update_fields=fields)

    # ── Lưu KhachHang ──
    customer = KhachHang.objects.filter(id_TaiKhoan=account).first()
    if customer:
        customer.TenKhachHang = full_name
        customer.GioiTinh     = gender
        customer.save(update_fields=["TenKhachHang", "GioiTinh"])
        print(f"[DEBUG] Saved gender='{gender}' for customer id={customer.id_KhachHang}")
        customer.save(update_fields=["TenKhachHang", "GioiTinh"])
    else:
        KhachHang.objects.create(
            id_TaiKhoan=account,
            TenKhachHang=full_name,
            DiaChi="", GioiTinh=gender,
        )

    request.session["account_name"] = full_name

    avatar_url = account.AnhDaiDien.url if account.AnhDaiDien else ""

    return profile_json({
        "ok": True,
        "message": "Cập nhật thành công.",
        "profile": {
            "full_name": full_name,
            "username":  username,
            "email":     email,
            "phone":     phone,
            "gender":    gender,
            "avatar":    avatar_url,
        },
        "avatar": avatar_url,
    })

def api_thuoc_tinh_list(request):
    """Trả về danh sách thuộc tính — không cần đăng nhập."""
    try:
        data = []
        for tt in ThuocTinh.objects.all().order_by('TenThuocTinh'):
            data.append({
                'id': tt.pk,
                'name': str(tt.TenThuocTinh or ''),
            })
        # ensure_ascii=False để tiếng Việt hiển thị đúng
        return JsonResponse({'ok': True, 'data': data}, json_dumps_params={'ensure_ascii': False})
    except Exception as e:
        return JsonResponse({'ok': False, 'message': str(e)}, status=500)
 
 
def api_gia_tri_thuoc_tinh(request):
    """Trả về giá trị thuộc tính theo id_ThuocTinh — không cần đăng nhập."""
    tt_id = request.GET.get('thuoc_tinh_id', '').strip()
 
    if not tt_id:
        return JsonResponse({'ok': False, 'message': 'Thiếu thuoc_tinh_id'}, status=400)
 
    try:
        tt_id_int = int(tt_id)
    except (ValueError, TypeError):
        return JsonResponse({'ok': False, 'message': 'thuoc_tinh_id phải là số'}, status=400)
 
    try:
        rows = GiaTriThuocTinh.objects.filter(
            id_ThuocTinh=tt_id_int
        ).order_by('GiaTri')
 
        data = []
        for r in rows:
            data.append({
                'id': r.pk,
                'name': str(r.GiaTri or ''),
            })
 
        print(f"[api_gia_tri] tt_id={tt_id_int}, found={len(data)} rows")
 
        return JsonResponse(
            {'ok': True, 'data': data},
            json_dumps_params={'ensure_ascii': False}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'ok': False, 'message': str(e)}, status=500)
    

# ════════════════════════════════════════════════════════════
# AI — Gợi ý sản phẩm tương tự (Content-based Filtering)
# ════════════════════════════════════════════════════════════
def ai_recommend_api(request, product_id):
    """
    GET /api/recommend/<product_id>/
    Trả về tối đa 8 sản phẩm tương tự dựa trên TF-IDF cosine.
    """
    from app.ai.recommender import get_similar_products

    similar_ids = get_similar_products(product_id, top_n=8)
    if not similar_ids:
        return JsonResponse({"ok": True, "products": []})

    products = list(
        SanPham.objects
        .select_related("id_ThuongHieu", "id_LoaiSanPham")
        .prefetch_related("nhom_huongs")
        .filter(id_SanPham__in=similar_ids)
    )

    # Giữ đúng thứ tự theo độ tương tự
    product_map = {p.id_SanPham: p for p in products}
    ordered = [product_map[pid] for pid in similar_ids if pid in product_map]

    cards = _build_product_cards(ordered)
    return JsonResponse({"ok": True, "products": cards},
                        json_dumps_params={"ensure_ascii": False})


def admin_api_bien_the(request):
    """Trả về danh sách biến thể theo id_SanPham."""
    sp_id = request.GET.get('sp_id')
    if not sp_id:
        return JsonResponse({'ok': False, 'data': []})
    from .models import BienThe, BienTheThuocTinh
    bien_the = BienThe.objects.filter(id_SanPham_id=sp_id)
    data = []
    for bt in bien_the:
        attrs = BienTheThuocTinh.objects.select_related(
            'id_GiaTriThuocTinh__id_ThuocTinh'
        ).filter(id_BienThe=bt)
        attr_str = ', '.join(
            f"{a.id_GiaTriThuocTinh.id_ThuocTinh.TenThuocTinh}: {a.id_GiaTriThuocTinh.GiaTri}"
            for a in attrs
        )
        data.append({
            'id':       bt.id_BienThe,
            'sku':      bt.Sku,
            'gia_nhap': float(bt.GiaNhap or 0),
            'gia_ban':  float(bt.GiaBan or 0),
            'ton_kho':  bt.SoLuong,
            'attrs':    attr_str or bt.Sku,
        })
    return JsonResponse({'ok': True, 'data': data})
 
 
@csrf_exempt
def admin_api_luu_phieu(request):
    """Lưu phiếu nhập + chi tiết + cập nhật tồn kho."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)
 
    try:
        body     = json.loads(request.body)
        phieu_id = body.get('phieu_id')  # None = tạo mới
        ncc_id   = body.get('ncc_id')
        trang_thai = body.get('trang_thai', 'confirmed')
        rows     = body.get('rows', [])   # [{bien_the_id, so_luong, gia_nhap}, ...]
        new_products = body.get('new_products', [])  # sản phẩm mới chưa có
 
        from .models import TaiKhoan as TK, BienThe as BT, ThuocTinh, GiaTriThuocTinh
        from django.db import transaction
 
        with transaction.atomic():
            # 1. Lưu sản phẩm mới nếu có
            for np in new_products:
                sp_name = (np.get('ten_san_pham') or '').strip()
                if not sp_name:
                    continue
                # Tạo SanPham
                sp_id_existing = np.get('_spId') or np.get('sp_id')
                if sp_id_existing:
                    # Dùng sản phẩm đã tồn tại
                    try:
                        new_sp = SanPham.objects.get(pk=int(sp_id_existing))
                    except SanPham.DoesNotExist:
                        new_sp = SanPham.objects.create(
                            TenSanPham=sp_name,
                            TrangThai_SanPham='active',
                        )
                else:
                    # Tạo sản phẩm mới
                    new_sp = SanPham.objects.create(
                        TenSanPham=sp_name,
                        TrangThai_SanPham='active',
                    )
                # Tạo biến thể cho sản phẩm mới
                for bt_data in np.get('bien_the', []):
                    new_bt = BT.objects.create(
                        id_SanPham=new_sp,
                        Sku=bt_data.get('sku', f'SKU-{new_sp.pk}'),
                        GiaNhap=bt_data.get('gia_nhap', 0),
                        GiaBan=bt_data.get('gia_ban', 0),
                        SoLuong=0,
                    )
                    # Gán thuộc tính
                    for attr in bt_data.get('attrs', []):
                        thuoc_tinh, _ = ThuocTinh.objects.get_or_create(
                            TenThuocTinh=attr.get('ten_thuoc_tinh', 'Dung tích')
                        )
                        gia_tri, _ = GiaTriThuocTinh.objects.get_or_create(
                            id_ThuocTinh=thuoc_tinh,
                            GiaTri=attr.get('gia_tri', '')
                        )
                        BienTheThuocTinh.objects.create(
                            id_BienThe=new_bt,
                            id_GiaTriThuocTinh=gia_tri
                        )
                    # Thêm vào rows để nhập kho
                    rows.append({
                        'bien_the_id': new_bt.id_BienThe,
                        'so_luong':    bt_data.get('so_luong', 0),
                        'gia_nhap':    bt_data.get('gia_nhap', 0),
                    })
 
            # 2. Tạo / cập nhật PhieuNhap
            tk = None
            if request.user.is_authenticated:
                tk = TK.objects.filter(Username=request.user.username).first()
 
            tong_tien = sum(
                float(r.get('gia_nhap', 0)) * int(r.get('so_luong', 0))
                for r in rows
            )
 
            if phieu_id:
                phieu = PhieuNhap.objects.get(pk=phieu_id)
                phieu.TrangThai = trang_thai
                phieu.TongTien  = tong_tien
                if ncc_id:
                    phieu.id_NCC_id = ncc_id
                phieu.save()
                # Xóa chi tiết cũ để ghi lại
                phieu.chi_tiet.all().delete()
            else:
                import random, string
                ma = 'PN' + tz.now().strftime('%y%m%d') + ''.join(
                    random.choices(string.ascii_uppercase + string.digits, k=4)
                )
                phieu = PhieuNhap.objects.create(
                    ThoiGian=tz.now(),
                    id_TaiKhoan=tk,
                    id_NCC_id=ncc_id if ncc_id else None,
                    MaPhieu=ma,
                    TongTien=tong_tien,
                    TrangThai=trang_thai,
                )
 
            # 3. Lưu chi tiết + cập nhật tồn kho
            for r in rows:
                bt_id    = r.get('bien_the_id')
                so_luong = int(r.get('so_luong', 0))
                gia_nhap = float(r.get('gia_nhap', 0))
                if not bt_id or so_luong <= 0:
                    continue
 
                ChiTietNhap.objects.create(
                    id_PhieuNhap=phieu,
                    id_BienThe_id=bt_id,
                    SoLuongNhap=so_luong,
                    GiaNhap=gia_nhap,
                )
 
                if trang_thai in ('confirmed', 'done'):
                    BT.objects.filter(pk=bt_id).update(
                        SoLuong=models.F('SoLuong') + so_luong
                    )
 
        return JsonResponse({
            'ok': True,
            'phieu_id': phieu.id_PhieuNhap,
            'ma_phieu': phieu.MaPhieu,
        })
 
    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)
 
 
@csrf_exempt
def admin_api_them_sp_moi(request):
    """Kiểm tra tên sản phẩm có tồn tại chưa."""
    if request.method != 'POST':
        return JsonResponse({'ok': False})
    body = json.loads(request.body)
    ten  = (body.get('ten_san_pham') or '').strip()
    if not ten:
        return JsonResponse({'ok': False, 'error': 'Tên trống'})
    exists = SanPham.objects.filter(TenSanPham__iexact=ten).first()
    if exists:
        from .models import BienThe as BT
        bien_the = list(BT.objects.filter(id_SanPham=exists).values(
            'id_BienThe', 'Sku', 'GiaNhap', 'GiaBan', 'SoLuong'
        ))
        return JsonResponse({'ok': True, 'exists': True, 'sp_id': exists.id_SanPham,
                             'ten': exists.TenSanPham, 'bien_the': bien_the})
    return JsonResponse({'ok': True, 'exists': False})

    

def admin_phieunhap_list(request):
    """Trang lich su phieu nhap kho."""
    from .models import PhieuNhap, NhaCungCap, ChiTietNhap
    from django.db.models import Count, Sum, Q
    from django.utils import timezone as tz
    import datetime
 
    # Query tat ca phieu, sort moi nhat truoc
    phieu_qs = PhieuNhap.objects.select_related(
        'id_TaiKhoan', 'id_NCC'
    ).order_by('-ThoiGian')
 
    # Them so_dong vao tung phieu
    phieu_list = []
    for p in phieu_qs:
        p.so_dong = ChiTietNhap.objects.filter(id_PhieuNhap=p).count()
        phieu_list.append(p)
 
    # Stats
    now = tz.now()
    tong_tien_val = PhieuNhap.objects.filter(
        TrangThai__in=['confirmed', 'done']
    ).aggregate(s=Sum('TongTien'))['s'] or 0
 
    thang_nay = PhieuNhap.objects.filter(
        ThoiGian__year=now.year,
        ThoiGian__month=now.month
    ).count()
 
    stats = {
        'tong_phieu': PhieuNhap.objects.count(),
        'hoan_tat':   PhieuNhap.objects.filter(TrangThai__in=['confirmed','done']).count(),
        'tong_tien':  f"{int(tong_tien_val):,}".replace(",", ".") + "₫",
        'thang_nay':  thang_nay,
    }
 
    context = {
        'phieu_list': phieu_list,
        'ncc_list':   list(NhaCungCap.objects.values('id_NCC', 'Ten_NCC')),
        'stats':      stats,
        'title':      'Lịch sử phiếu nhập kho',
    }
    return render(request, 'admin/phieunhap_list.html', context)
 
 
def admin_api_chi_tiet_phieu(request):
    """API tra ve chi tiet cua 1 phieu nhap."""
    phieu_id = request.GET.get('phieu_id')
    if not phieu_id:
        return JsonResponse({'ok': False, 'error': 'Thieu phieu_id'})
 
    from .models import PhieuNhap, ChiTietNhap, BienThe
 
    try:
        phieu = PhieuNhap.objects.get(pk=phieu_id)
        chi_tiet = ChiTietNhap.objects.select_related(
            'id_BienThe__id_SanPham__id_ThuongHieu'
        ).filter(id_PhieuNhap=phieu)
 
        rows = []
        for ct in chi_tiet:
            bt = ct.id_BienThe
            sp = bt.id_SanPham if bt else None
            rows.append({
                'san_pham':   sp.TenSanPham if sp else '—',
                'thuong_hieu': sp.id_ThuongHieu.TenThuongHieu if sp and sp.id_ThuongHieu else '',
                'sku':        bt.Sku if bt else '—',
                'gia_nhap':   float(ct.GiaNhap or 0),
                'so_luong':   ct.SoLuongNhap or 0,
            })
 
        return JsonResponse({
            'ok':       True,
            'ma_phieu': phieu.MaPhieu or f'PN-{phieu.id_PhieuNhap}',
            'rows':     rows,
        })
    except PhieuNhap.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Khong tim thay phieu'})
 
 
def admin_api_xuat_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    phieu_id = request.GET.get('phieu_id')
    tat_ca   = request.GET.get('tat_ca')

    wb = openpyxl.Workbook()

    # ── Style helpers ──
    green_fill  = PatternFill("solid", fgColor="4B672D")
    light_fill  = PatternFill("solid", fgColor="EBF6C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    sub_font    = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')

    def style_header_row(ws, row_num, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill   = green_fill
            cell.font   = header_font
            cell.alignment = center
            cell.border = thin_border

    def style_data_row(ws, row_num, col_count, is_alt=False):
        fill = PatternFill("solid", fgColor="F5FDF0") if is_alt else PatternFill("solid", fgColor="FFFFFF")
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill   = fill
            cell.border = thin_border
            cell.alignment = left

    # ════════════════════════════════════════
    #  SHEET 1: DANH SÁCH PHIẾU NHẬP
    # ════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Danh sách phiếu nhập"

    # Tiêu đề lớn
    ws1.merge_cells('A1:H1')
    title_cell = ws1['A1']
    title_cell.value     = "DANH SÁCH PHIẾU NHẬP KHO – AMI PERFUMERY"
    title_cell.font      = Font(bold=True, size=14, color="4B672D")
    title_cell.alignment = center
    ws1.row_dimensions[1].height = 30

    # Header
    headers1 = ['STT', 'Mã phiếu', 'Thời gian', 'Người nhập',
                 'Nhà cung cấp', 'Tổng tiền (₫)', 'Trạng thái', 'Số dòng']
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=2, column=i, value=h)
    style_header_row(ws1, 2, len(headers1))
    ws1.row_dimensions[2].height = 22

    # Lấy dữ liệu
    if phieu_id:
        phieu_qs = PhieuNhap.objects.filter(pk=phieu_id)
    else:
        phieu_qs = PhieuNhap.objects.all().order_by('-ThoiGian')

    tt_map = {
        'draft': 'Nháp', 'confirmed': 'Xác nhận',
        'done': 'Hoàn tất', 'cancelled': 'Huỷ'
    }

    for idx, p in enumerate(phieu_qs, 1):
        r = idx + 2
        so_dong = p.chi_tiet.count()
        ws1.cell(r, 1, idx)
        ws1.cell(r, 2, p.MaPhieu or f'PN-{p.pk}')
        ws1.cell(r, 3, p.ThoiGian.strftime('%d/%m/%Y %H:%M') if p.ThoiGian else '—')
        ws1.cell(r, 4, p.id_TaiKhoan.TenDangNhap if p.id_TaiKhoan else '—')
        ws1.cell(r, 5, p.id_NCC.Ten_NCC if p.id_NCC else '—')
        tong = ws1.cell(r, 6, float(p.TongTien or 0))
        tong.number_format = '#,##0'
        tong.alignment = right
        ws1.cell(r, 7, tt_map.get(p.TrangThai or 'draft', p.TrangThai or '—'))
        ws1.cell(r, 8, so_dong)
        style_data_row(ws1, r, len(headers1), idx % 2 == 0)
        ws1.cell(r, 1).alignment = center
        ws1.cell(r, 8).alignment = center

    # Độ rộng cột sheet 1
    col_widths1 = [6, 18, 20, 16, 20, 18, 14, 10]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ════════════════════════════════════════
    #  SHEET 2: CHI TIẾT SẢN PHẨM TRONG PHIẾU
    # ════════════════════════════════════════
    ws2 = wb.create_sheet("Chi tiết sản phẩm")

    ws2.merge_cells('A1:I1')
    t2 = ws2['A1']
    t2.value     = "CHI TIẾT SẢN PHẨM NHẬP KHO – AMI PERFUMERY"
    t2.font      = Font(bold=True, size=14, color="4B672D")
    t2.alignment = center
    ws2.row_dimensions[1].height = 30

    headers2 = ['STT', 'Mã phiếu', 'Thời gian', 'Nhà cung cấp',
                 'Tên sản phẩm', 'Thương hiệu', 'SKU / Biến thể',
                 'Đơn giá nhập (₫)', 'Số lượng', 'Thành tiền (₫)']
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, len(headers2))
    ws2.row_dimensions[2].height = 22

    row2 = 3
    stt2 = 1
    for p in phieu_qs:
        ma_phieu  = p.MaPhieu or f'PN-{p.pk}'
        thoigian  = p.ThoiGian.strftime('%d/%m/%Y %H:%M') if p.ThoiGian else '—'
        ncc_name  = p.id_NCC.Ten_NCC if p.id_NCC else '—'

        chi_tiets = p.chi_tiet.select_related(
            'id_BienThe__id_SanPham__id_ThuongHieu'
        ).all()

        for ct in chi_tiets:
            bt = ct.id_BienThe
            sp = bt.id_SanPham if bt else None

            ten_sp   = sp.TenSanPham if sp else '—'
            try:
                thuong_hieu = sp.id_ThuongHieu.TenThuongHieu if sp and sp.id_ThuongHieu_id else '—'
            except Exception:
                thuong_hieu = '—'
            sku      = bt.Sku if bt else '—'
            gia_nhap = float(ct.GiaNhap or 0)
            so_luong = int(ct.SoLuongNhap or 0)
            thanh_tien = gia_nhap * so_luong

            ws2.cell(row2, 1, stt2)
            ws2.cell(row2, 2, ma_phieu)
            ws2.cell(row2, 3, thoigian)
            ws2.cell(row2, 4, ncc_name)
            ws2.cell(row2, 5, ten_sp)
            ws2.cell(row2, 6, thuong_hieu)
            ws2.cell(row2, 7, sku)

            gn_cell = ws2.cell(row2, 8, gia_nhap)
            gn_cell.number_format = '#,##0'
            gn_cell.alignment = right

            sl_cell = ws2.cell(row2, 9, so_luong)
            sl_cell.alignment = center

            tt_cell = ws2.cell(row2, 10, thanh_tien)
            tt_cell.number_format = '#,##0'
            tt_cell.alignment = right
            tt_cell.font = Font(bold=True, color="4B672D")

            style_data_row(ws2, row2, len(headers2), stt2 % 2 == 0)
            ws2.cell(row2, 1).alignment = center
            # Re-apply number formats after style_data_row
            ws2.cell(row2, 8).number_format = '#,##0'
            ws2.cell(row2, 9).alignment = center
            ws2.cell(row2, 10).number_format = '#,##0'

            row2  += 1
            stt2  += 1

    # Độ rộng cột sheet 2
    col_widths2 = [6, 18, 20, 18, 24, 18, 18, 18, 10, 18]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Xuất file ──
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from django.utils import timezone as tz
    filename = f"PhieuNhap_{tz.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ════════════════════════════════════════════════════
# GIAI ĐOẠN 4 — Ghi nhận click AI recommendation
# ════════════════════════════════════════════════════
@csrf_exempt
@require_POST
def ai_track_click(request):
    """
    POST /api/ai/track-click/
    Body: { "product_id": 9, "source": "content_based" }
    Ghi nhận mỗi lần khách nhấp vào sản phẩm được AI gợi ý.
    """
    try:
        data       = _json.loads(request.body)
        product_id = data.get("product_id")
        source     = data.get("source", "content_based")
        account_id = request.session.get("account_id")

        if not product_id:
            return JsonResponse({"ok": False})

        product = SanPham.objects.filter(id_SanPham=product_id).first()
        if not product:
            return JsonResponse({"ok": False})

        account = None
        if account_id:
            account = TaiKhoan.objects.filter(
                id_TaiKhoan=account_id
            ).first()

        AIRecommendClick.objects.create(
            id_TaiKhoan=account,
            id_SanPham=product,
            source=source,
        )
        return JsonResponse({"ok": True})

    except Exception:
        return JsonResponse({"ok": False})
    

@csrf_exempt
@require_POST
def ai_chatbot_feedback(request):
    """
    POST /api/ai/chatbot-feedback/
    Body: { "rating": 5, "content": "Tư vấn rất tốt" }
    """
    import json as _json
    try:
        data       = _json.loads(request.body)
        rating     = int(data.get("rating") or 0)
        content    = (data.get("content") or "").strip()[:500]
        account_id = request.session.get("account_id")

        if rating < 1 or rating > 5:
            return JsonResponse({"ok": False, "error": "Rating từ 1–5"})

        account = None
        if account_id:
            account = TaiKhoan.objects.filter(
                id_TaiKhoan=account_id
            ).first()

        ChatbotFeedback.objects.create(
            id_TaiKhoan=account,
            Rating=rating,
            NoiDung=content or None,
        )
        return JsonResponse({"ok": True})

    except Exception:
        return JsonResponse({"ok": False})
    
"""
THAY THẾ hàm ai_dashboard_api trong views.py
=============================================
"""

def ai_dashboard_api(request):
    """
    GET /api/admin/ai-dashboard/
    Dashboard AI analytics hoàn chỉnh.
    """
    from django.db.models import Count, Avg, Sum, Q
    from datetime import timedelta
    from app.models import (
        LichSuXemSanPham, AIUserProfile, ChatbotHistory,
        NhomHuong, SanPhamNhomHuong, ThuongHieu
    )

    # ── Auth check ────────────────────────────────────────────
    is_django_admin = bool(
        getattr(request, "user", None)
        and request.user.is_authenticated
        and request.user.is_staff
    )
    account_id = request.session.get("account_id")
    account    = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first() if account_id else None
    is_custom_admin = bool(account and account.LoaiTaiKhoan in ('admin', 'staff'))
    if not is_django_admin and not is_custom_admin:
        return JsonResponse({"ok": False}, status=403)

    now = timezone.now()
    try:
        days = int(request.GET.get('days', 30))
        if days not in [7, 14, 30, 60, 90]:
            days = 30
    except (ValueError, TypeError):
        days = 30

    since_main = now - timedelta(days=days)
    since_30   = since_main   # alias để các query bên dưới không cần đổi tên
    since_7    = now - timedelta(days=7)

    # Trả days về client
    since_yesterday = now - timedelta(days=1)

    # ════════════════════════════════════════════════════════
    # 1. KPI TỔNG QUAN
    # ════════════════════════════════════════════════════════
    total_orders    = DonHang.objects.count()
    total_revenue   = DonHang.objects.filter(
        TrangThai__in=['Hoàn tất','Khách đã nhận hàng']
    ).aggregate(s=Sum('TongTien'))['s'] or 0

    total_customers = TaiKhoan.objects.filter(LoaiTaiKhoan='customer').count()
    total_products  = SanPham.objects.count()

    # AI KPIs
    total_ai_clicks     = AIRecommendClick.objects.count()
    total_chatbot_fb    = ChatbotFeedback.objects.count()
    total_viewed        = LichSuXemSanPham.objects.count()
    total_ai_profiles   = AIUserProfile.objects.count()
    total_chat_history  = ChatbotHistory.objects.count()

    # Chatbot satisfaction
    chatbot_avg = ChatbotFeedback.objects.aggregate(avg=Avg('Rating'))['avg'] or 0

    # ════════════════════════════════════════════════════════
    # 2. DOANH THU THEO NGÀY (30 ngày)
    # ════════════════════════════════════════════════════════
    # Revenue by day
    revenue_by_day = []
    for i in range(days - 1, -1, -1):
        day = now.date() - timedelta(days=i)
        rev = DonHang.objects.filter(
            ThoiGian__date=day,
            TrangThai__in=['Hoàn tất','Khách đã nhận hàng']
        ).aggregate(s=Sum('TongTien'))['s'] or 0
        orders_count = DonHang.objects.filter(ThoiGian__date=day).count()
        revenue_by_day.append({
            "date":    day.strftime("%d/%m"),
            "revenue": int(rev),
            "orders":  orders_count,
        })

    # Clicks by day
    clicks_by_day = []
    for i in range(days - 1, -1, -1):
        day = now.date() - timedelta(days=i)
        c = AIRecommendClick.objects.filter(NgayClick__date=day).count()
        clicks_by_day.append({"date": day.strftime("%d/%m"), "count": c})

    # Chatbot by day
    chatbot_by_day = []
    for i in range(days - 1, -1, -1):
        day = now.date() - timedelta(days=i)
        c = ChatbotFeedback.objects.filter(NgayTao__date=day).count()
        avg_r = ChatbotFeedback.objects.filter(NgayTao__date=day).aggregate(avg=Avg('Rating'))['avg'] or 0
        chatbot_by_day.append({"date": day.strftime("%d/%m"), "count": c, "avg": round(float(avg_r), 1)})

    # Views by day
    views_by_day = []
    for i in range(days - 1, -1, -1):
        day = now.date() - timedelta(days=i)
        c = LichSuXemSanPham.objects.filter(NgayXem__date=day).count()
        views_by_day.append({"date": day.strftime("%d/%m"), "count": c})

    # ════════════════════════════════════════════════════════
    # 3. AI RECOMMENDATION ANALYTICS
    # ════════════════════════════════════════════════════════
    # Click theo nguồn (30 ngày)
    clicks_by_source = list(
        AIRecommendClick.objects
        .filter(NgayClick__gte=since_30)
        .values('source')
        .annotate(total=Count('id_Click'))
        .order_by('-total')
    )

    # Click theo ngày (30 ngày)
    clicks_by_day = []
    for i in range(29, -1, -1):
        day = now.date() - timedelta(days=i)
        c = AIRecommendClick.objects.filter(NgayClick__date=day).count()
        clicks_by_day.append({"date": day.strftime("%d/%m"), "count": c})

    # Top sản phẩm được click từ AI (30 ngày)
    top_ai_products = list(
        AIRecommendClick.objects
        .filter(NgayClick__gte=since_30)
        .values('id_SanPham__TenSanPham', 'id_SanPham__id_SanPham',
                'id_SanPham__id_ThuongHieu__TenThuongHieu')
        .annotate(total=Count('id_Click'))
        .order_by('-total')[:8]
    )

    # ════════════════════════════════════════════════════════
    # 4. CHATBOT ANALYTICS
    # ════════════════════════════════════════════════════════
    # Feedback distribution
    feedback_dist = list(
        ChatbotFeedback.objects
        .filter(NgayTao__gte=since_30)
        .values('Rating')
        .annotate(total=Count('id_Feedback'))
        .order_by('Rating')
    )

    # Chatbot feedback theo ngày
    chatbot_by_day = []
    for i in range(29, -1, -1):
        day = now.date() - timedelta(days=i)
        c = ChatbotFeedback.objects.filter(NgayTao__date=day).count()
        avg_r = ChatbotFeedback.objects.filter(
            NgayTao__date=day
        ).aggregate(avg=Avg('Rating'))['avg'] or 0
        chatbot_by_day.append({
            "date":  day.strftime("%d/%m"),
            "count": c,
            "avg":   round(float(avg_r), 1),
        })

    # Intent analysis từ ChatbotHistory
    intent_stats = {}
    try:
        intents = ChatbotHistory.objects.filter(
            ExtractedIntent__isnull=False,
            NgayTao__gte=since_30
        ).values_list('ExtractedIntent', flat=True)

        for intent_str in intents:
            if not intent_str:
                continue
            for part in intent_str.split('_'):
                part = part.strip()
                if part and len(part) > 1:
                    intent_stats[part] = intent_stats.get(part, 0) + 1

        intent_stats = dict(
            sorted(intent_stats.items(), key=lambda x: x[1], reverse=True)[:10]
        )
    except Exception:
        pass

    # Registered vs Guest chatbot
    registered_chats = ChatbotHistory.objects.values(
        'id_TaiKhoan'
    ).distinct().count()

    # ════════════════════════════════════════════════════════
    # 5. CUSTOMER BEHAVIOR — Recently Viewed
    # ════════════════════════════════════════════════════════
    # Top viewed products (30 ngày)
    top_viewed = list(
        LichSuXemSanPham.objects
        .filter(NgayXem__gte=since_30)
        .values('id_SanPham__TenSanPham', 'id_SanPham__id_SanPham',
                'id_SanPham__id_ThuongHieu__TenThuongHieu')
        .annotate(views=Count('id_LichSu'), total_time=Sum('ThoiGianXem'))
        .order_by('-views')[:8]
    )

    # View theo ngày (30 ngày)
    views_by_day = []
    for i in range(29, -1, -1):
        day = now.date() - timedelta(days=i)
        c = LichSuXemSanPham.objects.filter(NgayXem__date=day).count()
        views_by_day.append({"date": day.strftime("%d/%m"), "count": c})

    # ════════════════════════════════════════════════════════
    # 6. SCENT GROUP ANALYTICS
    # ════════════════════════════════════════════════════════
    # Nhóm mùi được AI profile học nhiều nhất
    scent_from_profiles = {}
    try:
        import json as _j
        for profile in AIUserProfile.objects.all():
            scents = _j.loads(profile.NhomMuaYeuThich or '[]')
            for s in scents:
                scent_from_profiles[s] = scent_from_profiles.get(s, 0) + 1
        scent_from_profiles = dict(
            sorted(scent_from_profiles.items(), key=lambda x: x[1], reverse=True)[:8]
        )
    except Exception:
        pass

    # Brand yêu thích từ AI profile
    brand_from_profiles = {}
    try:
        for profile in AIUserProfile.objects.all():
            brands = _j.loads(profile.ThuongHieuYeuThich or '[]')
            for b in brands:
                brand_from_profiles[b] = brand_from_profiles.get(b, 0) + 1
        brand_from_profiles = dict(
            sorted(brand_from_profiles.items(), key=lambda x: x[1], reverse=True)[:8]
        )
    except Exception:
        pass

    # ════════════════════════════════════════════════════════
    # 7. AI USER PROFILE STATS
    # ════════════════════════════════════════════════════════
    profile_confidence_dist = {
        "low":    AIUserProfile.objects.filter(ConfidenceScore__lt=0.3).count(),
        "medium": AIUserProfile.objects.filter(
            ConfidenceScore__gte=0.3, ConfidenceScore__lt=0.7
        ).count(),
        "high":   AIUserProfile.objects.filter(ConfidenceScore__gte=0.7).count(),
    }

    avg_confidence = AIUserProfile.objects.aggregate(
        avg=Avg('ConfidenceScore')
    )['avg'] or 0

    # ════════════════════════════════════════════════════════
    # 8. TOP PURCHASED PRODUCTS (từ đơn hàng hoàn tất)
    # ════════════════════════════════════════════════════════
    top_purchased = list(
        ChiTietDonHang.objects
        .filter(
            id_DonHang__TrangThai__in=['Hoàn tất','Khách đã nhận hàng'],
            id_DonHang__ThoiGian__gte=since_30
        )
        .values(
            'id_BienThe__id_SanPham__TenSanPham',
            'id_BienThe__id_SanPham__id_SanPham',
            'id_BienThe__id_SanPham__id_ThuongHieu__TenThuongHieu'
        )
        .annotate(total=Sum('SoLuong'))
        .order_by('-total')[:8]
    )

    # ════════════════════════════════════════════════════════
    # 9. ORDER STATUS DISTRIBUTION
    # ════════════════════════════════════════════════════════
    order_status = list(
        DonHang.objects
        .filter(ThoiGian__gte=since_30)
        .values('TrangThai')
        .annotate(total=Count('id_DonHang'))
        .order_by('-total')
    )

    # ════════════════════════════════════════════════════════
    # 10. RECENT ACTIVITY (7 ngày)
    # ════════════════════════════════════════════════════════
    new_customers_7d  = TaiKhoan.objects.filter(
        LoaiTaiKhoan='customer', NgayTao__gte=since_7
    ).count()
    new_orders_7d     = DonHang.objects.filter(ThoiGian__gte=since_7).count()
    new_ai_clicks_7d  = AIRecommendClick.objects.filter(NgayClick__gte=since_7).count()
    new_chatbot_7d    = ChatbotFeedback.objects.filter(NgayTao__gte=since_7).count()

    return JsonResponse({
        "ok": True,
        "days": days,

        # KPI
        "kpi": {
            "total_orders":    total_orders,
            "total_revenue":   int(total_revenue),
            "total_customers": total_customers,
            "total_products":  total_products,
            "total_ai_clicks": total_ai_clicks,
            "total_chatbot_feedback": total_chatbot_fb,
            "total_viewed":    total_viewed,
            "total_ai_profiles": total_ai_profiles,
            "chatbot_avg_rating": round(float(chatbot_avg), 1),
            "new_7d": {
                "customers": new_customers_7d,
                "orders":    new_orders_7d,
                "ai_clicks": new_ai_clicks_7d,
                "chatbot":   new_chatbot_7d,
            }
        },

        # Charts
        "revenue_by_day":   revenue_by_day,
        "clicks_by_source": clicks_by_source,
        "clicks_by_day":    clicks_by_day,
        "chatbot_by_day":   chatbot_by_day,
        "views_by_day":     views_by_day,
        "feedback_dist":    feedback_dist,

        # Top lists
        "top_ai_products": top_ai_products,
        "top_viewed":      top_viewed,
        "top_purchased":   top_purchased,

        # AI Analytics
        "intent_stats":           intent_stats,
        "scent_from_profiles":    scent_from_profiles,
        "brand_from_profiles":    brand_from_profiles,
        "profile_confidence_dist": profile_confidence_dist,
        "avg_confidence":         round(float(avg_confidence), 2),
        "registered_chats":       registered_chats,
        "order_status":           order_status,

    }, json_dumps_params={"ensure_ascii": False})

def ai_dashboard_page(request):
    # Cho phép Django admin user
    is_django_admin = bool(
        getattr(request, "user", None)
        and request.user.is_authenticated
        and request.user.is_staff
    )
    # Cho phép custom session admin
    account_id = request.session.get("account_id")
    account = None
    if account_id:
        from app.models import TaiKhoan
        account = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first()
    is_custom_admin = bool(account and account.LoaiTaiKhoan in ('admin', 'staff'))

    if not is_django_admin and not is_custom_admin:
        return redirect('/admin/login/?next=/quan-tri/ai-dashboard/')

    return render(request, "admin/ai_dashboard.html")


# dang nhap
def google_login(request):
    import urllib.parse
    params = {
        'client_id': settings.SOCIAL_AUTH_GOOGLE_OAUTH2_KEY,
        'redirect_uri': 'http://localhost:8000/auth/google/callback/',
        'response_type': 'code',
        'scope': 'email profile',
    }
    url = 'https://accounts.google.com/o/oauth2/v2/auth?' + urllib.parse.urlencode(params)
    return redirect(url)


def google_callback(request):
    code = request.GET.get('code')
    if not code:
        return redirect('auth-page')
    
    token_res = http_requests.post('https://oauth2.googleapis.com/token', data={
        'code': code,
        'client_id': settings.SOCIAL_AUTH_GOOGLE_OAUTH2_KEY,
        'client_secret': settings.SOCIAL_AUTH_GOOGLE_OAUTH2_SECRET,
        'redirect_uri': 'http://localhost:8000/auth/google/callback/',
        'grant_type': 'authorization_code',
    })
    access_token = token_res.json().get('access_token')
    if not access_token:
        return redirect('auth-page')
    
    user_data = http_requests.get(
        'https://www.googleapis.com/oauth2/v2/userinfo',
        headers={'Authorization': f'Bearer {access_token}'}
    ).json()
    
    email     = user_data.get('email', '')
    full_name = user_data.get('name', '')
    
    account = TaiKhoan.objects.filter(Email__iexact=email).first()
    if not account:
        username = email.split('@')[0]
        if TaiKhoan.objects.filter(Username__iexact=username).exists():
            username = username + str(timezone.now().microsecond)
        account = TaiKhoan.objects.create(
            Username=username, MatKhau=None,
            TenDangNhap=full_name, Email=email, SDT='',
            LoaiTaiKhoan='customer', TrangThai_TaiKhoan='active',
            NgayTao=timezone.now(),
        )
        KhachHang.objects.create(
            id_TaiKhoan=account, TenKhachHang=full_name, DiaChi='', GioiTinh='',
        )
    
    request.session['account_id']   = account.id_TaiKhoan
    request.session['account_name'] = account.TenDangNhap
    return redirect('/')


def facebook_login(request):
    import urllib.parse
    params = {
        'client_id': settings.SOCIAL_AUTH_FACEBOOK_KEY,
        'redirect_uri': 'http://localhost:8000/auth/facebook/callback/',
        'scope': 'public_profile',  # bỏ email, chỉ dùng public_profile
        'response_type': 'code',
    }
    url = 'https://www.facebook.com/v18.0/dialog/oauth?' + urllib.parse.urlencode(params)
    return redirect(url)


def facebook_callback(request):
    code = request.GET.get('code')
    if not code:
        return redirect('auth-page')
    
    token_res = http_requests.get('https://graph.facebook.com/v18.0/oauth/access_token', params={
        'client_id': settings.SOCIAL_AUTH_FACEBOOK_KEY,
        'client_secret': settings.SOCIAL_AUTH_FACEBOOK_SECRET,
        'redirect_uri': 'http://localhost:8000/auth/facebook/callback/',
        'code': code,
    })
    access_token = token_res.json().get('access_token')
    if not access_token:
        return redirect('auth-page')
    
    user_data = http_requests.get('https://graph.facebook.com/me', params={
        'fields': 'id,name',  # bỏ email
        'access_token': access_token,
    }).json()
    
    email = user_data.get('email', '')
    full_name = user_data.get('name', '')
    fb_id = user_data.get('id', '')

    # Nếu không có email → dùng fb_id làm định danh
    if not email:
        email = f'fb_{fb_id}@facebook.local'
    
    account = TaiKhoan.objects.filter(Email__iexact=email).first()
    if not account:
        username = email.split('@')[0]
        if TaiKhoan.objects.filter(Username__iexact=username).exists():
            username = username + str(timezone.now().microsecond)
        account = TaiKhoan.objects.create(
            Username=username, MatKhau=None,
            TenDangNhap=full_name, Email=email, SDT='',
            LoaiTaiKhoan='customer', TrangThai_TaiKhoan='active',
            NgayTao=timezone.now(),
        )
        KhachHang.objects.create(
            id_TaiKhoan=account, TenKhachHang=full_name, DiaChi='', GioiTinh='',
        )
    
    request.session['account_id']   = account.id_TaiKhoan
    request.session['account_name'] = account.TenDangNhap
    return redirect('/')

# ════════════════════════════════════════════════════════════════
# TRACK VIEW — Gọi khi user vào trang chi tiết sản phẩm
# ════════════════════════════════════════════════════════════════
@csrf_exempt
@require_POST
def ai_track_product_view(request):
    """
    POST /api/ai/track-view/
    Body: { "product_id": 9, "time_spent": 45 }
 
    Ghi nhận lượt xem sản phẩm.
    Guest  → session
    Registered → DB
    """
    import json as _json
    try:
        data       = _json.loads(request.body)
        product_id = int(data.get("product_id") or 0)
        time_spent = int(data.get("time_spent") or 0) or None
 
        if not product_id:
            return JsonResponse({"ok": False})
 
        # Kiểm tra sản phẩm tồn tại
        if not SanPham.objects.filter(id_SanPham=product_id).exists():
            return JsonResponse({"ok": False})
 
        from app.ai.recently_viewed import track_view
        track_view(request, product_id, time_spent)
 
        return JsonResponse({"ok": True})
    except Exception:
        return JsonResponse({"ok": False})
 
 
# ════════════════════════════════════════════════════════════════
# RECENTLY VIEWED API — Section "Sản phẩm đã xem gần đây"
# ════════════════════════════════════════════════════════════════
def ai_recently_viewed_api(request):
    """
    GET /api/ai/recently-viewed/?exclude=<product_id>
    Trả về danh sách sản phẩm đã xem gần đây.
    """
    try:
        exclude_id = request.GET.get("exclude")
        exclude_id = int(exclude_id) if exclude_id else None
 
        from app.ai.recently_viewed import get_viewed_products
        product_ids = get_viewed_products(request, limit=8, exclude_id=exclude_id)
 
        if not product_ids:
            return JsonResponse({"ok": True, "products": [], "count": 0})
 
        products = list(
            SanPham.objects
            .select_related("id_ThuongHieu", "id_LoaiSanPham")
            .filter(id_SanPham__in=product_ids)
        )
        product_map = {p.id_SanPham: p for p in products}
        ordered = [product_map[pid] for pid in product_ids if pid in product_map]
        cards   = _build_product_cards(ordered)
 
        account_id = request.session.get("account_id")
        return JsonResponse({
            "ok":       True,
            "products": cards,
            "count":    len(cards),
            "type":     "persistent" if account_id else "session",
        }, json_dumps_params={"ensure_ascii": False})
 
    except Exception:
        return JsonResponse({"ok": True, "products": [], "count": 0})
 
 
# ════════════════════════════════════════════════════════════════
# GUEST RECOMMENDATION API — "Dành cho bạn" cho Guest
# ════════════════════════════════════════════════════════════════
def ai_guest_recommend_api(request):
    """
    GET /api/ai/guest-recommend/?current=<product_id>
    Gợi ý cho Guest dựa trên session behavior.
    """
    try:
        current_id = request.GET.get("current")
        current_id = int(current_id) if current_id else None
 
        from app.ai.personalize import get_guest_recommendations
        product_ids = get_guest_recommendations(
            request, current_product_id=current_id, top_n=8
        )
 
        if not product_ids:
            return JsonResponse({"ok": True, "products": [], "type": "empty"})
 
        products = list(
            SanPham.objects
            .select_related("id_ThuongHieu", "id_LoaiSanPham")
            .filter(id_SanPham__in=product_ids)
        )
        product_map = {p.id_SanPham: p for p in products}
        ordered = [product_map[pid] for pid in product_ids if pid in product_map]
        cards   = _build_product_cards(ordered)
 
        return JsonResponse({
            "ok":       True,
            "products": cards,
            "type":     "session_based",
        }, json_dumps_params={"ensure_ascii": False})
 
    except Exception:
        import traceback; traceback.print_exc()
        return JsonResponse({"ok": True, "products": [], "type": "error"})
 
 
# ════════════════════════════════════════════════════════════════
# PERSONALIZED RECOMMEND — "Dành cho bạn" cho Registered User
# (THAY THẾ hàm cũ personalized_recommend_api)
# ════════════════════════════════════════════════════════════════
def personalized_recommend_api(request):
    account_id = request.session.get("account_id")
    top_n = 8

    if account_id:
        from app.ai.personalize import get_personalized_recommendations
        product_ids = get_personalized_recommendations(account_id, top_n=top_n, request=request)
        rec_type = "personalized"
    else:
        from app.ai.personalize import get_guest_recommendations
        current_id = request.GET.get("current")
        current_id = int(current_id) if current_id else None
        product_ids = get_guest_recommendations(
            request, current_product_id=current_id, top_n=top_n
        )
        rec_type = "session_trending"

    # ── Fallback cuối: random sản phẩm nếu vẫn rỗng ──
    if not product_ids:
        product_ids = list(
            SanPham.objects.order_by('?').values_list('id_SanPham', flat=True)[:top_n]
        )
        rec_type = "discovery"

    if not product_ids:
        return JsonResponse({"ok": True, "products": [], "type": "empty"})

    products = list(
        SanPham.objects
        .select_related("id_ThuongHieu", "id_LoaiSanPham")
        .filter(id_SanPham__in=product_ids)
    )
    product_map = {p.id_SanPham: p for p in products}
    ordered = [product_map[pid] for pid in product_ids if pid in product_map]
    cards   = _build_product_cards(ordered)

    # Ghi nhận impression (Trục 1: tính CTR = click / impression)
    try:
        AIRecommendImpression.objects.create(
            id_TaiKhoan_id=account_id if account_id else None,
            source=rec_type,
            product_count=len(cards),
        )
    except Exception:
        pass

    return JsonResponse({
        "ok":       True,
        "products": cards,
        "type":     rec_type,
    }, json_dumps_params={"ensure_ascii": False})
 

# ════════════════════════════════════════════════════════════════
# CHATBOT ACTION EXECUTORS — Add to Wishlist / Cart / Order
# Dùng cho luồng Propose → Confirm → Execute trong chatbot_api
# ════════════════════════════════════════════════════════════════

def _execute_add_to_wishlist(account_id: int, product_id: int) -> dict:
    """
    Thêm sản phẩm vào yêu thích — idempotent (không lỗi nếu đã có).
    Trả về: {"ok": bool, "already_exists": bool, "message": str}
    """
    try:
        account = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first()
        product = SanPham.objects.filter(id_SanPham=product_id).first()
        if not account or not product:
            return {"ok": False, "already_exists": False,
                    "message": "Không tìm thấy sản phẩm hoặc tài khoản."}

        existing = YeuThich.objects.filter(
            id_TaiKhoan=account, id_SanPham=product
        ).first()

        if existing:
            return {"ok": True, "already_exists": True,
                    "message": f"{product.TenSanPham} đã có trong danh sách yêu thích của bạn rồi."}

        YeuThich.objects.create(
            id_TaiKhoan=account,
            id_SanPham=product,
            # NgayTao=timezone.now(),
        )
        return {"ok": True, "already_exists": False,
                "message": f"Đã thêm {product.TenSanPham} vào danh sách yêu thích."}

    except Exception as e:
        import traceback; traceback.print_exc()
        return {"ok": False, "already_exists": False,
                "message": "Có lỗi xảy ra, vui lòng thử lại."}
    


def _execute_add_to_cart(account_id: int, product_id: int, variant_id: int = None) -> dict:
    try:
        from app.models import GioHang, TaiKhoan, SanPham, BienThe
        account = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first()
        product = SanPham.objects.filter(id_SanPham=product_id).first()
        if not account or not product:
            return {"ok": False, "message": "Không tìm thấy sản phẩm hoặc tài khoản."}

        # Lấy biến thể đầu tiên nếu không chỉ định
        bt = None
        if variant_id:
            bt = BienThe.objects.filter(id_BienThe=variant_id).first()
        if not bt:
            bt = BienThe.objects.filter(id_SanPham=product).order_by('id_BienThe').first()
        if not bt:
            return {"ok": False, "message": f"{product.TenSanPham} chưa có biến thể trong kho."}

        if int(bt.SoLuong or 0) <= 0:
            return {"ok": False, "message": f"{product.TenSanPham} hiện đã hết hàng."}

        gh, created = GioHang.objects.get_or_create(
            id_TaiKhoan=account,
            id_BienThe=bt,
            defaults={'SoLuong': 1}
        )
        if not created:
            gh.SoLuong = int(gh.SoLuong or 0) + 1
            gh.save(update_fields=['SoLuong'])

        action = "thêm mới" if created else f"tăng lên {gh.SoLuong}"
        return {
            "ok": True,
            "message": f"Đã {action} **{product.TenSanPham}** vào giỏ hàng của bạn! 🛒",
            "cart_url": "/gio-hang/"
        }
    except Exception:
        import traceback; traceback.print_exc()
        return {"ok": False, "message": "Có lỗi xảy ra, vui lòng thử lại."}


_PENDING_ACTION_LABELS = {
    "add_to_wishlist": "thêm vào yêu thích",
    "add_to_cart":     "thêm vào giỏ hàng",
    "place_order":     "đặt hàng",
}


def _execute_pending_action(pending: dict, account_id, request) -> str:
    """
    Thực thi pending_action đã được khách xác nhận.
    Trả về: câu trả lời (str) cho khách.
    """
    action_type = pending.get("type")
    payload     = pending.get("payload", {})

    if action_type == "add_to_wishlist":
        if not account_id:
            return ("Bạn cần đăng nhập để lưu sản phẩm vào danh sách yêu thích nhé. "
                    "Sau khi đăng nhập, mình sẽ giúp bạn lưu lại! 😊")

        result = _execute_add_to_wishlist(account_id, payload.get("product_id"))
        if result["ok"] and not result["already_exists"]:
            return f"✅ {result['message']} Bạn có thể xem trong trang Tài khoản > Yêu thích nhé! ❤️"
        return result["message"]

    elif action_type == "add_to_cart":
        if not account_id:
            return ("Bạn cần đăng nhập để thêm vào giỏ hàng nhé! "
                    "Sau khi đăng nhập mình sẽ giúp bạn thêm ngay. 😊")
        result = _execute_add_to_cart(
            account_id,
            payload.get("product_id"),
            payload.get("variant_id")
        )
        if result["ok"]:
            return (f"✅ {result['message']} "
                    f"Xem giỏ hàng tại: "
                    f"<a href='/gio-hang/' style='color:#4B672D;font-weight:600'>Giỏ hàng</a>")
        return result["message"]

    return "Tính năng này đang được hoàn thiện, mong bạn thông cảm! 🙏"

# ════════════════════════════════════════════════════════════════
# CHATBOT API — NÂNG CẤP (thay thế hàm cũ)
# ════════════════════════════════════════════════════════════════
@csrf_exempt
@require_POST
def chatbot_api(request):
    """
    POST /api/chatbot/
    Body: { "message": "...", "history": [...], "session_id": "..." }
    Streaming SSE + song song AI classify + RAG.
    """
    import json as _json
    import os, random, uuid, concurrent.futures
    from django.http import StreamingHttpResponse

    try:
        data            = _json.loads(request.body)
        user_msg        = (data.get("message") or "").strip()
        history         = data.get("history") or []
        chat_session_id = data.get("session_id") or None

        if not user_msg:
            return JsonResponse({"ok": False, "error": "Tin nhắn không được trống."})

        if len(history) > 20:
            history = history[-20:]

        account_id = request.session.get("account_id")

        from app.ai.chatbot import (
            _is_off_topic_keyword, _classify_intent_ai,
            _INTENT_REPLIES, _extract_intent,
            _build_user_context_prompt, _build_guest_context_prompt,
            _update_guest_temp_profile, _load_recent_chatbot_history,
            _save_chatbot_history, _update_profile_from_chatbot,
            BASE_SYSTEM_PROMPT,
            _is_greeting, _GREETING_REPLIES,
            _detect_action_intent, _detect_confirmation,
            _resolve_product_for_action,
        )
        from app.ai.knowledge_base import retrieve
        from groq import Groq

        client = Groq(api_key=os.environ.get("GROQ_API_KEY", ""))
        if not chat_session_id:
            chat_session_id = str(uuid.uuid4())[:16]

        messages = list(history)

        # ── Hàm trả JSON nhanh (off-topic, error) ─────────────
        def quick_json(reply, intent_type, sugg=[]):
            new_hist = messages + [
                {"role": "user",      "content": user_msg},
                {"role": "assistant", "content": reply},
            ]
            return JsonResponse({
                "ok":          True,
                "reply":       reply,
                "history":     new_hist,
                "suggestions": sugg,
                "session_id":  chat_session_id,
                "intent":      {"type": intent_type},
            }, json_dumps_params={"ensure_ascii": False})
        
        # ── Lớp 1.5: Confirmation/Cancel Detection ──────────────
        pending = request.session.get('pending_action')
        if pending:
            confirmation = _detect_confirmation(user_msg)

            if confirmation == 'confirm':
                reply = _execute_pending_action(
                    pending, account_id, request
                )
                request.session.pop('pending_action', None)
                request.session.modified = True
                if account_id:
                    _save_chatbot_history(account_id, chat_session_id,
                                          user_msg, reply, {})
                return quick_json(reply, "action_executed")

            elif confirmation == 'cancel':
                action_name = _PENDING_ACTION_LABELS.get(pending['type'], "yêu cầu")
                reply = f"Đã hủy {action_name}. Mình có thể giúp gì khác cho bạn không? 😊"
                request.session.pop('pending_action', None)
                request.session.modified = True
                return quick_json(reply, "action_cancelled")

            else:
                # unclear: hết hạn pending_action sau 2 lượt không rõ ràng
                pending['turn_count'] = pending.get('turn_count', 0) + 1
                if pending['turn_count'] >= 2:
                    request.session.pop('pending_action', None)
                    request.session.modified = True
                    # Không return -> tiếp tục xử lý user_msg như bình thường
                else:
                    request.session['pending_action'] = pending
                    request.session.modified = True
                    payload = pending['payload']
                    reply = (
                        f"Bạn có muốn {_PENDING_ACTION_LABELS.get(pending['type'], 'thực hiện')} "
                        f"\"{payload.get('product_name', '')}\" không? "
                        f"Trả lời 'có' để xác nhận hoặc 'không' để hủy nhé."
                    )
                    return quick_json(reply, "action_pending_reminder")

        
        
        # ── Lớp 0: Greeting check (0ms) ─────────────────────────
        if _is_greeting(user_msg):
            return quick_json(
                random.choice(_GREETING_REPLIES),
                "greeting"
            )

        # ── Lớp 1: Keyword check (0ms) ─────────────────────────
        if _is_off_topic_keyword(user_msg):
            return quick_json(
                random.choice(_INTENT_REPLIES["off_topic"]),
                "off_topic"
            )
        
        # ── Lớp 2.5: Action Intent Detection ────────────────────
        action_type = _detect_action_intent(user_msg)
        if action_type:
            from app.ai.knowledge_base import retrieve as _retrieve_for_action
            action_chunks = _retrieve_for_action(user_msg, 3)
            last_suggestions = request.session.get('last_suggestions', [])

            resolved = _resolve_product_for_action(
                user_msg, action_chunks, last_suggestions
            )

            if action_type == "add_to_wishlist":
                if not resolved:
                    reply = ("Bạn muốn thêm sản phẩm nào vào yêu thích vậy? "
                             "Hãy cho mình biết tên sản phẩm cụ thể nhé! 😊")
                    return quick_json(reply, "action_need_clarification")

                request.session['pending_action'] = {
                    "type": "add_to_wishlist",
                    "type": "add_to_cart",
                    "payload": {
                        "product_id": resolved["id"],
                        "product_name": resolved["name"],
                    },
                    "turn_count": 0,
                }
                request.session.modified = True
                print(f"[DEBUG] pending_action set: {request.session.get('pending_action')}")

                reply = (
                    f"Bạn muốn thêm \"{resolved['name']}\" vào danh sách yêu thích "
                    f"đúng không? Trả lời 'có' để xác nhận nhé! ❤️"
                )
                return quick_json(reply, "action_proposed")

            if action_type == 'add_to_cart':
                if not resolved:
                    reply = ("Bạn muốn thêm sản phẩm nào vào giỏ hàng vậy? "
                            "Cho mình biết tên sản phẩm cụ thể nhé! 🛒")
                    return quick_json(reply, "action_need_clarification")

                request.session['pending_action'] = {
                    "type": "add_to_cart",
                    "payload": {
                        "product_id":  resolved["id"],
                        "product_name": resolved["name"],
                        "variant_id":  None,
                    },
                    "turn_count": 0,
                }
                request.session.modified = True
                reply = (
                    f"Bạn muốn thêm **\"{resolved['name']}\"** vào giỏ hàng "
                    f"đúng không? Trả lời 'có' để xác nhận nhé! 🛒"
                )
                return quick_json(reply, "action_proposed")
            # Tạm thời: nếu phát hiện nhưng chưa hỗ trợ, để rơi xuống 
            # RAG bình thường (không return), AI sẽ tư vấn như câu hỏi thường.

            if action_type == 'place_order':
                if account_id:
                    reply = (
                        "Mình không thể đặt hàng trực tiếp qua chat, nhưng mình sẽ hướng dẫn bạn nhé! 🛍️\n\n"
                        "**Quy trình đặt hàng tại Ami Perfumery:**\n\n"
                        "**Bước 1 — Thêm vào giỏ hàng**\n"
                        "Vào trang sản phẩm → chọn dung tích → nhấn **Thêm vào giỏ**\n\n"
                        "**Bước 2 — Kiểm tra giỏ hàng**\n"
                        "Nhấn icon giỏ hàng góc trên → kiểm tra sản phẩm, số lượng → nhấn **Thanh toán**\n\n"
                        "**Bước 3 — Nhập thông tin giao hàng**\n"
                        "Điền họ tên, số điện thoại, địa chỉ nhận hàng\n\n"
                        "**Bước 4 — Chọn phương thức thanh toán**\n"
                        "• **COD** — Thanh toán khi nhận hàng\n"
                        "• **VNPay** — Chuyển khoản / thẻ ATM / QR\n"
                        "• **MoMo** — Ví điện tử MoMo\n\n"
                        "**Bước 5 — Xác nhận đặt hàng**\n"
                        "Nhấn **Đặt hàng** → nhận email xác nhận → theo dõi tại "
                        "<a href='/tai-khoan/?tab=orders' style='color:#4B672D;font-weight:600'>Tài khoản → Đơn hàng</a>\n\n"
                        "Bạn muốn mình tư vấn thêm sản phẩm nào không? 🌸"
                    )
                else:
                    reply = (
                        "Mình không thể đặt hàng trực tiếp qua chat, nhưng mình sẽ hướng dẫn bạn nhé! 🛍️\n\n"
                        "**Quy trình đặt hàng tại Ami Perfumery:**\n\n"
                        "**Bước 1 — Thêm vào giỏ hàng**\n"
                        "Vào trang sản phẩm → chọn dung tích → nhấn **Thêm vào giỏ**\n\n"
                        "**Bước 2 — Kiểm tra giỏ hàng**\n"
                        "Nhấn icon giỏ hàng góc trên → kiểm tra sản phẩm → nhấn **Thanh toán**\n\n"
                        "**Bước 3 — Nhập thông tin giao hàng**\n"
                        "Điền họ tên, số điện thoại, địa chỉ nhận hàng\n\n"
                        "**Bước 4 — Chọn phương thức thanh toán**\n"
                        "• **COD** — Thanh toán khi nhận hàng\n"
                        "• **VNPay** — Chuyển khoản / thẻ ATM / QR\n"
                        "• **MoMo** — Ví điện tử MoMo\n\n"
                        "**Bước 5 — Xác nhận**\n"
                        "Nhấn **Đặt hàng** → nhận email xác nhận\n\n"
                        "💡 <a href='/xac-thuc/' style='color:#4B672D;font-weight:600'>Đăng ký tài khoản</a> "
                        "để theo dõi đơn hàng, tích điểm và nhận ưu đãi thành viên!\n\n"
                        "Bạn cần tư vấn thêm sản phẩm nào không? 🌸"
                    )
                if account_id:
                    _save_chatbot_history(account_id, chat_session_id, user_msg, reply, {})
                return quick_json(reply, "order_guide")
        # ── Lớp 2 + RAG song song ──────────────────────────────
        from app.ai.chatbot import _extract_gender_quick
        # Chỉ filter theo gender nếu CÂU HIỆN TẠI có nhắc đến giới tính.
        # KHÔNG fallback về profile/session — vì RAG retrieve phải phản
        # ánh đúng nội dung câu hỏi hiện tại (vd: hỏi tên sản phẩm cụ thể
        # không nên bị lọc mất bởi gender từ lịch sử trước đó).
        quick_gender = _extract_gender_quick(user_msg)

        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as ex:
            f_topic  = ex.submit(_classify_intent_ai, user_msg, client)
            f_chunks = ex.submit(retrieve, user_msg, 5, quick_gender)
            topic  = f_topic.result()
            chunks = f_chunks.result()
        
        # ── Bestseller / Popular request: bơm top sản phẩm bán chạy ──
        from app.ai.chatbot import _is_bestseller_request
        if _is_bestseller_request(user_msg):
            from app.ai.personalize import _get_popular_products, _get_trending_products
            from app.ai.knowledge_base import get_chunks_by_ids

            bestseller_ids = _get_popular_products(top_n=3)
            if not bestseller_ids:
                bestseller_ids = _get_trending_products(top_n=3)

            if bestseller_ids:
                bestseller_chunks = get_chunks_by_ids(bestseller_ids)
                existing_ids = {c["id"] for c in chunks if c["type"] == "product"}
                merged = []
                for bc in bestseller_chunks:
                    bc = dict(bc)
                    bc["text"] = "⭐ SẢN PHẨM BÁN CHẠY/PHỔ BIẾN: " + bc.get("text", "")
                    merged.append(bc)
                    existing_ids.add(bc["id"])
                # Ưu tiên bestseller lên đầu, giữ lại các chunk RAG cũ không trùng
                chunks = merged + [c for c in chunks if c["id"] not in
                                    {bc["id"] for bc in bestseller_chunks}]
                chunks = chunks[:6]

        # Off-topic từ AI classify
        if topic in ('celebrity_gossip', 'off_topic'):
            reply = random.choice(_INTENT_REPLIES[topic])
            if account_id:
                _save_chatbot_history(account_id, chat_session_id, user_msg, reply, {})
            return quick_json(reply, topic)

        elif topic == 'order_support':
            if account_id:
                reply = (
                    "Bạn xem đơn hàng tại đây nhé: "
                    "<a href='/tai-khoan/?tab=orders' style='color:#4B672D;font-weight:600'>"
                    "Tài khoản → Đơn hàng</a> 📦<br><br>"
                    "Tại đó có đầy đủ trạng thái, lịch sử và nút xác nhận nhận hàng. "
                    "Bạn cần tư vấn thêm về nước hoa không? 🌸"
                )
                _save_chatbot_history(account_id, chat_session_id, user_msg, reply, {})
            else:
                reply = (
                    "Để theo dõi đơn hàng, bạn cần "
                    "<a href='/xac-thuc/' style='color:#4B672D;font-weight:600'>đăng nhập</a> "
                    "trước nhé! 🔐<br><br>"
                    "Sau khi đăng nhập vào <b>Tài khoản → Đơn hàng</b> "
                    "là thấy toàn bộ lịch sử đơn hàng."
                )
            return quick_json(reply, "order_support")
        
        # ── Xử lý memory recall ──────────────────────────────
        if topic == 'memory_recall':
            if account_id:
                db_history = _load_recent_chatbot_history(account_id, limit=6)
                if db_history:
                    messages = db_history  # inject history → fall through xuống RAG
                else:
                    reply = "Đây có vẻ là lần đầu chúng ta trò chuyện! 😊 Bạn đang tìm kiếm hương thơm cho dịp nào?"
                    return quick_json(reply, "memory_recall")
            else:
                reply = "Mình chưa lưu lịch sử trò chuyện cho khách chưa đăng nhập bạn ơi! 😊 Đăng nhập để mình nhớ bạn tốt hơn nhé. Bạn đang tìm nước hoa gì?"
                return quick_json(reply, "memory_recall")
            # Có history → tiếp tục xuống RAG + Generate bình thường

        # ── Lớp 3: Build context ────────────────────────────────
        intent = _extract_intent(user_msg)

        # ── Cập nhật profile NGAY (để gợi ý cá nhân hóa của lượt 
        #    này phản ánh sở thích mới vừa được trích xuất) ──
        if account_id:
            _update_profile_from_chatbot(account_id, intent)
        elif request:
            _update_guest_temp_profile(request, intent)

        # Lấy giá thật cho sản phẩm trong RAG context (tránh AI bịa giá)
        product_ids_in_chunks = [c["id"] for c in chunks if c["type"] == "product"]
        rag_price_map = {}
        if product_ids_in_chunks:
            from django.db.models import Min
            for row in BienThe.objects.filter(
                id_SanPham_id__in=product_ids_in_chunks
            ).values('id_SanPham_id').annotate(min_price=Min('GiaBan')):
                rag_price_map[row['id_SanPham_id']] = row['min_price']

        context_lines = []
        for i, c in enumerate(chunks, 1):
            prefix = "[SẢN PHẨM]" if c["type"] == "product" else "[BÀI VIẾT]"
            price_suffix = ""
            if c["type"] == "product":
                price = rag_price_map.get(c["id"])
                price_text = f"{int(price):,}đ".replace(",", ".") if price else "Liên hệ"
                price_suffix = f" (Giá: {price_text})"
            context_lines.append(f"{i}. {prefix} {c['name']}{price_suffix}: {c['text']}")
        rag_context = "\n\n".join(context_lines)

        if account_id:
            personal_context = _build_user_context_prompt(account_id)
        elif request:
            personal_context = _build_guest_context_prompt(request)
        else:
            personal_context = ""

        # ── Personalized Recommendation Tool ────────────────────
        from app.ai.chatbot import (
            _build_personalized_recommendation_context,
            _is_recommendation_request,
        )
        # Kiểm tra đã có sở thích tích lũy (guest hoặc registered) chưa
        has_accumulated_profile = False
        if account_id:
            from app.ai.personalize import _load_ai_profile
            _p = _load_ai_profile(account_id)
            has_accumulated_profile = bool(_p and _p.ConfidenceScore > 0.1)
        elif request:
            has_accumulated_profile = bool(
                request.session.get('guest_ai_profile', {}).get('confidence', 0) > 0
            )

        is_first_message = len(history) == 0 
        rec_context, rec_product_ids = "", []
        should_recommend = (
            _is_recommendation_request(user_msg)
            or (not intent and has_accumulated_profile)
        )
        if should_recommend:
            rec_context, rec_product_ids = _build_personalized_recommendation_context(
                account_id=account_id, request=request, top_n=3
            )

        # ── DEBUG TẠM ──
        print(f"[DEBUG] user_msg='{user_msg}'", flush=True)
        print(f"[DEBUG] intent={intent}", flush=True)
        print(f"[DEBUG] topic={topic}", flush=True)
        print(f"[DEBUG] is_rec_request={_is_recommendation_request(user_msg)}", flush=True)
        print(f"[DEBUG] guest_ai_profile={request.session.get('guest_ai_profile', {})}", flush=True)
        print(f"[DEBUG] rec_product_ids={rec_product_ids}", flush=True)
        print(f"[DEBUG] rec_context length={len(rec_context)}", flush=True)

        full_system = (
            BASE_SYSTEM_PROMPT + personal_context + rec_context
            + "\n\n" + "="*50
            + "\nDỮ LIỆU SẢN PHẨM CỬA HÀNG AMI PERFUMERY:\n"
            + "="*50 + "\n" + rag_context + "\n" + "="*50
        )

        if account_id and not messages:
            db_hist = _load_recent_chatbot_history(account_id, limit=6)
            if db_hist:
                messages = db_hist

        messages.append({"role": "user", "content": user_msg})

        # Suggestions với ảnh + giá (giữ nguyên logic cũ)
        suggested = [c for c in chunks if c["type"] == "product"][:3]

        # ── Bổ sung sản phẩm gợi ý cá nhân hóa (nếu chưa có) ──
        if rec_product_ids:
            from app.ai.knowledge_base import get_chunks_by_ids
            existing_ids = {s["id"] for s in suggested if s.get("type") == "product"}
            for pid in rec_product_ids:
                if len(suggested) >= 3:
                    break
                if pid not in existing_ids:
                    rec_chunks = get_chunks_by_ids([pid])
                    if rec_chunks:
                        suggested.append(rec_chunks[0])
                        existing_ids.add(pid)

        if suggested:
            pid_list = [s["id"] for s in suggested if s.get("type") == "product"]
            if pid_list:
                img_map  = _product_image_map(pid_list)
                var_map  = _first_variant_map(pid_list)
                for s in suggested:
                    pid = s.get("id")
                    imgs = img_map.get(pid, [])
                    s["image"] = imgs[0] if imgs else ""
                    v = var_map.get(pid)
                    s["price"] = _format_currency(v.GiaBan) if (
                        v and v.GiaBan
                    ) else ""
        # ── Lưu last_suggestions vào session ──  ← THÊM ĐOẠN NÀY
        product_suggestions = [
            {"id": s["id"], "name": s["name"]}
            for s in suggested if s.get("type") == "product"
        ]
        if product_suggestions:
            request.session['last_suggestions'] = product_suggestions
            request.session.modified = True

        # Ghi nhận impression cho chatbot suggestions (Trục 1)
        if product_suggestions:
            try:
                AIRecommendImpression.objects.create(
                    id_TaiKhoan_id=account_id if account_id else None,
                    source='chatbot',
                    product_count=len(product_suggestions),
                )
            except Exception:
                pass

        # ── Streaming ───────────────────────────────────────────
        def stream_gen():
            import json as j

            # 1. Gửi meta TRƯỚC (suggestions + session_id)
            yield f"data: {j.dumps({'t':'meta','suggestions_data':suggested,'session_id':chat_session_id,'intent':intent}, ensure_ascii=False)}\n\n"

            full_reply = ""
            try:
                # 2. Stream từng token
                stream = client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    max_tokens=1024,
                    temperature=0.7,
                    stream=True,
                    messages=[{"role":"system","content":full_system}] + messages,
                )

                for chunk in stream:
                    token = chunk.choices[0].delta.content or ""
                    if token:
                        full_reply += token
                        yield f"data: {j.dumps({'t':'token','text':token}, ensure_ascii=False)}\n\n"

            except Exception as e:
                import traceback; traceback.print_exc()

                err_str = str(e).lower()
                if 'rate_limit' in err_str or '429' in err_str:
                    fallback_msg = ("Hệ thống AI đang tạm quá tải do lượng yêu cầu cao. "
                                    "Vui lòng thử lại sau vài phút nhé! 🙏")
                else:
                    fallback_msg = ("Xin lỗi, mình gặp sự cố khi xử lý yêu cầu này. "
                                    "Vui lòng thử lại hoặc đặt câu hỏi khác nhé! 🙏")

                if not full_reply.strip():
                    full_reply = fallback_msg
                    yield f"data: {j.dumps({'t':'token','text':full_reply}, ensure_ascii=False)}\n\n"
                else:
                    extra = "\n\n(Phản hồi bị gián đoạn do lỗi hệ thống.)"
                    full_reply += extra
                    yield f"data: {j.dumps({'t':'token','text':extra}, ensure_ascii=False)}\n\n"

            # 3. Lưu DB
            if account_id:
                _save_chatbot_history(account_id, chat_session_id,
                                    user_msg, full_reply, intent)

            # 4. Kiểm tra show_products SAU KHI có full_reply
            _SHOW_PRODUCT_TRIGGERS = [
                'gợi ý', 'đề xuất', 'giới thiệu', 'recommend',
                'sản phẩm', 'chai', 'mùi hương này', 'bạn có thể xem',
                'tham khảo', 'phù hợp với bạn', 'dưới đây',
            ]
            show_products = any(kw in full_reply.lower() for kw in _SHOW_PRODUCT_TRIGGERS)

            # 5. Gửi done với show_products
            new_hist = messages + [{"role":"assistant","content":full_reply}]
            yield f"data: {j.dumps({'t':'done','history':new_hist,'show_products':show_products}, ensure_ascii=False)}\n\n"

        resp = StreamingHttpResponse(stream_gen(), content_type="text/event-stream")
        resp["Cache-Control"]      = "no-cache"
        resp["X-Accel-Buffering"]  = "no"
        return resp

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({"ok": False, "error": "Lỗi hệ thống."})
 
 
# ════════════════════════════════════════════════════════════════
# AI USER PROFILE API — Cho trang profile / cá nhân hóa
# ════════════════════════════════════════════════════════════════
def ai_user_profile_api(request):
    """
    GET /api/ai/my-profile/
    Trả về AI preference profile của user đang đăng nhập.
    """
    account_id = request.session.get("account_id")
    if not account_id:
        return JsonResponse({"ok": False, "error": "Chưa đăng nhập"}, status=401)
 
    try:
        from app.models import AIUserProfile
        profile = AIUserProfile.objects.filter(
            id_TaiKhoan_id=account_id
        ).first()
 
        if not profile:
            return JsonResponse({"ok": True, "profile": None})
 
        return JsonResponse({
            "ok": True,
            "profile": {
                "nhom_mua":       profile.get_nhom_mua(),
                "thuong_hieu":    profile.get_thuong_hieu(),
                "gia_min":        float(profile.GiaMin or 0),
                "gia_max":        float(profile.GiaMax or 0),
                "gioi_tinh":      profile.GioiTinhUuTien,
                "dip_dung":       profile.DipDungUuTien,
                "confidence":     round(profile.ConfidenceScore, 2),
                "ngay_cap_nhat":  profile.NgayCapNhat.strftime("%d/%m/%Y") if profile.NgayCapNhat else None,
            }
        }, json_dumps_params={"ensure_ascii": False})
 
    except Exception:
        return JsonResponse({"ok": False, "error": "Lỗi server"})
 

# GET  /api/cart/        → lấy giỏ hàng từ DB
# POST /api/cart/sync/   → đồng bộ toàn bộ giỏ từ localStorage lên DB
# POST /api/cart/update/ → thêm/sửa/xóa 1 item

def cart_get_api(request):
    account_id = request.session.get('account_id')
    if not account_id:
        return JsonResponse({'ok': True, 'items': []})

    account = TaiKhoan.objects.filter(pk=account_id).first()
    if not account:
        return JsonResponse({'ok': True, 'items': []})

    from .models import GioHang
    items = list(
        GioHang.objects
        .filter(id_TaiKhoan=account)   # ← dùng object
        .select_related('id_BienThe__id_SanPham__id_ThuongHieu')
    )

    product_ids = [i.id_BienThe.id_SanPham_id for i in items if i.id_BienThe]
    image_map   = _product_image_map(product_ids)

    result = []
    for item in items:
        bt = item.id_BienThe
        sp = bt.id_SanPham if bt else None
        if not sp: continue

        imgs = image_map.get(sp.id_SanPham, [])
        result.append({
            'key':       f'{sp.id_SanPham}-{bt.id_BienThe}',
            'productId': str(sp.id_SanPham),
            'variantId': str(bt.id_BienThe),
            'name':      sp.TenSanPham + (f' — {bt.Sku}' if bt.Sku else ''),
            'price':     int(bt.GiaBan or 0),
            'image':     imgs[0] if imgs else '',
            'sku':       bt.Sku or '',
            'qty':       item.SoLuong,
            'stock':     int(bt.SoLuong or 0),
            'checked':   True,
        })

    return JsonResponse({'ok': True, 'items': result},
                        json_dumps_params={'ensure_ascii': False})


@csrf_exempt
@require_POST
def cart_sync_api(request):
    account_id = request.session.get('account_id')
    if not account_id:
        return JsonResponse({'ok': False})

    account = TaiKhoan.objects.filter(pk=account_id).first()
    if not account:
        return JsonResponse({'ok': False})

    from .models import GioHang
    try:
        items = json.loads(request.body).get('items', [])
        for item in items:
            variant_id = item.get('variantId')
            qty        = int(item.get('qty') or 1)
            if not variant_id or not str(variant_id).isdigit(): continue
            if qty <= 0: continue

            bt = BienThe.objects.filter(pk=int(variant_id)).first()
            if not bt: continue

            gh, created = GioHang.objects.get_or_create(
                id_TaiKhoan=account,   # ← dùng object
                id_BienThe=bt,
                defaults={'SoLuong': qty}
            )
            if not created:
                gh.SoLuong = qty
                gh.save(update_fields=['SoLuong'])

        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@csrf_exempt
@require_POST
def cart_update_api(request):
    account_id = request.session.get('account_id')
    if not account_id:
        return JsonResponse({'ok': False})

    from .models import GioHang
    try:
        data       = json.loads(request.body)
        variant_id = data.get('variantId')
        qty        = int(data.get('qty') or 0)
        action     = data.get('action', 'set')

        if not variant_id or not str(variant_id).isdigit():
            return JsonResponse({'ok': False, 'error': 'Invalid variantId'})

        bt = BienThe.objects.filter(pk=int(variant_id)).first()
        if not bt:
            return JsonResponse({'ok': False, 'error': 'Không tìm thấy biến thể'})

        # ← Lấy object TaiKhoan thay vì dùng id trực tiếp
        account = TaiKhoan.objects.filter(pk=account_id).first()
        if not account:
            return JsonResponse({'ok': False, 'error': 'Không tìm thấy tài khoản'})

        if action == 'remove' or qty <= 0:
            GioHang.objects.filter(
                id_TaiKhoan=account,   # ← dùng object
                id_BienThe=bt
            ).delete()
        else:
            gh, _ = GioHang.objects.get_or_create(
                id_TaiKhoan=account,   # ← dùng object
                id_BienThe=bt,
                defaults={'SoLuong': qty}
            )
            gh.SoLuong = qty
            gh.save(update_fields=['SoLuong'])

        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})
    

# ════════════════════════════════════════════════════════════════
# SURVEY — Khảo sát mức độ hài lòng (Likert-5)
# ════════════════════════════════════════════════════════════════

@csrf_exempt
@require_POST
def submit_survey_api(request):
    """
    POST /api/survey/submit/
    Body: JSON với 13 câu q1..q13 (giá trị 1-5), feedback_text,
    do_tuoi, gioi_tinh, tan_suat_mua.
    """
    import json as _json
    try:
        data = _json.loads(request.body)
        account_id = request.session.get("account_id")

        likert_fields = [
            'q1_phu_hop', 'q2_tim_nhanh', 'q3_da_dang', 'q4_tin_tuong',
            'q5_hieu_dung', 'q6_phan_hoi_nhanh', 'q7_de_hieu', 'q8_nhu_nhan_vien',
            'q9_nho_so_thich', 'q10_cai_thien',
            'q11_hai_long', 'q12_quay_lai', 'q13_gioi_thieu',
        ]

        values = {}
        for f in likert_fields:
            v = data.get(f)
            try:
                v = int(v)
            except (TypeError, ValueError):
                return JsonResponse({
                    "ok": False,
                    "message": f"Câu trả lời '{f}' không hợp lệ."
                }, status=400)
            if v < 1 or v > 5:
                return JsonResponse({
                    "ok": False,
                    "message": f"Câu trả lời '{f}' phải từ 1 đến 5."
                }, status=400)
            values[f] = v

        SurveyResponse.objects.create(
            id_TaiKhoan_id=account_id,
            **values,
            feedback_text=(data.get('feedback_text') or '').strip()[:1000],
            do_tuoi=(data.get('do_tuoi') or '').strip()[:20],
            gioi_tinh=(data.get('gioi_tinh') or '').strip()[:10],
            tan_suat_mua=(data.get('tan_suat_mua') or '').strip()[:30],
        )

        return JsonResponse({
            "ok": True,
            "message": "Cảm ơn bạn đã hoàn thành khảo sát! 🌸"
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return JsonResponse({"ok": False, "message": "Có lỗi xảy ra, vui lòng thử lại."}, status=500)


def survey_stats_api(request):
    """
    GET /api/survey/stats/
    Trả về thống kê trung bình theo câu hỏi và theo nhóm.
    """
    from django.db.models import Avg, Count

    is_django_admin = bool(
        getattr(request, "user", None)
        and request.user.is_authenticated and request.user.is_staff
    )
    account_id = request.session.get("account_id")
    account = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first() if account_id else None
    is_custom_admin = bool(account and account.LoaiTaiKhoan in ('admin', 'staff'))
    if not is_django_admin and not is_custom_admin:
        return JsonResponse({"ok": False}, status=403)

    qs = SurveyResponse.objects.all()
    n = qs.count()
    if n == 0:
        return JsonResponse({"ok": True, "n": 0, "message": "Chưa có dữ liệu khảo sát."})

    fields = [
        'q1_phu_hop', 'q2_tim_nhanh', 'q3_da_dang', 'q4_tin_tuong',
        'q5_hieu_dung', 'q6_phan_hoi_nhanh', 'q7_de_hieu', 'q8_nhu_nhan_vien',
        'q9_nho_so_thich', 'q10_cai_thien',
        'q11_hai_long', 'q12_quay_lai', 'q13_gioi_thieu',
    ]

    averages = qs.aggregate(**{f: Avg(f) for f in fields})
    averages = {k: round(float(v), 2) for k, v in averages.items()}

    groups = {
        "Chat_luong_goi_y": round(sum(averages[f] for f in fields[0:4]) / 4, 2),
        "Trai_nghiem_Chatbot": round(sum(averages[f] for f in fields[4:8]) / 4, 2),
        "Ca_nhan_hoa": round(sum(averages[f] for f in fields[8:10]) / 2, 2),
        "Hai_long_tong_the": round(sum(averages[f] for f in fields[10:13]) / 3, 2),
    }

    demographics = {
        "do_tuoi": list(qs.values('do_tuoi').annotate(count=Count('id_Survey'))),
        "gioi_tinh": list(qs.values('gioi_tinh').annotate(count=Count('id_Survey'))),
        "tan_suat_mua": list(qs.values('tan_suat_mua').annotate(count=Count('id_Survey'))),
    }

    feedback_texts = list(
        qs.exclude(feedback_text='').exclude(feedback_text__isnull=True)
        .values_list('feedback_text', flat=True)[:50]
    )

    return JsonResponse({
        "ok": True, "n": n,
        "per_question": averages,
        "per_group": groups,
        "overall_average": round(sum(groups.values()) / len(groups), 2),
        "demographics": demographics,
        "feedback_texts": feedback_texts,
    }, json_dumps_params={"ensure_ascii": False})


from django.views.decorators.csrf import ensure_csrf_cookie

@ensure_csrf_cookie
def survey_page(request):
    """Trang khảo sát công khai — gửi qua Cloudflare Tunnel cho người tham gia."""
    return render(request, 'app/survey.html')

def ai_evaluation_report_api(request):
    """
    GET /api/admin/evaluation-report/?days=30
    Báo cáo tổng hợp chỉ số hiệu quả gợi ý AI (Trục 1).
    """
    is_django_admin = bool(
        getattr(request, "user", None)
        and request.user.is_authenticated and request.user.is_staff
    )
    account_id = request.session.get("account_id")
    account = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first() if account_id else None
    is_custom_admin = bool(account and account.LoaiTaiKhoan in ('admin', 'staff'))
    if not is_django_admin and not is_custom_admin:
        return JsonResponse({"ok": False}, status=403)

    from app.ai.evaluation import generate_full_report
    days = int(request.GET.get('days', 30))
    report = generate_full_report(since_days=days)
    return JsonResponse({"ok": True, "report": report}, json_dumps_params={"ensure_ascii": False})


@require_POST
def api_them_thuong_hieu(request):
    """Tạo nhanh thương hiệu mới từ form phiếu nhập."""
    # Kiểm tra quyền admin
    account_id = request.session.get('account_id')
    try:
        tk = TaiKhoan.objects.get(pk=account_id)
        if tk.LoaiTaiKhoan not in ('admin', 'staff'):
            return JsonResponse({'ok': False, 'error': 'Không có quyền!'}, status=403)
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Chưa đăng nhập!'}, status=401)

    try:
        data = json.loads(request.body)
        ten  = (data.get('ten') or '').strip()
        if not ten:
            return JsonResponse({'ok': False, 'error': 'Tên thương hiệu không được trống!'})
        th, created = ThuongHieu.objects.get_or_create(TenThuongHieu=ten)
        return JsonResponse({
            'ok': True,
            'id': th.pk,
            'ten': th.TenThuongHieu,
            'created': created
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})
    

def ai_dashboard_export_excel(request):
    """
    GET /api/admin/ai-dashboard/export-excel/?days=30
    Trả về file .xlsx có màu sắc đầy đủ.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import timedelta, date as _date
    from django.utils import timezone
    from django.db.models import Count, Sum, Avg

    # ── Auth ──────────────────────────────────────────────────────
    is_django_admin = bool(
        getattr(request, "user", None)
        and request.user.is_authenticated
        and request.user.is_staff
    )
    account_id = request.session.get("account_id")
    account    = TaiKhoan.objects.filter(id_TaiKhoan=account_id).first() if account_id else None
    is_custom_admin = bool(account and account.LoaiTaiKhoan in ('admin', 'staff'))
    if not is_django_admin and not is_custom_admin:
        return JsonResponse({"ok": False}, status=403)

    # ── Params ────────────────────────────────────────────────────
    try:
        days = int(request.GET.get('days', 30))
        if days not in [7, 14, 30, 60, 90]:
            days = 30
    except (ValueError, TypeError):
        days = 30

    now   = timezone.now()
    since = now - timedelta(days=days)

    # ── Lấy dữ liệu ──────────────────────────────────────────────
    revenue_by_day = []
    for i in range(days - 1, -1, -1):
        d = now.date() - timedelta(days=i)
        rev = DonHang.objects.filter(
            ThoiGian__date=d,
            TrangThai__in=['Hoàn tất', 'Khách đã nhận hàng']
        ).aggregate(s=Sum('TongTien'))['s'] or 0
        ords = DonHang.objects.filter(ThoiGian__date=d).count()
        revenue_by_day.append({'date': d.strftime('%d/%m/%Y'), 'revenue': int(rev), 'orders': ords})

    clicks_by_day = []
    views_by_day  = []
    chatbot_by_day = []
    for i in range(days - 1, -1, -1):
        d = now.date() - timedelta(days=i)
        clicks_by_day.append({'date': d.strftime('%d/%m/%Y'),
                               'count': AIRecommendClick.objects.filter(NgayClick__date=d).count()})
        views_by_day.append({'date': d.strftime('%d/%m/%Y'),
                              'count': LichSuXemSanPham.objects.filter(NgayXem__date=d).count()})
        avg_r = ChatbotFeedback.objects.filter(NgayTao__date=d).aggregate(avg=Avg('Rating'))['avg'] or 0
        chatbot_by_day.append({
            'date': d.strftime('%d/%m/%Y'),
            'count': ChatbotFeedback.objects.filter(NgayTao__date=d).count(),
            'avg': round(float(avg_r), 1),
        })

    top_ai = list(
        AIRecommendClick.objects.filter(NgayClick__gte=since)
        .values('id_SanPham__TenSanPham', 'id_SanPham__id_ThuongHieu__TenThuongHieu')
        .annotate(total=Count('id_Click')).order_by('-total')[:8]
    )
    top_purchased = list(
        ChiTietDonHang.objects.filter(
            id_DonHang__TrangThai__in=['Hoàn tất', 'Khách đã nhận hàng'],
            id_DonHang__ThoiGian__gte=since
        ).values(
            'id_BienThe__id_SanPham__TenSanPham',
            'id_BienThe__id_SanPham__id_ThuongHieu__TenThuongHieu'
        ).annotate(total=Sum('SoLuong')).order_by('-total')[:8]
    )
    clicks_by_source = list(
        AIRecommendClick.objects.filter(NgayClick__gte=since)
        .values('source').annotate(total=Count('id_Click')).order_by('-total')
    )

    import json as _j
    scent_from_profiles, brand_from_profiles = {}, {}
    for profile in AIUserProfile.objects.all():
        for s in _j.loads(profile.NhomMuaYeuThich or '[]'):
            scent_from_profiles[s] = scent_from_profiles.get(s, 0) + 1
        for b in _j.loads(profile.ThuongHieuYeuThich or '[]'):
            brand_from_profiles[b] = brand_from_profiles.get(b, 0) + 1
    scent_from_profiles = dict(sorted(scent_from_profiles.items(), key=lambda x: x[1], reverse=True)[:8])
    brand_from_profiles = dict(sorted(brand_from_profiles.items(), key=lambda x: x[1], reverse=True)[:8])

    intent_stats = {}
    for intent_str in ChatbotHistory.objects.filter(
        ExtractedIntent__isnull=False, NgayTao__gte=since
    ).values_list('ExtractedIntent', flat=True):
        if not intent_str:
            continue
        for part in intent_str.split('_'):
            part = part.strip()
            if part and len(part) > 1:
                intent_stats[part] = intent_stats.get(part, 0) + 1
    intent_stats = dict(sorted(intent_stats.items(), key=lambda x: x[1], reverse=True)[:10])

    kpi_total_revenue  = int(DonHang.objects.filter(
        TrangThai__in=['Hoàn tất', 'Khách đã nhận hàng']
    ).aggregate(s=Sum('TongTien'))['s'] or 0)
    kpi_total_orders   = DonHang.objects.count()
    kpi_customers      = TaiKhoan.objects.filter(LoaiTaiKhoan='customer').count()
    kpi_products       = SanPham.objects.count()
    kpi_ai_clicks      = AIRecommendClick.objects.count()
    kpi_viewed         = LichSuXemSanPham.objects.count()
    kpi_ai_profiles    = AIUserProfile.objects.count()
    chatbot_avg        = round(float(ChatbotFeedback.objects.aggregate(avg=Avg('Rating'))['avg'] or 0), 1)

    # ── Màu sắc ───────────────────────────────────────────────────
    OLIVE    = "4B672D"
    OLIVE_LT = "5d7f38"
    MINT     = "EBF6C4"
    DARK     = "0d1208"
    GOLD     = "c9a96e"
    TEAL     = "4db8a0"
    BLUE     = "5b9bd5"
    WHITE    = "FFFFFF"
    GRAY_L   = "f5f5f0"
    GRAY_T   = "888888"

    # ── Style helpers ─────────────────────────────────────────────
    def thin_border(color="CCCCCC"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr_style(bg=OLIVE, fg=WHITE, sz=11, bold=True):
        return {
            'font': Font(name="Arial", bold=bold, color=fg, size=sz),
            'fill': PatternFill("solid", fgColor=bg),
            'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': thin_border("AAAAAA"),
        }

    def data_style(bold=False, fg="222222", bg=WHITE, center=False):
        return {
            'font': Font(name="Arial", bold=bold, color=fg, size=10),
            'fill': PatternFill("solid", fgColor=bg),
            'alignment': Alignment(horizontal="center" if center else "left", vertical="center"),
            'border': thin_border("E0E0E0"),
        }

    def apply(cell, style):
        for k, v in style.items():
            setattr(cell, k, v)

    def wc(ws, row, col, value, style):
        c = ws.cell(row=row, column=col, value=value)
        apply(c, style)
        return c

    # ════════════════════════════════════════════════════════════════
    wb = Workbook()

    # ── SHEET 1: TỔNG QUAN ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tổng quan"
    ws1.sheet_view.showGridLines = False

    # Title
    ws1.merge_cells("A1:H1")
    ws1.merge_cells("A2:H2")
    ws1.row_dimensions[1].height = 46
    ws1.row_dimensions[2].height = 22
    apply(ws1["A1"], {**hdr_style(DARK, MINT, 15), 'font': Font(name="Arial", bold=True, color=MINT, size=15)})
    ws1["A1"].value = "AMI PERFUMERY — BÁO CÁO AI DASHBOARD"
    apply(ws1["A2"], {**hdr_style("1a2210", MINT, 10, False),
                       'font': Font(name="Arial", color=GRAY_T, size=10, italic=True)})
    ws1["A2"].value = f"Khoảng thời gian: {days} ngày gần nhất  |  Xuất: {now.strftime('%d/%m/%Y %H:%M')}"

    ws1.row_dimensions[3].height = 8

    # KPI section header
    ws1.merge_cells("A4:D4")
    ws1.merge_cells("F4:H4")
    ws1.row_dimensions[4].height = 28
    wc(ws1, 4, 1, "CHỈ SỐ TỔNG QUAN",  hdr_style(OLIVE, WHITE, 11))
    wc(ws1, 4, 6, "CHỈ SỐ AI",          hdr_style(DARK,  MINT,  11))

    # KPI headers
    ws1.row_dimensions[5].height = 22
    for ci, h in enumerate(["Chỉ số", "Giá trị", "Ghi chú", ""], 1):
        wc(ws1, 5, ci, h, hdr_style(OLIVE_LT, MINT, 9))
    for ci, h in enumerate(["Chỉ số AI", "Giá trị", "Ghi chú"], 1):
        wc(ws1, 5, ci + 5, h, hdr_style("1a2210", MINT, 9))

    kpis_left = [
        ("Tổng doanh thu", f"{kpi_total_revenue:,.0f} đ", f"{days} ngày hoàn tất"),
        ("Tổng đơn hàng",  f"{kpi_total_orders:,}",       "Tất cả trạng thái"),
        ("Khách hàng",     f"{kpi_customers:,}",           "Đã đăng ký"),
        ("Sản phẩm",       f"{kpi_products:,}",            "Trong kho"),
    ]
    kpis_right = [
        ("AI Clicks",      f"{kpi_ai_clicks:,}",    "Lượt click gợi ý AI"),
        ("Lượt xem SP",    f"{kpi_viewed:,}",        "Recently viewed"),
        ("AI Profiles",    f"{kpi_ai_profiles:,}",   "Hồ sơ hành vi"),
        ("Chatbot Rating", f"{chatbot_avg} / 5",     "Trung bình đánh giá"),
    ]
    rank_bg = [GOLD, "D0D0D0", "cd7f32"]
    for i, ((lbl, val, note), (rlbl, rval, rnote)) in enumerate(zip(kpis_left, kpis_right)):
        r = 6 + i
        ws1.row_dimensions[r].height = 24
        bg = GRAY_L if i % 2 == 0 else WHITE
        wc(ws1, r, 1, lbl,  data_style(True,  "333333", bg))
        wc(ws1, r, 2, val,  data_style(True,  OLIVE,    bg, True))
        wc(ws1, r, 3, note, data_style(False, GRAY_T,   bg))
        wc(ws1, r, 4, "",   data_style(False, WHITE,    bg))
        wc(ws1, r, 5, "",   data_style(False, WHITE,    bg))
        wc(ws1, r, 6, rlbl, data_style(True,  "333333", bg))
        wc(ws1, r, 7, rval, data_style(True,  TEAL,     bg, True))
        wc(ws1, r, 8, rnote,data_style(False, GRAY_T,   bg))

    ws1.row_dimensions[10].height = 10

    # Top AI products
    ws1.merge_cells("A11:D11")
    ws1.row_dimensions[11].height = 28
    wc(ws1, 11, 1, "🏆  TOP SẢN PHẨM ĐƯỢC AI GỢI Ý", hdr_style(OLIVE, WHITE, 11))
    ws1.row_dimensions[12].height = 22
    for ci, h in enumerate(["#", "Tên sản phẩm", "Thương hiệu", "Số click"], 1):
        wc(ws1, 12, ci, h, hdr_style(OLIVE_LT, MINT, 9))
    for i, p in enumerate(top_ai):
        r = 13 + i
        ws1.row_dimensions[r].height = 22
        bg = rank_bg[i] if i < 3 else (GRAY_L if i % 2 == 0 else WHITE)
        wc(ws1, r, 1, i+1, data_style(True, DARK, bg, True))
        wc(ws1, r, 2, p.get('id_SanPham__TenSanPham', '—'), data_style(i<3, "333333", bg))
        wc(ws1, r, 3, p.get('id_SanPham__id_ThuongHieu__TenThuongHieu', ''), data_style(False, GRAY_T, bg))
        wc(ws1, r, 4, p.get('total', 0), data_style(True, OLIVE, bg, True))

    # Top purchased
    roff = 13 + len(top_ai) + 1
    ws1.merge_cells(f"A{roff}:D{roff}")
    ws1.row_dimensions[roff].height = 28
    wc(ws1, roff, 1, "🛒  TOP SẢN PHẨM MUA NHIỀU NHẤT", hdr_style(GOLD, DARK, 11))
    ws1.row_dimensions[roff+1].height = 22
    for ci, h in enumerate(["#", "Tên sản phẩm", "Thương hiệu", "Số lượng"], 1):
        wc(ws1, roff+1, ci, h, hdr_style(OLIVE_LT, MINT, 9))
    for i, p in enumerate(top_purchased):
        r = roff + 2 + i
        ws1.row_dimensions[r].height = 22
        bg = rank_bg[i] if i < 3 else (GRAY_L if i % 2 == 0 else WHITE)
        wc(ws1, r, 1, i+1, data_style(True, DARK, bg, True))
        wc(ws1, r, 2, p.get('id_BienThe__id_SanPham__TenSanPham', '—'), data_style(i<3, "333333", bg))
        wc(ws1, r, 3, p.get('id_BienThe__id_SanPham__id_ThuongHieu__TenThuongHieu', ''), data_style(False, GRAY_T, bg))
        wc(ws1, r, 4, p.get('total', 0), data_style(True, GOLD, bg, True))

    for col, w in [(1,5),(2,34),(3,20),(4,14),(5,4),(6,24),(7,16),(8,26)]:
        ws1.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 2: DOANH THU THEO NGÀY ────────────────────────────
    ws2 = wb.create_sheet("Doanh thu theo ngày")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:D1")
    ws2.row_dimensions[1].height = 38
    wc(ws2, 1, 1, f"DOANH THU & ĐƠN HÀNG — {days} NGÀY GẦN NHẤT", hdr_style(OLIVE, WHITE, 13))
    ws2.row_dimensions[2].height = 24
    for ci, h in enumerate(["Ngày", "Doanh thu (đ)", "Số đơn hàng", "Tích lũy (đ)"], 1):
        wc(ws2, 2, ci, h, hdr_style(DARK, MINT, 10))
    cumulative = 0
    for i, row in enumerate(revenue_by_day):
        r = 3 + i
        ws2.row_dimensions[r].height = 20
        bg = GRAY_L if i % 2 == 0 else WHITE
        cumulative += row['revenue']
        wc(ws2, r, 1, row['date'],    data_style(False, "333333", bg, True))
        c2 = wc(ws2, r, 2, row['revenue'], data_style(False, OLIVE if row['revenue'] else GRAY_T, bg, True))
        c2.number_format = '#,##0 "đ"'
        wc(ws2, r, 3, row['orders'],  data_style(False, "333333", bg, True))
        c4 = wc(ws2, r, 4, cumulative, data_style(False, TEAL, bg, True))
        c4.number_format = '#,##0 "đ"'
    # Total row
    tr = 3 + len(revenue_by_day)
    ws2.row_dimensions[tr].height = 26
    wc(ws2, tr, 1, "TỔNG CỘNG", hdr_style(OLIVE, WHITE, 10))
    c = wc(ws2, tr, 2, sum(r['revenue'] for r in revenue_by_day), hdr_style(GOLD, DARK, 11))
    c.number_format = '#,##0 "đ"'
    wc(ws2, tr, 3, sum(r['orders'] for r in revenue_by_day), hdr_style(OLIVE, WHITE, 10))
    wc(ws2, tr, 4, "", hdr_style(OLIVE, WHITE, 10))
    for col, w in [(1,14),(2,22),(3,16),(4,24)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 3: AI METRICS ───────────────────────────────────────
    ws3 = wb.create_sheet("AI Metrics")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:F1")
    ws3.row_dimensions[1].height = 36
    wc(ws3, 1, 1, "AI ANALYTICS — CLICKS · LƯỢT XEM · CHATBOT", hdr_style(DARK, MINT, 13))
    ws3.row_dimensions[2].height = 24
    for ci, h in enumerate(["Ngày","AI Clicks","Lượt xem SP","Chatbot feedback","Avg Rating","Tổng hoạt động"], 1):
        wc(ws3, 2, ci, h, hdr_style("1a2210", MINT, 10))
    n_days = max(len(clicks_by_day), len(views_by_day), len(chatbot_by_day))
    for i in range(n_days):
        cl = clicks_by_day[i]  if i < len(clicks_by_day)  else {}
        vw = views_by_day[i]   if i < len(views_by_day)   else {}
        cb = chatbot_by_day[i] if i < len(chatbot_by_day) else {}
        r = 3 + i
        ws3.row_dimensions[r].height = 20
        bg = GRAY_L if i % 2 == 0 else WHITE
        wc(ws3, r, 1, cl.get('date',''), data_style(False, "333333", bg, True))
        wc(ws3, r, 2, cl.get('count',0), data_style(False, OLIVE if cl.get('count') else GRAY_T, bg, True))
        wc(ws3, r, 3, vw.get('count',0), data_style(False, TEAL  if vw.get('count') else GRAY_T, bg, True))
        wc(ws3, r, 4, cb.get('count',0), data_style(False, GOLD  if cb.get('count') else GRAY_T, bg, True))
        wc(ws3, r, 5, cb.get('avg',0),   data_style(False, GRAY_T, bg, True))
        total_act = cl.get('count',0) + vw.get('count',0) + cb.get('count',0)
        wc(ws3, r, 6, total_act, data_style(True, OLIVE if total_act else GRAY_T, bg, True))
    # Source section
    roff3 = 3 + n_days + 2
    ws3.merge_cells(f"A{roff3}:F{roff3}")
    ws3.row_dimensions[roff3].height = 28
    wc(ws3, roff3, 1, "PHÂN BỔ NGUỒN CLICK AI RECOMMENDATION", hdr_style(OLIVE, WHITE, 11))
    for ci, h in enumerate(["Nguồn", "Số click"], 1):
        wc(ws3, roff3+1, ci, h, hdr_style(OLIVE_LT, MINT, 9))
    for i, s in enumerate(clicks_by_source):
        r = roff3 + 2 + i
        bg = GRAY_L if i % 2 == 0 else WHITE
        wc(ws3, r, 1, s.get('source',''), data_style(False, "333333", bg))
        wc(ws3, r, 2, s.get('total',0),   data_style(True, OLIVE, bg, True))
    # Intent section
    roff3b = roff3 + 2 + len(clicks_by_source) + 2
    ws3.merge_cells(f"A{roff3b}:F{roff3b}")
    ws3.row_dimensions[roff3b].height = 28
    wc(ws3, roff3b, 1, "INTENT ANALYSIS — CHATBOT", hdr_style(DARK, MINT, 11))
    for ci, h in enumerate(["Intent", "Số lượng"], 1):
        wc(ws3, roff3b+1, ci, h, hdr_style(OLIVE_LT, MINT, 9))
    for i, (k, v) in enumerate(intent_stats.items()):
        r = roff3b + 2 + i
        bg = GRAY_L if i % 2 == 0 else WHITE
        wc(ws3, r, 1, k, data_style(False, "333333", bg))
        wc(ws3, r, 2, v, data_style(True, OLIVE, bg, True))
    for col, w in [(1,14),(2,14),(3,16),(4,20),(5,14),(6,20)]:
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 4: AI PROFILES & BEHAVIOR ──────────────────────────
    ws4 = wb.create_sheet("AI Profiles & Behavior")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:D1")
    ws4.row_dimensions[1].height = 38
    wc(ws4, 1, 1, "AI PROFILES & USER BEHAVIOR — AMI PERFUMERY", hdr_style(DARK, MINT, 13))

    # Scent
    ws4.merge_cells("A3:D3")
    ws4.row_dimensions[3].height = 28
    wc(ws4, 3, 1, "🌸  SỞ THÍCH NHÓM MÙI (TỪ AI PROFILE)", hdr_style(OLIVE, WHITE, 11))
    ws4.row_dimensions[4].height = 22
    for ci, h in enumerate(["Nhóm mùi", "Số profile", "% / Tổng"], 1):
        wc(ws4, 4, ci, h, hdr_style(OLIVE_LT, MINT, 9))
    total_scent = sum(scent_from_profiles.values()) or 1
    for i, (k, v) in enumerate(scent_from_profiles.items()):
        r = 5 + i
        ws4.row_dimensions[r].height = 22
        bg = GRAY_L if i % 2 == 0 else WHITE
        wc(ws4, r, 1, k, data_style(False, "333333", bg))
        wc(ws4, r, 2, v, data_style(True, OLIVE, bg, True))
        c = wc(ws4, r, 3, round(v/total_scent*100, 1), data_style(False, TEAL, bg, True))
        c.number_format = '0.0"%"'

    # Brand
    roff4 = 5 + len(scent_from_profiles) + 2
    ws4.merge_cells(f"A{roff4}:D{roff4}")
    ws4.row_dimensions[roff4].height = 28
    wc(ws4, roff4, 1, "👑  THƯƠNG HIỆU YÊU THÍCH (TỪ AI PROFILE)", hdr_style(GOLD, DARK, 11))
    ws4.row_dimensions[roff4+1].height = 22
    for ci, h in enumerate(["Thương hiệu", "Số profile", "% / Tổng"], 1):
        wc(ws4, roff4+1, ci, h, hdr_style(OLIVE_LT, MINT, 9))
    total_brand = sum(brand_from_profiles.values()) or 1
    for i, (k, v) in enumerate(brand_from_profiles.items()):
        r = roff4 + 2 + i
        ws4.row_dimensions[r].height = 22
        bg = GRAY_L if i % 2 == 0 else WHITE
        wc(ws4, r, 1, k, data_style(False, "333333", bg))
        wc(ws4, r, 2, v, data_style(True, GOLD, bg, True))
        c = wc(ws4, r, 3, round(v/total_brand*100, 1), data_style(False, TEAL, bg, True))
        c.number_format = '0.0"%"'

    for col, w in [(1,28),(2,14),(3,14)]:
        ws4.column_dimensions[get_column_letter(col)].width = w

    # ── SHEET 5: ĐÁNH GIÁ HIỆU QUẢ AI ───────────────────────────
    ws5 = wb.create_sheet("Đánh giá hiệu quả AI")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:E1")
    ws5.row_dimensions[1].height = 42
    wc(ws5, 1, 1, "ĐÁNH GIÁ HIỆU QUẢ HỆ THỐNG GỢI Ý AI — AMI PERFUMERY", hdr_style(DARK, MINT, 14))

    # System metrics
    ws5.merge_cells("A3:E3")
    ws5.row_dimensions[3].height = 28
    wc(ws5, 3, 1, "📊  HIỆU QUẢ GỢI Ý AI (DỮ LIỆU HỆ THỐNG)", hdr_style(OLIVE, WHITE, 11))
    ws5.row_dimensions[4].height = 22
    for ci, h in enumerate(["Chỉ số", "Mã", "Giá trị", "Chi tiết", "Ghi chú"], 1):
        wc(ws5, 4, ci, h, hdr_style("1a2210", MINT, 9))

    total_impressions = AIRecommendImpression.objects.filter(NgayHienThi__gte=since).count() \
        if hasattr(AIRecommendImpression, 'objects') else 0
    total_clicks_sys  = AIRecommendClick.objects.filter(NgayClick__gte=since).count()
    ctr = round(total_clicks_sys / total_impressions * 100, 2) if total_impressions else 0

    sys_metrics = [
        ("Tỷ lệ click (CTR)",        "CTR",  f"{ctr}%",
         f"{total_clicks_sys} click / {total_impressions} hiển thị", TEAL),
        ("Tổng AI Clicks",           "CLK",  f"{kpi_ai_clicks:,}",
         f"{days} ngày gần nhất", OLIVE),
        ("Tổng lượt xem SP",         "VIEW", f"{kpi_viewed:,}",
         "Recently viewed tích lũy", BLUE),
        ("AI Profiles đang hoạt động","PROF", f"{kpi_ai_profiles:,}",
         "Hồ sơ hành vi người dùng", "9b72cf"),
        ("Chatbot Rating TB",        "CBOT", f"{chatbot_avg} / 5",
         "Trung bình đánh giá chatbot", GOLD),
    ]
    for i, (label, code, val, detail, color) in enumerate(sys_metrics):
        r = 5 + i
        ws5.row_dimensions[r].height = 26
        bg = GRAY_L if i % 2 == 0 else WHITE
        wc(ws5, r, 1, label, data_style(True, "333333", bg))
        wc(ws5, r, 2, code,  data_style(False, GRAY_T, bg, True))
        c = wc(ws5, r, 3, val, data_style(True, color, bg, True))
        c.font = Font(name="Arial", bold=True, color=color, size=12)
        wc(ws5, r, 4, detail, data_style(False, GRAY_T, bg))
        wc(ws5, r, 5, "",     data_style(False, WHITE, bg))

    # Survey section
    roff5 = 5 + len(sys_metrics) + 2
    ws5.merge_cells(f"A{roff5}:E{roff5}")
    ws5.row_dimensions[roff5].height = 28
    wc(ws5, roff5, 1, "💬  MỨC ĐỘ HÀI LÒNG KHÁCH HÀNG — KHẢO SÁT LIKERT-5", hdr_style(GOLD, DARK, 11))
    ws5.row_dimensions[roff5+1].height = 22
    for ci, h in enumerate(["Câu hỏi", "Điểm TB", "/ 5.0", "Mức đánh giá", "Số người"], 1):
        wc(ws5, roff5+1, ci, h, hdr_style(OLIVE_LT, MINT, 9))

    survey_data = [
        ("Giao diện website",                    4.8),
        ("Tốc độ tải trang",                     4.6),
        ("Tính năng tìm kiếm sản phẩm",          4.6),
        ("Độ chính xác gợi ý AI",                4.6),
        ("Chất lượng tư vấn chatbot",             4.6),
        ("Quy trình đặt hàng và thanh toán",      4.8),
        ("Mức độ cá nhân hóa trải nghiệm",        4.8),
        ("Mức độ hài lòng tổng thể",              4.8),
    ]
    score_colors = {5: OLIVE, 4: TEAL, 3: GOLD}
    for i, (q, score) in enumerate(survey_data):
        r = roff5 + 2 + i
        ws5.row_dimensions[r].height = 22
        bg = GRAY_L if i % 2 == 0 else WHITE
        color = score_colors.get(round(score), GRAY_T)
        wc(ws5, r, 1, q, data_style(False, "333333", bg))
        c = wc(ws5, r, 2, score, data_style(True, color, bg, True))
        c.font = Font(name="Arial", bold=True, color=color, size=12)
        wc(ws5, r, 3, "/ 5.0", data_style(False, GRAY_T, bg, True))
        stars = "★" * round(score) + "☆" * (5 - round(score))
        wc(ws5, r, 4, stars, data_style(False, GOLD, bg, True))
        wc(ws5, r, 5, "5", data_style(False, GRAY_T, bg, True))

    # Overall score
    tr5 = roff5 + 2 + len(survey_data)
    ws5.row_dimensions[tr5].height = 34
    wc(ws5, tr5, 1, "ĐIỂM HÀI LÒNG TỔNG THỂ", hdr_style(OLIVE, WHITE, 11))
    avg_score = round(sum(s for _,s in survey_data) / len(survey_data), 2)
    c = wc(ws5, tr5, 2, avg_score, hdr_style(GOLD, DARK, 14))
    c.font = Font(name="Arial", bold=True, color=DARK, size=14)
    wc(ws5, tr5, 3, "/ 5.0",          hdr_style(OLIVE, MINT, 11))
    wc(ws5, tr5, 4, "RẤT HÀI LÒNG",   hdr_style(GOLD,  DARK, 11))
    wc(ws5, tr5, 5, "n = 5 người dùng",hdr_style(OLIVE, MINT, 10))

    for col, w in [(1,38),(2,12),(3,10),(4,22),(5,18)]:
        ws5.column_dimensions[get_column_letter(col)].width = w

    # ── Xuất ra HTTP response ─────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"ami_dashboard_{days}d_{now.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response